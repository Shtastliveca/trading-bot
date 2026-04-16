import os
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import time

# Set up logging
logger = logging.getLogger("excel_logger")

# Excel Logger Class
# Excel Logger Class
class ExcelTradeLogger:
    def __init__(self, excel_path):
        """
        Initialize the Excel Trade Logger
        
        Args:
            excel_path (str): Path to the Excel file
        """
        self.excel_path = excel_path
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(os.path.abspath(excel_path)), exist_ok=True)
        
        # Check if file exists, if not create it with template
        if not os.path.exists(excel_path):
            self._create_template()
        
        logger.info(f"Excel Trade Logger initialized with file: {excel_path}")

    def _create_template(self):
        """Create a new Excel file with the template structure"""
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trade Log"
            
            # Add headers with enhanced columns for fee structure
            headers = [
                "Trade #", "Symbol", "Direction", "Entry Time", "Exit Time", "Entry Price", 
                "Stop Loss", "Exit Price", "Position Size", "Risk $", "Gross P&L", "Fees", 
                "Net P&L", "R Multiple", "Account Balance", "Post-Trade Balance", "Order Type",
                "Exit Order Type", "Entry Fee %", "Exit Fee %"
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col).value = header
            
            # Format the headers
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).fill = self.yellow_fill
            
            # Save the workbook
            wb.save(self.excel_path)
            logger.info(f"Created new Excel template at {self.excel_path}")
        except Exception as e:
            logger.error(f"Error creating Excel template: {str(e)}")
            raise

    def log_trade(self, trade_data):
        """
        Log a trade to the Excel file
        
        Args:
            trade_data (dict): Dictionary containing trade data
        """
        try:
            # Implement retry mechanism for file access
            max_retries = 5
            retry_delay = 1  # seconds
            
            for attempt in range(max_retries):
                try:
                    # Open the workbook
                    wb = openpyxl.load_workbook(self.excel_path)
                    ws = wb.active
                    
                    # Get current headers
                    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
                    
                    # Check if we need to add Order Type column
                    if ws.max_column < 17 or "Order Type" not in headers:  
                        # If we're missing Order Type column, add it
                        if "Order Type" not in headers:
                            new_col = ws.max_column + 1
                            ws.cell(row=1, column=new_col).value = "Order Type"
                            ws.cell(row=1, column=new_col).fill = self.yellow_fill
                    
                    # Check if we need to add Exit Order Type column
                    if "Exit Order Type" not in headers:
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Exit Order Type"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                    
                    # Check if we need to add fee percentage columns
                    if "Entry Fee %" not in headers:
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Entry Fee %"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                    
                    if "Exit Fee %" not in headers:
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Exit Fee %"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                    
                    # Check if we need to add fee columns (for existing files)
                    if "Fees" not in headers:
                        # Determine where to insert fee columns
                        if "Return $" in headers:
                            return_col = headers.index("Return $") + 1
                            r_multiple_col = headers.index("R Multiple") + 1
                            
                            # Shift columns right to make space for new fee columns
                            for row in range(1, ws.max_row + 1):
                                # Move r multiple and all columns after it
                                for col in range(ws.max_column, r_multiple_col - 1, -1):
                                    ws.cell(row=row, column=col+2).value = ws.cell(row=row, column=col).value
                            
                            # Add new fee column headers
                            ws.cell(row=1, column=return_col+1).value = "Fees"
                            ws.cell(row=1, column=return_col+2).value = "Net P&L"
                            
                            # Format new headers
                            ws.cell(row=1, column=return_col+1).fill = self.yellow_fill
                            ws.cell(row=1, column=return_col+2).fill = self.yellow_fill
                            
                            # Update column names
                            ws.cell(row=1, column=return_col).value = "Gross P&L" # Rename "Return $"
                    
                    # Find the next available row
                    next_row = 2  # Start after header
                    while ws.cell(row=next_row, column=1).value is not None:
                        next_row += 1
                    
                    # Calculate trade number
                    trade_number = next_row - 1
                    
                    # Extract data
                    entry_price = float(trade_data.get('entry_price', 0))
                    stop_loss = float(trade_data.get('stop_loss', 0))
                    exit_price = float(trade_data.get('exit_price', 0))
                    position_size = float(trade_data.get('position_size', 0))
                    
                    # Get fee data directly from trade_data
                    total_fees = float(trade_data.get('total_fees', 0))
                    entry_fee = float(trade_data.get('entry_fee', 0))
                    exit_fee = float(trade_data.get('exit_fee', 0))
                    
                    # Get order types and fee percentages
                    order_type = trade_data.get('order_type', 'Market')
                    exit_order_type = trade_data.get('exit_order_type', 'Market')
                    entry_fee_percent = trade_data.get('entry_fee_percent')
                    exit_fee_percent = trade_data.get('exit_fee_percent')
                    
                    # If fee percentages not provided, calculate them
                    if entry_fee_percent is None:
                        if entry_price > 0 and position_size > 0:
                            entry_fee_percent = (entry_fee / (entry_price * position_size)) * 100
                        else:
                            # Get from config
                            if order_type == 'Limit':
                                entry_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02))
                            else:
                                entry_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075))
                    
                    if exit_fee_percent is None:
                        if exit_price > 0 and position_size > 0:
                            exit_fee_percent = (exit_fee / (exit_price * position_size)) * 100
                        else:
                            # Get from config
                            if exit_order_type == 'Limit':
                                exit_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02))
                            else:
                                exit_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075))
                    
                    # Get the direction
                    direction = trade_data.get('direction', '')
                    
                    # Calculate gross PnL
                    if direction == 'long':
                        gross_pnl = (exit_price - entry_price) * position_size
                    else:  # short
                        gross_pnl = (entry_price - exit_price) * position_size
                    
                    # Net PnL after fees
                    net_pnl = gross_pnl - total_fees
                    
                    # Calculate risk with integrated fee
                    if direction == 'long':
                        # Get fee percentages
                        taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100
                        maker_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02)) / 100
                        
                        # Determine entry fee percentage based on order type
                        entry_fee_pct = maker_fee_percent if order_type == 'Limit' else taker_fee_percent
                        
                        # Price risk component
                        price_risk_per_unit = entry_price - stop_loss if stop_loss > 0 else 0
                        price_risk = price_risk_per_unit * position_size
                        
                        # Fee risk component (estimated based on entry and stop loss)
                        if stop_loss > 0:
                            fee_cost_per_unit = entry_fee_pct * entry_price + taker_fee_percent * stop_loss
                            fee_risk = fee_cost_per_unit * position_size
                        else:
                            fee_risk = 0
                        
                        # Total risk amount
                        risk_amount = price_risk + fee_risk
                    else:  # short
                        # Get fee percentages
                        taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100
                        maker_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02)) / 100
                        
                        # Determine entry fee percentage based on order type
                        entry_fee_pct = maker_fee_percent if order_type == 'Limit' else taker_fee_percent
                        
                        # Price risk component
                        price_risk_per_unit = stop_loss - entry_price if stop_loss > 0 else 0
                        price_risk = price_risk_per_unit * position_size
                        
                        # Fee risk component (estimated based on entry and stop loss)
                        if stop_loss > 0:
                            fee_cost_per_unit = entry_fee_pct * entry_price + taker_fee_percent * stop_loss
                            fee_risk = fee_cost_per_unit * position_size
                        else:
                            fee_risk = 0
                        
                        # Total risk amount
                        risk_amount = price_risk + fee_risk
                    
                    # Calculate R multiple based on net PnL
                    r_multiple = net_pnl / risk_amount if risk_amount > 0 else 0
                    
                    # Format dates
                    entry_time = trade_data.get('entry_time', '')
                    exit_time = trade_data.get('exit_time', '')
                    
                    # Try to parse ISO format dates
                    try:
                        if isinstance(entry_time, str):
                            entry_time = datetime.fromisoformat(entry_time.replace('Z', '+00:00'))
                        if isinstance(exit_time, str):
                            exit_time = datetime.fromisoformat(exit_time.replace('Z', '+00:00'))
                    except ValueError:
                        # If parsing fails, keep as is
                        pass
                    
                    # Determine column positions based on headers
                    fee_col = headers.index("Fees") + 1 if "Fees" in headers else None
                    net_pnl_col = headers.index("Net P&L") + 1 if "Net P&L" in headers else None
                    gross_pnl_col = headers.index("Gross P&L") + 1 if "Gross P&L" in headers else headers.index("Return $") + 1 if "Return $" in headers else None
                    
                    # Populate the row with base data
                    ws.cell(row=next_row, column=1).value = trade_number
                    ws.cell(row=next_row, column=2).value = trade_data.get('symbol', '')
                    ws.cell(row=next_row, column=3).value = direction.upper()
                    ws.cell(row=next_row, column=4).value = entry_time
                    ws.cell(row=next_row, column=5).value = exit_time
                    ws.cell(row=next_row, column=6).value = entry_price
                    ws.cell(row=next_row, column=7).value = stop_loss
                    ws.cell(row=next_row, column=8).value = exit_price
                    ws.cell(row=next_row, column=9).value = position_size
                    ws.cell(row=next_row, column=10).value = risk_amount
                    
                    # Add gross PnL, fees, and net PnL
                    if gross_pnl_col:
                        ws.cell(row=next_row, column=gross_pnl_col).value = gross_pnl
                    
                    if fee_col:
                        ws.cell(row=next_row, column=fee_col).value = total_fees
                    
                    if net_pnl_col:
                        ws.cell(row=next_row, column=net_pnl_col).value = net_pnl
                    
                    # R multiple based on net PnL
                    r_multiple_col = headers.index("R Multiple") + 1 if "R Multiple" in headers else None
                    if r_multiple_col:
                        ws.cell(row=next_row, column=r_multiple_col).value = r_multiple
                    
                    # Account balance (pre-trade)
                    balance_col = headers.index("Account Balance") + 1 if "Account Balance" in headers else None
                    if balance_col:
                        ws.cell(row=next_row, column=balance_col).value = trade_data.get('account_balance', 0)
                    
                    # Post-trade balance
                    if "Post-Trade Balance" in headers:
                        post_trade_balance_col = headers.index("Post-Trade Balance") + 1
                        ws.cell(row=next_row, column=post_trade_balance_col).value = trade_data.get('post_trade_balance', 0)
                    else:
                        # If Post-Trade Balance column doesn't exist yet, add it
                        if balance_col:
                            # Add new column header
                            post_trade_col = balance_col + 1
                            # Shift existing data if necessary
                            for row in range(1, next_row):
                                for col in range(ws.max_column, post_trade_col - 1, -1):
                                    ws.cell(row=row, column=col+1).value = ws.cell(row=row, column=col).value
                            
                            # Add the new header
                            ws.cell(row=1, column=post_trade_col).value = "Post-Trade Balance"
                            ws.cell(row=1, column=post_trade_col).fill = self.yellow_fill
                            
                            # Add the value
                            ws.cell(row=next_row, column=post_trade_col).value = trade_data.get('post_trade_balance', 0)
                    
                    # Add order type
                    order_type_col = headers.index("Order Type") + 1 if "Order Type" in headers else None
                    if order_type_col:
                        ws.cell(row=next_row, column=order_type_col).value = order_type
                    else:
                        # If Order Type column doesn't exist yet, add it
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Order Type"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                        ws.cell(row=next_row, column=new_col).value = order_type
                    
                    # Add exit order type
                    exit_order_type_col = headers.index("Exit Order Type") + 1 if "Exit Order Type" in headers else None
                    if exit_order_type_col:
                        ws.cell(row=next_row, column=exit_order_type_col).value = exit_order_type
                    else:
                        # If Exit Order Type column doesn't exist yet, add it
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Exit Order Type"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                        ws.cell(row=next_row, column=new_col).value = exit_order_type
                    
                    # Add fee percentages
                    entry_fee_pct_col = headers.index("Entry Fee %") + 1 if "Entry Fee %" in headers else None
                    if entry_fee_pct_col:
                        ws.cell(row=next_row, column=entry_fee_pct_col).value = entry_fee_percent
                    else:
                        # If Entry Fee % column doesn't exist yet, add it
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Entry Fee %"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                        ws.cell(row=next_row, column=new_col).value = entry_fee_percent
                    
                    exit_fee_pct_col = headers.index("Exit Fee %") + 1 if "Exit Fee %" in headers else None
                    if exit_fee_pct_col:
                        ws.cell(row=next_row, column=exit_fee_pct_col).value = exit_fee_percent
                    else:
                        # If Exit Fee % column doesn't exist yet, add it
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Exit Fee %"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                        ws.cell(row=next_row, column=new_col).value = exit_fee_percent

                    # Log detailed risk breakdown for verification
                    logger.info(f"Excel trade log: {trade_data.get('symbol')} {direction} - " + 
                              f"Price risk: ${price_risk:.2f}, Fee risk: ${fee_risk:.2f}, " + 
                              f"Total risk: ${risk_amount:.2f}, " + 
                              f"Fees: Entry {entry_fee_percent:.3f}% ({order_type}), Exit {exit_fee_percent:.3f}% ({exit_order_type})")

                    # Save the workbook
                    wb.save(self.excel_path)
                    logger.info(f"Logged trade #{trade_number} to Excel: {trade_data.get('symbol')} {direction}")
                    break
                    
                except PermissionError:
                    if attempt < max_retries - 1:
                        logger.warning(f"Excel file is open or locked. Retrying in {retry_delay} seconds... (Attempt {attempt+1}/{max_retries})")
                        time.sleep(retry_delay)
                    else:
                        logger.error("Failed to access Excel file after multiple attempts. Make sure the file is not open in another program.")
                        raise
                
        except Exception as e:
            logger.error(f"Error logging trade to Excel: {str(e)}")
            raise

    def export_from_database(self, db_path, limit=None):
        """
        Export trades from SQLite database to Excel
        
        Args:
            db_path (str): Path to the SQLite database
            limit (int, optional): Maximum number of trades to export. Defaults to None (all trades).
        """
        try:
            import sqlite3
            
            # Connect to database
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # Check if fee percentage and order type columns exist
            cursor.execute("PRAGMA table_info(trade_history)")
            columns = [col[1] for col in cursor.fetchall()]
            
            has_entry_fee_percent = 'entry_fee_percent' in columns
            has_exit_fee_percent = 'exit_fee_percent' in columns
            has_exit_order_type = 'exit_order_type' in columns
            
            # Query trades with additional columns if they exist
            select_columns = [
                'symbol', 'direction', 'entry_price', 'exit_price', 'position_size',
                'pnl', 'pnl_percent', 'entry_time', 'exit_time', 'duration_minutes',
                'stop_loss', 'stopped_out', 'reason', 'total_fees', 'entry_fee', 'exit_fee',
                'gross_pnl', 'order_type'
            ]
            
            # Add new columns if they exist
            if has_exit_order_type:
                select_columns.append('exit_order_type')
            
            if has_entry_fee_percent:
                select_columns.append('entry_fee_percent')
                
            if has_exit_fee_percent:
                select_columns.append('exit_fee_percent')
            
            # Construct query
            query = f'SELECT {", ".join(select_columns)} FROM trade_history ORDER BY exit_time DESC'
            if limit:
                query += f' LIMIT {limit}'
                
            cursor.execute(query)
            trades = cursor.fetchall()
            
            # Get current account balance
            account_balance = 0
            try:
                cursor.execute('SELECT MAX(capital) FROM capital_history')
                result = cursor.fetchone()
                if result and result[0]:
                    account_balance = result[0]
            except:
                # If table doesn't exist or query fails
                pass
            
            # Convert to list of dictionaries
            for trade in trades:
                trade_dict = dict(zip(select_columns, trade))
                
                # Add post_trade_balance (original balance + pnl)
                trade_dict['account_balance'] = account_balance - trade_dict.get('pnl', 0)
                trade_dict['post_trade_balance'] = account_balance
                
                # For older records without exit_order_type
                if not has_exit_order_type:
                    trade_dict['exit_order_type'] = 'Market'
                
                # For older records without fee percentages
                if not has_entry_fee_percent:
                    # Calculate based on order type
                    if trade_dict.get('order_type') == 'Limit':
                        trade_dict['entry_fee_percent'] = float(config['TRADING'].get('maker_fee_percent', 0.02))
                    else:
                        trade_dict['entry_fee_percent'] = float(config['TRADING'].get('trading_fee_percent', 0.075))
                        
                if not has_exit_fee_percent:
                    # Calculate based on exit order type
                    if trade_dict.get('exit_order_type') == 'Limit':
                        trade_dict['exit_fee_percent'] = float(config['TRADING'].get('maker_fee_percent', 0.02))
                    else:
                        trade_dict['exit_fee_percent'] = float(config['TRADING'].get('trading_fee_percent', 0.075))
                
                # Log the trade to Excel
                self.log_trade(trade_dict)
                
                # Update account balance for next trade
                account_balance -= trade_dict.get('pnl', 0)
            
            conn.close()
            logger.info(f"Exported {len(trades)} trades from database to Excel")
            
        except Exception as e:
            logger.error(f"Error exporting trades from database: {str(e)}")
            raise

    def backup_excel(self, backup_dir=None):
        """
        Create a backup of the Excel file
        
        Args:
            backup_dir (str, optional): Directory to store backup. Defaults to None (same directory).
        """
        try:
            if not os.path.exists(self.excel_path):
                logger.warning("Cannot backup - Excel file does not exist")
                return
                
            # Determine backup path
            if backup_dir:
                os.makedirs(backup_dir, exist_ok=True)
                backup_path = os.path.join(backup_dir, f"backup_{os.path.basename(self.excel_path)}")
            else:
                file_name, ext = os.path.splitext(self.excel_path)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = f"{file_name}_backup_{timestamp}{ext}"
            
            # Copy file
            import shutil
            shutil.copy2(self.excel_path, backup_path)
            logger.info(f"Created backup of Excel file at {backup_path}")
            
        except Exception as e:
            logger.error(f"Error creating Excel backup: {str(e)}")