import json
import logging
import os
import time
import sys
import psutil
import sqlite3
import requests
import threading
import configparser
import pybit.unified_trading
try:
    from hyperliquid_bridge import HyperliquidBridge, HyperliquidClient, get_hyperliquid_client
    HYPERLIQUID_AVAILABLE = True
except ImportError:
    HYPERLIQUID_AVAILABLE = False
import math
# For playing sound alarms on Windows (skip on Linux/Railway)
try:
    import winsound
    WINSOUND_AVAILABLE = True
except ImportError:
    WINSOUND_AVAILABLE = False
from datetime import datetime, timedelta
from flask import Flask, request, jsonify
from dotenv import load_dotenv
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
# Add additional imports for manual entry functionality
from flask import render_template, send_from_directory

# Import the database queue system
from db_queue import (
    queue_operation, 
    save_position as db_save_position,
    save_pending_order as db_save_pending_order,
    remove_pending_order as db_remove_pending_order,
    remove_position as db_remove_position,
    log_reconciliation_event as db_log_reconciliation_event,
    update_position_check_time as db_update_position_check_time,
    start_worker,
    stop_worker
)


# Excel logging module imports
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Load environment variables from .env file
load_dotenv()

# Set up logging with proper encoding
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("trading_bot.log", encoding='utf-8'),
        logging.StreamHandler(sys.stdout)  # This will help with console encoding
    ]
)
logger = logging.getLogger("bybit_trading_bot")

# Flask app for webhook endpoint
webhook_app = Flask("webhook_app")  # For the webhook only
app = Flask("main_app")             # For manual entry and other functionality

# Configuration
config = configparser.ConfigParser()
config_path = 'config.ini'

# Track bot start time for uptime reporting
start_time = datetime.now()

# Log config loading
logger.info(f"Loading configuration from {config_path}")

# Default configuration
if not os.path.exists(config_path):
    logger.warning(f"Config file {config_path} not found, creating with defaults")
    config['API'] = {
        'api_key': '',
        'api_secret': '',
        'testnet': 'False'  # Default to live trading
    }
    config['TRADING'] = {
        'default_symbol': 'BTCUSDT',
        'risk_percentage': '5',
        'initial_capital': '10000',
        'max_leverage': '5',
        'use_risk_only': 'False', 
	    'trading_fee_percent': '0.075',  # Default to 0.075% (typical Bybit fee)
	    'balance_buffer_percent': '5',    # Default to 5% buffer
        'use_limit_orders_entry': 'False',  # Default to market orders for entry
        'use_limit_orders_exit': 'False',   # Default to market orders for exit
        'limit_order_timeout_minutes': '5'  # Default timeout for limit orders
    }
    config['WEBHOOK'] = {
        'port': '5000',
        'endpoint': '/webhook'
    }
    config['MONITORING'] = {
        'enabled': 'True',
        'interval_seconds': '300',
        'reconciliation_interval_seconds': '60',  # Default to 1 minute
        'pending_orders_check_seconds': '30',      # Default to check pending orders every 30 seconds
        'reconciliation_schedule_mode': 'interval',       # NEW: 'interval' or 'fixed_second'
        'reconciliation_fixed_second': '30'                # NEW: 0-59, which second of the minute to trigger
    }
    config['NOTIFICATIONS'] = {
        'telegram_token': '',
        'telegram_chat_id': '',
        'notify_entries': 'True',
        'notify_exits': 'True',
        'notify_errors': 'True',
        'notify_system': 'False',
        'notify_reconciliation': 'False',
        'notify_order_status': 'True'
    }
    
    with open(config_path, 'w') as f:
        config.write(f)
else:
    logger.info(f"Reading existing config file")
    config.read(config_path)

# Override config with environment variables (for Railway deployment)
# API keys
if os.getenv('BYBIT_API_KEY'):
    config.set('API', 'api_key', os.getenv('BYBIT_API_KEY'))
if os.getenv('BYBIT_API_SECRET'):
    config.set('API', 'api_secret', os.getenv('BYBIT_API_SECRET'))

# Hyperliquid keys
if os.getenv('HYPERLIQUID_PRIVATE_KEY'):
    if 'HYPERLIQUID' not in config:
        config['HYPERLIQUID'] = {}
    config.set('HYPERLIQUID', 'private_key', os.getenv('HYPERLIQUID_PRIVATE_KEY'))
if os.getenv('HYPERLIQUID_WALLET_ADDRESS'):
    if 'HYPERLIQUID' not in config:
        config['HYPERLIQUID'] = {}
    config.set('HYPERLIQUID', 'wallet_address', os.getenv('HYPERLIQUID_WALLET_ADDRESS'))

# Telegram
if os.getenv('TELEGRAM_TOKEN'):
    config.set('NOTIFICATIONS', 'telegram_token', os.getenv('TELEGRAM_TOKEN'))
if os.getenv('TELEGRAM_CHAT_ID'):
    config.set('NOTIFICATIONS', 'telegram_chat_id', os.getenv('TELEGRAM_CHAT_ID'))`

# Ensure EXCHANGE section exists (for switching between Bybit/Hyperliquid)
if 'EXCHANGE' not in config:
    config['EXCHANGE'] = {
        'active_exchange': 'bybit'  # Options: 'bybit' or 'hyperliquid'
    }
    with open(config_path, 'w') as f:
        config.write(f)

# Ensure HYPERLIQUID section exists
if 'HYPERLIQUID' not in config:
    config['HYPERLIQUID'] = {
        'private_key': '',
        'wallet_address': '',
        'testnet': 'False',
        'trading_fee_percent': '0.035',
        'maker_fee_percent': '0.01'
    }
    with open(config_path, 'w') as f:
        config.write(f)

# Ensure the NOTIFICATIONS section exists
if 'NOTIFICATIONS' not in config:
    config['NOTIFICATIONS'] = {
        'telegram_token': '',
        'telegram_chat_id': '',
        'notify_entries': 'True',
        'notify_exits': 'True',
        'notify_errors': 'True',
        'notify_system': 'False',
        'notify_reconciliation': 'False',
        'notify_order_status': 'True'
    }
    with open(config_path, 'w') as f:
        config.write(f)

# Ensure the MONITORING section has reconciliation interval
if 'MONITORING' not in config:
    config['MONITORING'] = {
        'enabled': 'True',
        'interval_seconds': '300',
        'reconciliation_interval_seconds': '60',  # Default to 1 minute
        'pending_orders_check_seconds': '30'      # Default to check pending orders every 30 seconds
    }
    with open(config_path, 'w') as f:
        config.write(f)
elif 'reconciliation_interval_seconds' not in config['MONITORING']:
    config['MONITORING']['reconciliation_interval_seconds'] = '60'  # Default to 1 minute
    with open(config_path, 'w') as f:
        config.write(f)
elif 'pending_orders_check_seconds' not in config['MONITORING']:
    config['MONITORING']['pending_orders_check_seconds'] = '30'  # Default to 30 seconds
    with open(config_path, 'w') as f:
        config.write(f)
if 'reconciliation_schedule_mode' not in config['MONITORING']:
    config['MONITORING']['reconciliation_schedule_mode'] = 'interval'
    with open(config_path, 'w') as f:
        config.write(f)
if 'reconciliation_fixed_second' not in config['MONITORING']:
    config['MONITORING']['reconciliation_fixed_second'] = '30'
    with open(config_path, 'w') as f:
        config.write(f)

# Add sound alarm settings to config if they don't exist
if 'MONITORING' in config:
    if 'sound_alarm_enabled' not in config['MONITORING']:
        config['MONITORING']['sound_alarm_enabled'] = 'True'  # Default to enabled
    if 'sound_alarm_frequency' not in config['MONITORING']:
        config['MONITORING']['sound_alarm_frequency'] = '1000'  # Default: 1000 Hz
    if 'sound_alarm_duration' not in config['MONITORING']:
        config['MONITORING']['sound_alarm_duration'] = '500'  # Default: 500 ms
    with open(config_path, 'w') as f:
        config.write(f)

# Add Excel logging configuration section
if 'EXCEL_LOGGING' not in config:
    config['EXCEL_LOGGING'] = {
        'enabled': 'True',
        'excel_path': 'trading_log.xlsx',
        'auto_backup': 'True',
        'backup_interval_hours': '24'
    }
    with open(config_path, 'w') as f:
        config.write(f)

# Add limit order trading settings if they don't exist
if 'TRADING' in config:
    if 'use_limit_orders_entry' not in config['TRADING']:
        config['TRADING']['use_limit_orders_entry'] = 'False'
    if 'use_limit_orders_exit' not in config['TRADING']:
        config['TRADING']['use_limit_orders_exit'] = 'False'
    if 'limit_order_timeout_minutes' not in config['TRADING']:
        config['TRADING']['limit_order_timeout_minutes'] = '5'
    # Add TP adjustment check seconds if it doesn't exist
    if 'tp_adjustment_check_seconds' not in config['TRADING']:
        config['TRADING']['tp_adjustment_check_seconds'] = '10'
    with open(config_path, 'w') as f:
        config.write(f)

# Log important configuration values (masked for security)
api_key = config['API'].get('api_key') or os.getenv('BYBIT_API_KEY', '')
testnet_setting = config['API'].getboolean('testnet', False)
logger.info(f"API Key configured: {'Yes' if api_key else 'No'}")
logger.info(f"Using Testnet: {testnet_setting}")

# Global variables
active_positions = {}  # Track active positions
pending_orders = {}    # Track pending limit orders
previous_positions = {}  # Store the previous state for reconciliation comparison
api_errors = 0  # Track API errors for health monitoring
api_calls = 0  # Track total API calls
bot_status = "running"  # Track bot status
global_client = None  # Global Bybit client instance for reuse
excel_logger = None  # Global Excel logger instance
connection_loss_time = None  # Track when connection was lost
internet_status = "unknown"  # Track internet connectivity separately
exchange_status = "unknown"  # Track exchange connectivity separately
internet_loss_time = None    # Track when internet was lost
exchange_loss_time = None    # Track when exchange connection was lost
# Global variables for max daily loss notifications
last_max_loss_notification_time = None
last_warning_notification_time = None
max_loss_first_triggered_time = None
skipped_signals_count = 0
# Track recently closed positions to prevent double recording
recently_closed_positions = {}  # Format: {symbol: timestamp}
RECONCILIATION_COOLING_PERIOD = 20  # seconds (cooling period before reconciliation can re-process a position)
# Track orders that failed cancellation for retry
orders_pending_cancellation = {}  # Format: {order_id: {'symbol': symbol, 'retry_count': count, 'last_attempt': timestamp}}


# Excel Logger Class
# Excel Logger Class
class ExcelTradeLogger:
    def __init__(self, excel_path):
        """
        Initialize the Excel Trade Logger
        
        Args:
            excel_path (str): Path to the Excel file
        """
        self.excel_path = excel_path
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(os.path.abspath(excel_path)), exist_ok=True)
        
        # Check if file exists, if not create it with template
        if not os.path.exists(excel_path):
            self._create_template()
        
        logger.info(f"Excel Trade Logger initialized with file: {excel_path}")

    def _create_template(self):
        """Create a new Excel file with the template structure"""
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trade Log"
            
            # Add headers with enhanced columns for fee structure
            headers = [
                "Trade #", "Symbol", "Direction", "Entry Time", "Exit Time", "Entry Price", 
                "Stop Loss", "Exit Price", "Position Size", "Risk $", "Gross P&L", "Fees", 
                "Net P&L", "R Multiple", "Account Balance", "Post-Trade Balance", "Order Type",
                "Exit Order Type", "Entry Fee %", "Exit Fee %"
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col).value = header
            
            # Format the headers
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).fill = self.yellow_fill
            
            # Save the workbook
            wb.save(self.excel_path)
            logger.info(f"Created new Excel template at {self.excel_path}")
        except Exception as e:
            logger.error(f"Error creating Excel template: {str(e)}")
            raise

    def log_trade(self, trade_data):
        """
        Log a trade to the Excel file
        
        Args:
            trade_data (dict): Dictionary containing trade data
        """
        try:
            # Implement retry mechanism for file access
            max_retries = 5
            retry_delay = 1  # seconds
            
            for attempt in range(max_retries):
                try:
                    # Open the workbook
                    wb = openpyxl.load_workbook(self.excel_path)
                    ws = wb.active
                    
                    # Get current headers
                    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
                    
                    # Check if we need to add Order Type column
                    if ws.max_column < 17 or "Order Type" not in headers:  
                        # If we're missing Order Type column, add it
                        if "Order Type" not in headers:
                            new_col = ws.max_column + 1
                            ws.cell(row=1, column=new_col).value = "Order Type"
                            ws.cell(row=1, column=new_col).fill = self.yellow_fill
                    
                    # Check if we need to add Exit Order Type column
                    if "Exit Order Type" not in headers:
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Exit Order Type"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                    
                    # Check if we need to add fee percentage columns
                    if "Entry Fee %" not in headers:
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Entry Fee %"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                    
                    if "Exit Fee %" not in headers:
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Exit Fee %"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                    
                    # Check if we need to add fee columns (for existing files)
                    if "Fees" not in headers:
                        # Determine where to insert fee columns
                        if "Return $" in headers:
                            return_col = headers.index("Return $") + 1
                            r_multiple_col = headers.index("R Multiple") + 1
                            
                            # Shift columns right to make space for new fee columns
                            for row in range(1, ws.max_row + 1):
                                # Move r multiple and all columns after it
                                for col in range(ws.max_column, r_multiple_col - 1, -1):
                                    ws.cell(row=row, column=col+2).value = ws.cell(row=row, column=col).value
                            
                            # Add new fee column headers
                            ws.cell(row=1, column=return_col+1).value = "Fees"
                            ws.cell(row=1, column=return_col+2).value = "Net P&L"
                            
                            # Format new headers
                            ws.cell(row=1, column=return_col+1).fill = self.yellow_fill
                            ws.cell(row=1, column=return_col+2).fill = self.yellow_fill
                            
                            # Update column names
                            ws.cell(row=1, column=return_col).value = "Gross P&L" # Rename "Return $"
                    
                    # Find the next available row
                    next_row = 2  # Start after header
                    while ws.cell(row=next_row, column=1).value is not None:
                        next_row += 1
                    
                    # Calculate trade number
                    trade_number = next_row - 1
                    
                    # Extract data
                    entry_price = float(trade_data.get('entry_price', 0))
                    stop_loss = float(trade_data.get('stop_loss', 0))
                    exit_price = float(trade_data.get('exit_price', 0))
                    position_size = float(trade_data.get('position_size', 0))
                    
                    # Get fee data directly from trade_data
                    total_fees = float(trade_data.get('total_fees', 0))
                    entry_fee = float(trade_data.get('entry_fee', 0))
                    exit_fee = float(trade_data.get('exit_fee', 0))
                    
                    # Get order types and fee percentages
                    order_type = trade_data.get('order_type', 'Market')
                    exit_order_type = trade_data.get('exit_order_type', 'Market')
                    entry_fee_percent = trade_data.get('entry_fee_percent')
                    exit_fee_percent = trade_data.get('exit_fee_percent')
                    
                    # If fee percentages not provided, calculate them
                    if entry_fee_percent is None:
                        if entry_price > 0 and position_size > 0:
                            entry_fee_percent = (entry_fee / (entry_price * position_size)) * 100
                        else:
                            # Get from config
                            if order_type == 'Limit':
                                entry_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02))
                            else:
                                entry_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075))
                    
                    if exit_fee_percent is None:
                        if exit_price > 0 and position_size > 0:
                            exit_fee_percent = (exit_fee / (exit_price * position_size)) * 100
                        else:
                            # Get from config
                            if exit_order_type == 'Limit':
                                exit_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02))
                            else:
                                exit_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075))
                    
                    # Get the direction
                    direction = trade_data.get('direction', '')
                    
                    # Calculate gross PnL
                    if direction == 'long':
                        gross_pnl = (exit_price - entry_price) * position_size
                    else:  # short
                        gross_pnl = (entry_price - exit_price) * position_size
                    
                    # Net PnL after fees
                    net_pnl = gross_pnl - total_fees
                    
                    # Calculate risk with integrated fee
                    if direction == 'long':
                        # Get fee percentages
                        taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100
                        maker_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02)) / 100
                        
                        # Determine entry fee percentage based on order type
                        entry_fee_pct = maker_fee_percent if order_type == 'Limit' else taker_fee_percent
                        
                        # Price risk component
                        price_risk_per_unit = entry_price - stop_loss if stop_loss > 0 else 0
                        price_risk = price_risk_per_unit * position_size
                        
                        # Fee risk component (estimated based on entry and stop loss)
                        if stop_loss > 0:
                            fee_cost_per_unit = entry_fee_pct * entry_price + taker_fee_percent * stop_loss
                            fee_risk = fee_cost_per_unit * position_size
                        else:
                            fee_risk = 0
                        
                        # Total risk amount
                        risk_amount = price_risk + fee_risk
                    else:  # short
                        # Get fee percentages
                        taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100
                        maker_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02)) / 100
                        
                        # Determine entry fee percentage based on order type
                        entry_fee_pct = maker_fee_percent if order_type == 'Limit' else taker_fee_percent
                        
                        # Price risk component
                        price_risk_per_unit = stop_loss - entry_price if stop_loss > 0 else 0
                        price_risk = price_risk_per_unit * position_size
                        
                        # Fee risk component (estimated based on entry and stop loss)
                        if stop_loss > 0:
                            fee_cost_per_unit = entry_fee_pct * entry_price + taker_fee_percent * stop_loss
                            fee_risk = fee_cost_per_unit * position_size
                        else:
                            fee_risk = 0
                        
                        # Total risk amount
                        risk_amount = price_risk + fee_risk
                    
                    # Calculate R multiple based on net PnL
                    r_multiple = net_pnl / risk_amount if risk_amount > 0 else 0
                    
                    # Format dates
                    entry_time = trade_data.get('entry_time', '')
                    exit_time = trade_data.get('exit_time', '')
                    
                    # Try to parse ISO format dates
                    try:
                        if isinstance(entry_time, str):
                            entry_time = datetime.fromisoformat(entry_time.replace('Z', '+00:00'))
                        if isinstance(exit_time, str):
                            exit_time = datetime.fromisoformat(exit_time.replace('Z', '+00:00'))
                    except ValueError:
                        # If parsing fails, keep as is
                        pass
                    
                    # Determine column positions based on headers
                    fee_col = headers.index("Fees") + 1 if "Fees" in headers else None
                    net_pnl_col = headers.index("Net P&L") + 1 if "Net P&L" in headers else None
                    gross_pnl_col = headers.index("Gross P&L") + 1 if "Gross P&L" in headers else headers.index("Return $") + 1 if "Return $" in headers else None
                    
                    # Populate the row with base data
                    ws.cell(row=next_row, column=1).value = trade_number
                    ws.cell(row=next_row, column=2).value = trade_data.get('symbol', '')
                    ws.cell(row=next_row, column=3).value = direction.upper()
                    ws.cell(row=next_row, column=4).value = entry_time
                    ws.cell(row=next_row, column=5).value = exit_time
                    ws.cell(row=next_row, column=6).value = entry_price
                    ws.cell(row=next_row, column=7).value = stop_loss
                    ws.cell(row=next_row, column=8).value = exit_price
                    ws.cell(row=next_row, column=9).value = position_size
                    ws.cell(row=next_row, column=10).value = risk_amount
                    
                    # Add gross PnL, fees, and net PnL
                    if gross_pnl_col:
                        ws.cell(row=next_row, column=gross_pnl_col).value = gross_pnl
                    
                    if fee_col:
                        ws.cell(row=next_row, column=fee_col).value = total_fees
                    
                    if net_pnl_col:
                        ws.cell(row=next_row, column=net_pnl_col).value = net_pnl
                    
                    # R multiple based on net PnL
                    r_multiple_col = headers.index("R Multiple") + 1 if "R Multiple" in headers else None
                    if r_multiple_col:
                        ws.cell(row=next_row, column=r_multiple_col).value = r_multiple
                    
                    # Account balance (pre-trade)
                    balance_col = headers.index("Account Balance") + 1 if "Account Balance" in headers else None
                    if balance_col:
                        ws.cell(row=next_row, column=balance_col).value = trade_data.get('account_balance', 0)
                    
                    # Post-trade balance
                    if "Post-Trade Balance" in headers:
                        post_trade_balance_col = headers.index("Post-Trade Balance") + 1
                        ws.cell(row=next_row, column=post_trade_balance_col).value = trade_data.get('post_trade_balance', 0)
                    else:
                        # If Post-Trade Balance column doesn't exist yet, add it
                        if balance_col:
                            # Add new column header
                            post_trade_col = balance_col + 1
                            # Shift existing data if necessary
                            for row in range(1, next_row):
                                for col in range(ws.max_column, post_trade_col - 1, -1):
                                    ws.cell(row=row, column=col+1).value = ws.cell(row=row, column=col).value
                            
                            # Add the new header
                            ws.cell(row=1, column=post_trade_col).value = "Post-Trade Balance"
                            ws.cell(row=1, column=post_trade_col).fill = self.yellow_fill
                            
                            # Add the value
                            ws.cell(row=next_row, column=post_trade_col).value = trade_data.get('post_trade_balance', 0)
                    
                    # Add order type
                    order_type_col = headers.index("Order Type") + 1 if "Order Type" in headers else None
                    if order_type_col:
                        ws.cell(row=next_row, column=order_type_col).value = order_type
                    else:
                        # If Order Type column doesn't exist yet, add it
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Order Type"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                        ws.cell(row=next_row, column=new_col).value = order_type
                    
                    # Add exit order type
                    exit_order_type_col = headers.index("Exit Order Type") + 1 if "Exit Order Type" in headers else None
                    if exit_order_type_col:
                        ws.cell(row=next_row, column=exit_order_type_col).value = exit_order_type
                    else:
                        # If Exit Order Type column doesn't exist yet, add it
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Exit Order Type"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                        ws.cell(row=next_row, column=new_col).value = exit_order_type
                    
                    # Add fee percentages
                    entry_fee_pct_col = headers.index("Entry Fee %") + 1 if "Entry Fee %" in headers else None
                    if entry_fee_pct_col:
                        ws.cell(row=next_row, column=entry_fee_pct_col).value = entry_fee_percent
                    else:
                        # If Entry Fee % column doesn't exist yet, add it
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Entry Fee %"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                        ws.cell(row=next_row, column=new_col).value = entry_fee_percent
                    
                    exit_fee_pct_col = headers.index("Exit Fee %") + 1 if "Exit Fee %" in headers else None
                    if exit_fee_pct_col:
                        ws.cell(row=next_row, column=exit_fee_pct_col).value = exit_fee_percent
                    else:
                        # If Exit Fee % column doesn't exist yet, add it
                        new_col = ws.max_column + 1
                        ws.cell(row=1, column=new_col).value = "Exit Fee %"
                        ws.cell(row=1, column=new_col).fill = self.yellow_fill
                        ws.cell(row=next_row, column=new_col).value = exit_fee_percent

                    # Log detailed risk breakdown for verification
                    logger.info(f"Excel trade log: {trade_data.get('symbol')} {direction} - " + 
                              f"Price risk: ${price_risk:.2f}, Fee risk: ${fee_risk:.2f}, " + 
                              f"Total risk: ${risk_amount:.2f}, " + 
                              f"Fees: Entry {entry_fee_percent:.3f}% ({order_type}), Exit {exit_fee_percent:.3f}% ({exit_order_type})")

                    # Save the workbook
                    wb.save(self.excel_path)
                    logger.info(f"Logged trade #{trade_number} to Excel: {trade_data.get('symbol')} {direction}")
                    break
                    
                except PermissionError:
                    if attempt < max_retries - 1:
                        logger.warning(f"Excel file is open or locked. Retrying in {retry_delay} seconds... (Attempt {attempt+1}/{max_retries})")
                        time.sleep(retry_delay)
                    else:
                        logger.error("Failed to access Excel file after multiple attempts. Make sure the file is not open in another program.")
                        raise
                
        except Exception as e:
            logger.error(f"Error logging trade to Excel: {str(e)}")
            raise

    def export_from_database(self, db_path, limit=None):
        """
        Export trades from SQLite database to Excel
        
        Args:
            db_path (str): Path to the SQLite database
            limit (int, optional): Maximum number of trades to export. Defaults to None (all trades).
        """
        try:
            import sqlite3
            
            # Connect to database
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # Check if fee percentage and order type columns exist
            cursor.execute("PRAGMA table_info(trade_history)")
            columns = [col[1] for col in cursor.fetchall()]
            
            has_entry_fee_percent = 'entry_fee_percent' in columns
            has_exit_fee_percent = 'exit_fee_percent' in columns
            has_exit_order_type = 'exit_order_type' in columns
            
            # Query trades with additional columns if they exist
            select_columns = [
                'symbol', 'direction', 'entry_price', 'exit_price', 'position_size',
                'pnl', 'pnl_percent', 'entry_time', 'exit_time', 'duration_minutes',
                'stop_loss', 'stopped_out', 'reason', 'total_fees', 'entry_fee', 'exit_fee',
                'gross_pnl', 'order_type'
            ]
            
            # Add new columns if they exist
            if has_exit_order_type:
                select_columns.append('exit_order_type')
            
            if has_entry_fee_percent:
                select_columns.append('entry_fee_percent')
                
            if has_exit_fee_percent:
                select_columns.append('exit_fee_percent')
            
            # Construct query
            query = f'SELECT {", ".join(select_columns)} FROM trade_history ORDER BY exit_time DESC'
            if limit:
                query += f' LIMIT {limit}'
                
            cursor.execute(query)
            trades = cursor.fetchall()
            
            # Get current account balance
            account_balance = 0
            try:
                cursor.execute('SELECT MAX(capital) FROM capital_history')
                result = cursor.fetchone()
                if result and result[0]:
                    account_balance = result[0]
            except:
                # If table doesn't exist or query fails
                pass
            
            # Convert to list of dictionaries
            for trade in trades:
                trade_dict = dict(zip(select_columns, trade))
                
                # Add post_trade_balance (original balance + pnl)
                trade_dict['account_balance'] = account_balance - trade_dict.get('pnl', 0)
                trade_dict['post_trade_balance'] = account_balance
                
                # For older records without exit_order_type
                if not has_exit_order_type:
                    trade_dict['exit_order_type'] = 'Market'
                
                # For older records without fee percentages
                if not has_entry_fee_percent:
                    # Calculate based on order type
                    if trade_dict.get('order_type') == 'Limit':
                        trade_dict['entry_fee_percent'] = float(config['TRADING'].get('maker_fee_percent', 0.02))
                    else:
                        trade_dict['entry_fee_percent'] = float(config['TRADING'].get('trading_fee_percent', 0.075))
                        
                if not has_exit_fee_percent:
                    # Calculate based on exit order type
                    if trade_dict.get('exit_order_type') == 'Limit':
                        trade_dict['exit_fee_percent'] = float(config['TRADING'].get('maker_fee_percent', 0.02))
                    else:
                        trade_dict['exit_fee_percent'] = float(config['TRADING'].get('trading_fee_percent', 0.075))
                
                # Log the trade to Excel
                self.log_trade(trade_dict)
                
                # Update account balance for next trade
                account_balance -= trade_dict.get('pnl', 0)
            
            conn.close()
            logger.info(f"Exported {len(trades)} trades from database to Excel")
            
        except Exception as e:
            logger.error(f"Error exporting trades from database: {str(e)}")
            raise

    def backup_excel(self, backup_dir=None):
        """
        Create a backup of the Excel file
        
        Args:
            backup_dir (str, optional): Directory to store backup. Defaults to None (same directory).
        """
        try:
            if not os.path.exists(self.excel_path):
                logger.warning("Cannot backup - Excel file does not exist")
                return
                
            # Determine backup path
            if backup_dir:
                os.makedirs(backup_dir, exist_ok=True)
                backup_path = os.path.join(backup_dir, f"backup_{os.path.basename(self.excel_path)}")
            else:
                file_name, ext = os.path.splitext(self.excel_path)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = f"{file_name}_backup_{timestamp}{ext}"
            
            # Copy file
            import shutil
            shutil.copy2(self.excel_path, backup_path)
            logger.info(f"Created backup of Excel file at {backup_path}")
            
        except Exception as e:
            logger.error(f"Error creating Excel backup: {str(e)}")




# Initialize the Excel logger
def init_excel_logger():
    """Initialize the Excel Logger"""
    global excel_logger
    
    if config.getboolean('EXCEL_LOGGING', 'enabled', fallback=True):
        excel_path = config['EXCEL_LOGGING'].get('excel_path', 'trading_log.xlsx')
        try:
            # Make path absolute if it's relative
            if not os.path.isabs(excel_path):
                excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), excel_path)
                
            excel_logger = ExcelTradeLogger(excel_path)
            logger.info(f"Excel logging initialized with file: {excel_path}")
            
            # If auto-backup is enabled, set up a thread to periodically back up the Excel file
            if config.getboolean('EXCEL_LOGGING', 'auto_backup', fallback=True):
                backup_interval = int(config['EXCEL_LOGGING'].get('backup_interval_hours', 24)) * 3600  # Convert to seconds
                
                def backup_thread():
                    while True:
                        try:
                            time.sleep(backup_interval)
                            excel_logger.backup_excel()
                        except Exception as e:
                            logger.error(f"Error in Excel backup thread: {str(e)}")
                
                backup_thread = threading.Thread(target=backup_thread, daemon=True)
                backup_thread.start()
                logger.info(f"Excel auto-backup scheduled every {backup_interval//3600} hours")
                
        except Exception as e:
            logger.error(f"Failed to initialize Excel logging: {str(e)}")
    else:
        logger.info("Excel logging is disabled in configuration")


# Helper function to get the account balance for a position
def get_position_account_balance(symbol, entry_time):
    """Get the account balance at the time of position entry"""
    try:
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Try to get the account balance from the positions table
        cursor.execute('SELECT account_balance FROM positions WHERE symbol = ? AND entry_time = ?', 
                      (symbol, entry_time))
        result = cursor.fetchone()
        conn.close()
        
        if result and result[0]:
            return result[0]
        
        # If not found or NULL, return current capital
        return get_current_capital()
    except Exception as e:
        logger.error(f"Error getting position account balance: {str(e)}")
        return get_current_capital()



# Helper function to get detailed position information from Bybit
def get_bybit_position_details(client, symbol):
    """Get detailed position information from Bybit, including entry price and PnL"""
    try:
        response = client.get_positions(
            category="linear",
            symbol=symbol
        )
        
        if response.get('retCode') == 0:
            positions_list = response.get('result', {}).get('list', [])
            for pos in positions_list:
                if pos.get('symbol') == symbol and float(pos.get('size', 0)) > 0:
                    position_details = {
                        'direction': 'long' if pos.get('side') == 'Buy' else 'short',
                        'position_size': float(pos.get('size', 0)),
                        'entry_price': float(pos.get('avgPrice', 0)),
                        'unrealized_pnl': float(pos.get('unrealisedPnl', 0)),
                        'stop_loss': float(pos.get('stopLoss', 0)) if pos.get('stopLoss') else None,
                        'take_profit': float(pos.get('takeProfit', 0)) if pos.get('takeProfit') else None
                    }
                    return position_details
        
        logger.warning(f"No active position found for {symbol} on Bybit")
        return None
    except Exception as e:
        logger.error(f"Error getting position details from Bybit: {str(e)}")
        return None



def get_exact_position_size_from_exchange(client, symbol):
    """Get the exact position size from Bybit for a symbol"""
    try:
        response = client.get_positions(
            category="linear",
            symbol=symbol
        )
        
        if response.get('retCode') == 0:
            positions_list = response.get('result', {}).get('list', [])
            for pos in positions_list:
                if pos.get('symbol') == symbol and float(pos.get('size', 0)) > 0:
                    exact_size = float(pos.get('size', 0))
                    logger.info(f"Retrieved exact position size from Bybit for {symbol}: {exact_size}")
                    return exact_size
        
        logger.warning(f"No active position found on Bybit for {symbol}")
        return None
    except Exception as e:
        logger.error(f"Error getting exact position size from Bybit: {str(e)}")
        return None


# Helper function to get accurate exit price from Bybit
def get_accurate_bybit_exit_price(client, symbol, order_id=None, side=None):
    """Try multiple methods to get the most accurate exit price from Bybit"""
    try:
        # Method 1: Try to get from order details if we have the order ID
        if order_id:
            order_detail = client.get_order_history(
                category="linear",
                symbol=symbol,
                orderId=order_id
            )
            
            if order_detail.get('retCode') == 0:
                order_result = order_detail.get('result', {}).get('list', [{}])[0]
                if float(order_result.get('execQty', 0)) > 0:
                    return float(order_result.get('avgPrice', 0))
        
        # Method 2: Get execution history for stop loss cases (more accurate than order history for stops)
        try:
            # This uses the execution history endpoint which has the actual execution details
            executions = client.get_executions(
                category="linear",
                symbol=symbol,
                limit=20  # Increase limit to catch recent executions
            )
            
            if executions.get('retCode') == 0:
                exec_list = executions.get('result', {}).get('list', [])
                
                # Expected side will be opposite of position direction for exits
                expected_side = "Buy" if side == "short" else "Sell"
                
                # Sort by execution time, most recent first (if available)
                if exec_list and 'execTime' in exec_list[0]:
                    exec_list.sort(key=lambda x: int(x.get('execTime', 0)), reverse=True)
                
                # Find matching executions
                for exec_info in exec_list:
                    if exec_info.get('side') == expected_side and float(exec_info.get('execQty', 0)) > 0:
                        logger.info(f"Found actual execution price from execution history: {exec_info.get('execPrice')} for {symbol}")
                        return float(exec_info.get('execPrice', 0))
        except Exception as exec_error:
            logger.warning(f"Could not retrieve execution history: {str(exec_error)}, trying order history")
        
        # Method 3: Get recent orders for this symbol (fallback)
        recent_orders = client.get_order_history(
            category="linear",
            symbol=symbol,
            limit=20  # Increased from 10 to catch more orders
        )
        
        if recent_orders.get('retCode') == 0:
            orders = recent_orders.get('result', {}).get('list', [])
            # Look for the most recent order that matches the side (Buy for short exit, Sell for long exit)
            expected_side = "Buy" if side == "short" else "Sell"
            for order in orders:
                if (order.get('symbol') == symbol and 
                    order.get('side') == expected_side and 
                    float(order.get('execQty', 0)) > 0):
                    logger.info(f"Found actual exit price from order history: {order.get('avgPrice')} for {symbol}")
                    return float(order.get('avgPrice', 0))
        
        # Method 4: Try to get position close data from closed PnL record
        try:
            pnl_records = client.get_closed_pnl(
                category="linear",
                symbol=symbol,
                limit=10
            )
            
            if pnl_records.get('retCode') == 0:
                records = pnl_records.get('result', {}).get('list', [])
                if records:
                    # Records should be ordered by close time (most recent first)
                    for record in records:
                        # Check if this is the record we're looking for
                        if record.get('symbol') == symbol:
                            close_price = float(record.get('closePrice', 0))
                            if close_price > 0:
                                logger.info(f"Found exit price from closed PnL record: {close_price} for {symbol}")
                                return close_price
        except Exception as pnl_error:
            logger.warning(f"Could not retrieve closed PnL records: {str(pnl_error)}")
        
        # Method 5: Get current price as absolute last fallback
        ticker_response = client.get_tickers(category="linear", symbol=symbol)
        if ticker_response.get('retCode') == 0:
            last_price = float(ticker_response.get('result', {}).get('list', [{}])[0].get('lastPrice', 0))
            logger.warning(f"Could not get accurate exit price for {symbol}, using last traded price: {last_price}")
            return last_price
        
        logger.warning(f"Could not get accurate exit price for {symbol}, returning None")
        return None
    except Exception as e:
        logger.error(f"Error getting accurate exit price: {str(e)}")
        return None


# IMPROVEMENT: Position Persistence with SQLite
def init_database():
    """Initialize SQLite database for position persistence"""
    try:
        # Create positions table
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS positions (
            symbol TEXT PRIMARY KEY,
            direction TEXT,
            entry_price REAL,
            position_size REAL,
            stop_loss REAL,
            order_id TEXT,
            entry_time TEXT,
            sl_order_id TEXT
        )
        ''')
        
        # Create trade history table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS trade_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            symbol TEXT,
            direction TEXT,
            entry_price REAL,
            exit_price REAL,
            position_size REAL,
            pnl REAL,
            pnl_percent REAL,
            entry_time TEXT,
            exit_time TEXT,
            duration_minutes INTEGER,
            stop_loss REAL,
            stopped_out BOOLEAN
        )
        ''')
        
        # Create capital history table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS capital_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            capital REAL,
            change REAL,
            trade_id INTEGER,
            FOREIGN KEY (trade_id) REFERENCES trade_history (id)
        )
        ''')
        
        conn.commit()
        conn.close()
        logger.info("Database initialized successfully")
    except Exception as e:
        logger.error(f"Error initializing database: {str(e)}")



def init_enhanced_database():
    """Initialize enhanced database for improved reconciliation tracking"""
    try:
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Check if reason column exists in trade_history
        cursor.execute("PRAGMA table_info(trade_history)")
        columns = [col[1] for col in cursor.fetchall()]
        
        # Add reason column if it doesn't exist
        if 'reason' not in columns:
            logger.info("Adding 'reason' column to trade_history table")
            cursor.execute('ALTER TABLE trade_history ADD COLUMN reason TEXT')
            
        # Check if last_check_time column exists in positions
        cursor.execute("PRAGMA table_info(positions)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'last_check_time' not in columns:
            logger.info("Adding 'last_check_time' column to positions table")
            cursor.execute('ALTER TABLE positions ADD COLUMN last_check_time TEXT')
            
        if 'take_profit' not in columns:
            logger.info("Adding 'take_profit' column to positions table")
            cursor.execute('ALTER TABLE positions ADD COLUMN take_profit REAL')

        if 'account_balance' not in columns: 
            logger.info("Adding 'account_balance' column to positions table")
            cursor.execute('ALTER TABLE positions ADD COLUMN account_balance REAL')
            
        if 'order_type' not in columns:
            logger.info("Adding 'order_type' column to positions table")
            cursor.execute('ALTER TABLE positions ADD COLUMN order_type TEXT')
            
        if 'tp_order_type' not in columns:
            logger.info("Adding 'tp_order_type' column to positions table")
            cursor.execute('ALTER TABLE positions ADD COLUMN tp_order_type TEXT')
            
        # Add tp_order_id column to positions table if it doesn't exist
        if 'tp_order_id' not in columns:
            logger.info("Adding 'tp_order_id' column to positions table")
            cursor.execute('ALTER TABLE positions ADD COLUMN tp_order_id TEXT')
        
        # Check if necessary columns exist in pending_orders table
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pending_orders'")
        if cursor.fetchone():
            cursor.execute("PRAGMA table_info(pending_orders)")
            columns = [col[1] for col in cursor.fetchall()]
            
            # Add missing columns if needed
            if 'use_tp_limit_order' not in columns:
                cursor.execute('ALTER TABLE pending_orders ADD COLUMN use_tp_limit_order BOOLEAN')
            if 'stop_loss' not in columns:
                cursor.execute('ALTER TABLE pending_orders ADD COLUMN stop_loss REAL')
            if 'take_profit' not in columns:
                cursor.execute('ALTER TABLE pending_orders ADD COLUMN take_profit REAL')
        
        # Add order_type column to trade_history if it doesn't exist
        cursor.execute("PRAGMA table_info(trade_history)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'order_type' not in columns:
            logger.info("Adding 'order_type' column to trade_history table")
            cursor.execute('ALTER TABLE trade_history ADD COLUMN order_type TEXT')

        # Create reconciliation_log table for detailed tracking
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS reconciliation_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            event_type TEXT,
            symbol TEXT,
            details TEXT,
            position_data TEXT
        )
        ''')
        
        # Create pending_orders table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS pending_orders (
            order_id TEXT PRIMARY KEY,
            symbol TEXT,
            order_type TEXT,
            side TEXT,
            price REAL,
            qty REAL,
            status TEXT,
            created_time TEXT,
            purpose TEXT,
            related_position_symbol TEXT,
            expiry_time TEXT
        )
        ''')
        
        conn.commit()
        conn.close()
        
        # Initialize max loss feature tables
        init_max_loss_tables()
        
        logger.info("Enhanced database initialized successfully")
    except Exception as e:
        logger.error(f"Error initializing enhanced database: {str(e)}")




def add_balance_history_table(conn, cursor):
    """Create balance history table for tracking account balance over time"""
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS balance_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT,
        balance REAL,
        source TEXT
    )
    ''')
    
    # Create index for faster timestamp-based queries
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_balance_history_timestamp ON balance_history(timestamp)')
    
    logger.info("Balance history table created/verified")


def add_max_loss_events_table(conn, cursor):
    """Create max loss events table to record when max loss is triggered"""
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS max_loss_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT,
        current_balance REAL,
        reference_balance REAL,
        loss_percent REAL,
        positions_closed INTEGER,
        total_pnl REAL
    )
    ''')
    
    logger.info("Max loss events table created/verified")


def init_max_loss_tables():
    """Initialize database tables for max daily loss feature"""
    try:
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Add balance history table
        add_balance_history_table(conn, cursor)
        
        # Add max loss events table
        add_max_loss_events_table(conn, cursor)
        
        conn.commit()
        conn.close()
        logger.info("Max daily loss tables initialized")
    except Exception as e:
        logger.error(f"Error initializing max loss tables: {str(e)}")



def log_reconciliation_event(event_type, symbol, details, position_data=None):
    """Log reconciliation events to the database using the queue system"""
    # Queue the database operation
    def callback(success, _):
        if not success:
            logger.warning(f"Failed to log reconciliation event: {event_type} for {symbol}")
    
    # Add to the queue
    queue_operation(db_log_reconciliation_event, (event_type, symbol, details, position_data), callback)

def save_position(symbol, position_data):
    """Save position to SQLite database using the queue system"""
    # Update memory immediately
    global active_positions
    active_positions[symbol] = position_data
    
    # Queue the database operation
    def callback(success, _):
        if success:
            logger.info(f"Position saved to database: {symbol}")
        else:
            logger.warning(f"Failed to save position to database: {symbol}")
    
    # Make a deep copy of position_data to avoid potential modification during queue wait
    position_data_copy = position_data.copy()
    
    # Add to the queue
    queue_operation(db_save_position, (symbol, position_data_copy), callback)
    return True

def load_positions():
    """Load positions from SQLite database"""
    positions = {}
    try:
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Check table structure to determine how to load
        cursor.execute("PRAGMA table_info(positions)")
        columns = [col[1] for col in cursor.fetchall()]
        
        cursor.execute('SELECT * FROM positions')
        rows = cursor.fetchall()
        
        for row in rows:
            position_data = {
                'direction': row[1],
                'entry_price': row[2],
                'position_size': row[3],
                'stop_loss': row[4],
                'order_id': row[5],
                'entry_time': row[6],
                'sl_order_id': row[7]
            }
            
            # Add additional fields if they exist in the schema
            if 'last_check_time' in columns and len(row) > 8:
                position_data['last_check_time'] = row[8]
                
            if 'take_profit' in columns and len(row) > 9:
                position_data['take_profit'] = row[9]

            if 'account_balance' in columns and len(row) > 10:
                position_data['account_balance'] = row[10]
                
            if 'order_type' in columns and len(row) > 11:
                position_data['order_type'] = row[11]
                
            if 'tp_order_type' in columns and len(row) > 12:
                position_data['tp_order_type'] = row[12]

            if 'tp_order_id' in columns and len(row) > 13:
                position_data['tp_order_id'] = row[13]
  
            positions[row[0]] = position_data
            
        conn.close()
        logger.info(f"Loaded {len(positions)} positions from database")
        return positions
    except Exception as e:
        logger.error(f"Error loading positions from database: {str(e)}")
        return {}


def save_pending_order(order_data):
    """Save pending order to SQLite database using the queue system"""
    # Update memory immediately
    global pending_orders
    pending_orders[order_data['order_id']] = order_data
    
    # Queue the database operation
    def callback(success, _):
        if success:
            logger.info(f"Pending order saved to database: {order_data['order_id']} for {order_data['symbol']}")
        else:
            logger.info(f"Pending order saved to memory only: {order_data['order_id']} for {order_data['symbol']}")
    
    # Make a deep copy of order_data to avoid potential modification during queue wait
    order_data_copy = order_data.copy()
    
    # Add to the queue
    queue_operation(db_save_pending_order, (order_data_copy,), callback)
    return True

def load_pending_orders():
    """Load pending orders from SQLite database"""
    orders = {}
    try:
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Check if table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pending_orders'")
        if not cursor.fetchone():
            conn.close()
            return {}
        
        cursor.execute('SELECT * FROM pending_orders')
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        
        for row in rows:
            order_data = dict(zip(columns, row))
            orders[order_data['order_id']] = order_data
            
        conn.close()
        logger.info(f"Loaded {len(orders)} pending orders from database")
        return orders
    except Exception as e:
        logger.error(f"Error loading pending orders from database: {str(e)}")
        return {}

def verify_pending_orders_on_startup():
    """Verify all pending orders on bot startup"""
    try:
        logger.info("Verifying pending orders on startup")
        
        if not pending_orders:
            logger.info("No pending orders to verify")
            return
        
        client = initialize_bybit_client()
        current_time = datetime.now()
        verified = 0
        expired = 0
        not_found = 0
        
        for order_id, order_data in list(pending_orders.items()):
            # Check if already expired
            if 'expiry_time' in order_data and order_data['expiry_time']:
                try:
                    expiry_time = datetime.fromisoformat(order_data['expiry_time'])
                    if current_time > expiry_time:
                        logger.info(f"Order {order_id} for {order_data['symbol']} has expired, canceling")
                        cancel_order(client, order_id, order_data['symbol'])
                        remove_pending_order(order_id)
                        expired += 1
                        continue
                except Exception as e:
                    logger.warning(f"Error parsing expiry time for order {order_id}: {str(e)}")
            
            # Verify with Bybit that order still exists
            try:
                order_status = check_pending_order_status(client, order_id, order_data)
                if not order_status:
                    # Check if this might be a trigger order (SL/TP) that isn't visible
                    # in frontend_open_orders - don't remove it, just skip verification
                    is_trigger = (order_data.get('is_stop_loss', False) or 
                                  order_data.get('stop_order_type', '') in ['StopLoss', 'StopMarket', 'TakeProfit'])
                    if is_trigger:
                        logger.info(f"Order {order_id} is a trigger order not visible via API - keeping in tracking")
                        continue
                    logger.warning(f"Order {order_id} not found on Bybit, removing from tracking")
                    remove_pending_order(order_id)
                    not_found += 1
                    continue
                
                # Update order status
                order_data['status'] = order_status['status']
                save_pending_order(order_data)
                verified += 1
                
                # Process filled orders
                if order_status['status'] == 'Filled':
                    process_filled_order(order_data)
            except Exception as e:
                logger.warning(f"Error verifying order {order_id}: {str(e)}")
                # Don't remove the order yet - might be temporary connection issue
        
        logger.info(f"Pending orders verification complete: {verified} verified, {expired} expired, {not_found} not found")
    except Exception as e:
        logger.error(f"Error verifying pending orders: {str(e)}")



def remove_pending_order(order_id):
    """Remove pending order from database using the queue system"""
    # Remove from memory immediately
    global pending_orders
    if order_id in pending_orders:
        symbol = pending_orders[order_id].get('symbol', 'Unknown')
        del pending_orders[order_id]
    else:
        symbol = 'Unknown'
    
    # Queue the database operation
    def callback(success, _):
        if success:
            logger.info(f"Pending order removed from database: {order_id}")
        else:
            logger.warning(f"Failed to remove pending order from database: {order_id}")
    
    # Add to the queue
    queue_operation(db_remove_pending_order, (order_id,), callback)
    return True


def remove_position(symbol):
    """Remove position from database using the queue system"""
    # Remove from memory immediately
    global active_positions, recently_closed_positions
    
    if symbol in active_positions:
        del active_positions[symbol]
    
    # Mark position as recently closed with current timestamp
    recently_closed_positions[symbol] = datetime.now()
    
    # Queue the database operation
    def callback(success, _):
        if success:
            logger.info(f"Position removed from database: {symbol}")
        else:
            logger.warning(f"Failed to remove position from database: {symbol}")
    
    # Add to the queue
    queue_operation(db_remove_position, (symbol,), callback)
    return True




def update_position_check_time(symbol):
    """Update the last_check_time for a position using the queue system"""
    # Queue the database operation
    def callback(success, _):
        if not success:
            logger.debug(f"Failed to update position check time for {symbol}")
    
    # Add to the queue
    queue_operation(db_update_position_check_time, (symbol,), callback)





def record_completed_trade(symbol, direction, entry_price, exit_price, position_size, entry_time, stop_loss=None, stopped_out=False, reason=None, order_id=None, order_type='Market', is_partial_fill=False):
    """Record completed trade to trade history database with reason field support and accurate price data"""
    try:
        client = initialize_bybit_client()
        
        # Only cancel related orders if this is a FULL exit, not a partial fill
        if not is_partial_fill:
            # Make sure we cancel any related TP orders
            cancel_related_tp_orders(client, symbol)
            
            # LOCAL DATABASE CLEANUP ONLY: Remove stop loss orders from tracking without affecting the exchange
            remove_related_stop_loss_orders_from_database(symbol)
        
        # First priority: Try to get a more accurate exit price from Bybit
        if order_id:
            # Try using the order ID first if we have it
            accurate_exit_price = get_accurate_bybit_exit_price(client, symbol, order_id, direction)
            if accurate_exit_price:
                exit_price = accurate_exit_price
                logger.info(f"Using accurate exit price from order ID: {exit_price} for {symbol}")
        # If no order ID or couldn't get price from order ID, try again using just symbol and direction
        elif not order_id or exit_price == stop_loss:
            # This helps for stop loss cases where we don't have the order ID
            accurate_exit_price = get_accurate_bybit_exit_price(client, symbol, None, direction)
            if accurate_exit_price:
                exit_price = accurate_exit_price
                logger.info(f"Using accurate exit price from Bybit API: {exit_price} for {symbol}")
        
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Get entry order type from active positions or database
        entry_order_type = 'Market'  # Default
        exit_order_type = order_type  # Use the provided order_type for exit
        
        try:
            if symbol in active_positions:
                entry_order_type = active_positions[symbol].get('order_type', 'Market')
            else:
                # Try to get from database
                cursor.execute('SELECT order_type FROM positions WHERE symbol = ?', (symbol,))
                result = cursor.fetchone()
                if result and result[0]:
                    entry_order_type = result[0]
        except Exception as e:
            logger.warning(f"Error getting entry order type: {str(e)}, using default 'Market'")
        
        # Get fee percentages
        taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100  # For market orders
        maker_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02)) / 100     # For limit orders
        
        # Determine entry and exit fees based on order types
        entry_fee_percent = maker_fee_percent if entry_order_type == 'Limit' else taker_fee_percent
        exit_fee_percent = maker_fee_percent if exit_order_type == 'Limit' else taker_fee_percent
        
        # If stopped out, always use taker fee for exit (market order)
        if stopped_out:
            exit_fee_percent = taker_fee_percent
            exit_order_type = 'Market'
        
        # Calculate trading fees using the appropriate fee percentages
        entry_fee = entry_price * position_size * entry_fee_percent
        exit_fee = exit_price * position_size * exit_fee_percent
        total_fees = entry_fee + exit_fee
        
        logger.info(f"Trading fees for {symbol}: " + 
                   f"Entry ({entry_fee_percent*100:.3f}% {entry_order_type}): ${entry_fee:.2f}, " + 
                   f"Exit ({exit_fee_percent*100:.3f}% {exit_order_type}): ${exit_fee:.2f}, " + 
                   f"Total: ${total_fees:.2f}")
        
        # Calculate P&L (now subtracting fees for net PnL)
        if direction == 'long':
            gross_pnl = (exit_price - entry_price) * position_size
            net_pnl = gross_pnl - total_fees
        else:
            gross_pnl = (entry_price - exit_price) * position_size
            net_pnl = gross_pnl - total_fees
        
        # Calculate risk amount - now including fees in the calculation
        if direction == 'long':
            price_risk_per_unit = entry_price - (stop_loss or 0)
        else:
            price_risk_per_unit = (stop_loss or 0) - entry_price
            
        # Only calculate if stop_loss is provided
        if stop_loss:
            # Use the entry fee percentage determined above
            fee_cost_per_unit = entry_fee_percent * entry_price + taker_fee_percent * (stop_loss or 0)
            total_risk_per_unit = price_risk_per_unit + fee_cost_per_unit
            risk_amount = total_risk_per_unit * position_size
            
            # Calculate the components for display
            price_risk = price_risk_per_unit * position_size
            fee_risk = fee_cost_per_unit * position_size
        else:
            risk_amount = 0
            price_risk = 0
            fee_risk = 0
        
        # Calculate P&L percentage (relative to risk)
        pnl_percent = (net_pnl / risk_amount) * 100 if risk_amount > 0 else 0
        
        # Calculate duration
        try:
            entry_dt = datetime.fromisoformat(entry_time)
        except ValueError:
            # Handle ISO format with timezone
            entry_dt = datetime.fromisoformat(entry_time.replace('Z', '+00:00'))
        except Exception:
            # Fallback if entry_time format is invalid
            entry_dt = datetime.now() - timedelta(hours=1)
            
        exit_dt = datetime.now()
        duration_minutes = int((exit_dt - entry_dt).total_seconds() / 60)
        
        # Check if columns exist in trade_history
        cursor.execute("PRAGMA table_info(trade_history)")
        columns = [col[1] for col in cursor.fetchall()]
        
        # Add fee columns if they don't exist
        if 'total_fees' not in columns:
            cursor.execute('ALTER TABLE trade_history ADD COLUMN total_fees REAL')
            cursor.execute('ALTER TABLE trade_history ADD COLUMN entry_fee REAL')
            cursor.execute('ALTER TABLE trade_history ADD COLUMN exit_fee REAL')
            cursor.execute('ALTER TABLE trade_history ADD COLUMN gross_pnl REAL')
        
        # Add order_type column if it doesn't exist
        if 'order_type' not in columns:
            cursor.execute('ALTER TABLE trade_history ADD COLUMN order_type TEXT')
        
        # Add exit_order_type column if it doesn't exist
        if 'exit_order_type' not in columns:
            cursor.execute('ALTER TABLE trade_history ADD COLUMN exit_order_type TEXT')
        
        # Add fee percentage columns if they don't exist
        if 'entry_fee_percent' not in columns:
            cursor.execute('ALTER TABLE trade_history ADD COLUMN entry_fee_percent REAL')
        if 'exit_fee_percent' not in columns:
            cursor.execute('ALTER TABLE trade_history ADD COLUMN exit_fee_percent REAL')
        
        # Check if reason column exists
        if 'reason' in columns and reason:
            cursor.execute('''
            INSERT INTO trade_history (symbol, direction, entry_price, exit_price, position_size, 
                              pnl, pnl_percent, entry_time, exit_time, duration_minutes, 
                              stop_loss, stopped_out, reason, total_fees, entry_fee, exit_fee, gross_pnl, 
                              order_type, exit_order_type, entry_fee_percent, exit_fee_percent)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                symbol, direction, entry_price, exit_price, position_size,
                net_pnl, pnl_percent, entry_time, exit_dt.isoformat(), duration_minutes,
                stop_loss or 0, stopped_out, reason, total_fees, entry_fee, exit_fee, gross_pnl,
                entry_order_type, exit_order_type, entry_fee_percent * 100, exit_fee_percent * 100
            ))
        else:
            # Fallback if reason column doesn't exist or reason not provided
            cursor.execute('''
            INSERT INTO trade_history (symbol, direction, entry_price, exit_price, position_size, 
                              pnl, pnl_percent, entry_time, exit_time, duration_minutes, 
                              stop_loss, stopped_out, total_fees, entry_fee, exit_fee, gross_pnl, 
                              order_type, exit_order_type, entry_fee_percent, exit_fee_percent)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                symbol, direction, entry_price, exit_price, position_size,
                net_pnl, pnl_percent, entry_time, exit_dt.isoformat(), duration_minutes,
                stop_loss or 0, stopped_out, total_fees, entry_fee, exit_fee, gross_pnl,
                entry_order_type, exit_order_type, entry_fee_percent * 100, exit_fee_percent * 100
            ))
        
        # Get the newly inserted trade ID
        trade_id = cursor.lastrowid
        
        # Record capital update with net PnL (after fees)
        current_capital = get_current_capital()
        cursor.execute('''
        INSERT INTO capital_history (timestamp, capital, change, trade_id)
        VALUES (?, ?, ?, ?)
        ''', (
            exit_dt.isoformat(), current_capital, net_pnl, trade_id
        ))
        
        conn.commit()
        conn.close()
        
        close_reason = reason if reason else "Stop loss" if stopped_out else "Manual exit"
        logger.info(f"Trade recorded: {direction} {symbol}, Gross P&L: ${gross_pnl:.2f}, Fees: ${total_fees:.2f}, Net P&L: ${net_pnl:.2f}, P&L%: {pnl_percent:.2f}%, Reason: {close_reason}")
        
        # Log the event for audit trail if it's from reconciliation
        if reason and "reconciliation" in str(reason).lower():
            log_reconciliation_event(
                "TRADE_RECORDED", 
                symbol, 
                f"{direction.upper()} position closed. Gross P&L: ${gross_pnl:.2f}, Fees: ${total_fees:.2f}, Net P&L: ${net_pnl:.2f}, Reason: {close_reason}",
                {
                    "direction": direction,
                    "entry_price": entry_price,
                    "exit_price": exit_price,
                    "position_size": position_size,
                    "gross_pnl": gross_pnl,
                    "fees": total_fees,
                    "net_pnl": net_pnl,
                    "reason": close_reason,
                    "order_type": entry_order_type,
                    "exit_order_type": exit_order_type
                }
            )
        
        # Get post-trade balance (current balance after the trade is recorded)
        post_trade_balance = get_current_capital()
        
        # After recording to database, also log to Excel if enabled
        if excel_logger is not None:
            try:
                # Create trade data dict for Excel logging
                trade_data = {
                    'symbol': symbol,
                    'direction': direction,
                    'entry_price': entry_price,
                    'exit_price': exit_price,
                    'position_size': position_size,
                    'entry_time': entry_time,
                    'exit_time': exit_dt.isoformat(),
                    'stop_loss': stop_loss or 0,
                    'stopped_out': stopped_out,
                    'reason': reason,
                    'total_fees': total_fees,
                    'entry_fee': entry_fee,
                    'exit_fee': exit_fee,
                    'gross_pnl': gross_pnl,
                    'net_pnl': net_pnl,
                    'account_balance': get_position_account_balance(symbol, entry_time),
                    'post_trade_balance': post_trade_balance,
                    'order_type': entry_order_type,
                    'exit_order_type': exit_order_type,
                    'entry_fee_percent': entry_fee_percent * 100,
                    'exit_fee_percent': exit_fee_percent * 100
                }
                
                # Log to Excel in a separate thread to avoid blocking
                threading.Thread(target=excel_logger.log_trade, args=(trade_data,)).start()
                
            except Exception as excel_error:
                logger.error(f"Error logging trade to Excel: {str(excel_error)}")
                # Continue execution even if Excel logging fails
        
        # Send exit notification with integrated risk calculation
        if config.getboolean('NOTIFICATIONS', 'notify_exits', fallback=True):
            try:
                # Create emojis/badges for order types
                entry_order_badge = f"[Limit]" if entry_order_type == "Limit" else ""
                exit_order_badge = f"[Limit]" if exit_order_type == "Limit" else ""
                
                # Create emoji based on direction
                direction_emoji = "🟢" if direction == 'long' else "🔴"
                
                # Get account balance for notification
                account_balance = get_current_capital()
                
                # Calculate R multiple using total risk
                r_multiple = net_pnl / risk_amount if risk_amount > 0 else 0
                
                # Log for verification
                logger.info(f"Exit notification - Risk: Price ${price_risk:.2f}, Fees ${fee_risk:.2f}, Total ${risk_amount:.2f}, R multiple: {r_multiple:.2f}")
                
                # Send properly formatted notification with integrated risk
                send_telegram_message(f"{direction_emoji} <b>{direction.upper()} CLOSED</b> {exit_order_badge}\n"
                                    f"Symbol: {symbol}\n"
                                    f"Size: {position_size}\n"
                                    f"Risk $: ${risk_amount:.2f} (Price: ${price_risk:.2f}, Fees: ${fee_risk:.2f})\n"
                                    f"Entry: {entry_price} {entry_order_badge}\n"
                                    f"Exit: {exit_price} {exit_order_badge}\n"
                                    f"Gross P&L: ${gross_pnl:.2f}\n"
                                    f"Fees: ${total_fees:.2f} (Entry: ${entry_fee:.2f}, Exit: ${exit_fee:.2f})\n"
                                    f"Net P&L: ${net_pnl:.2f}\n"
                                    f"R multiple: {r_multiple:.2f}\n"
                                    f"Account balance: ${account_balance:.2f}\n"
                                    f"Reason: {close_reason}")
            except Exception as notification_error:
                logger.error(f"Error sending exit notification: {str(notification_error)}")
        
        return True
    except Exception as e:
        logger.error(f"Error recording trade: {str(e)}")
        return False





def get_performance_metrics():
    """Get trading performance metrics"""
    try:
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        metrics = {}
        
        # Total trades
        cursor.execute('SELECT COUNT(*) FROM trade_history')
        metrics['total_trades'] = cursor.fetchone()[0]
        
        if metrics['total_trades'] == 0:
            conn.close()
            return {
                'total_trades': 0,
                'winning_trades': 0,
                'win_rate': 0,
                'avg_profit': 0,
                'avg_loss': 0,
                'profit_factor': 0,
                'max_drawdown': 0,
                'avg_trade_duration': 0
            }
        
        # Winning trades
        cursor.execute('SELECT COUNT(*) FROM trade_history WHERE pnl > 0')
        metrics['winning_trades'] = cursor.fetchone()[0]
        
        # Win rate
        metrics['win_rate'] = (metrics['winning_trades'] / metrics['total_trades'] * 100) if metrics['total_trades'] > 0 else 0
        
        # Average profit, average loss
        cursor.execute('SELECT AVG(pnl) FROM trade_history WHERE pnl > 0')
        metrics['avg_profit'] = cursor.fetchone()[0] or 0
        
        cursor.execute('SELECT AVG(pnl) FROM trade_history WHERE pnl < 0')
        metrics['avg_loss'] = cursor.fetchone()[0] or 0
        
        # Profit factor
        cursor.execute('SELECT SUM(pnl) FROM trade_history WHERE pnl > 0')
        total_profit = cursor.fetchone()[0] or 0
        
        cursor.execute('SELECT SUM(pnl) FROM trade_history WHERE pnl < 0')
        total_loss = abs(cursor.fetchone()[0] or 0)
        
        metrics['profit_factor'] = total_profit / total_loss if total_loss > 0 else float('inf')
        
        # Average trade duration
        cursor.execute('SELECT AVG(duration_minutes) FROM trade_history')
        metrics['avg_trade_duration'] = cursor.fetchone()[0] or 0
        
        # Maximum drawdown calculation
        cursor.execute('SELECT capital, change FROM capital_history ORDER BY id')
        capital_history = cursor.fetchall()
        
        if capital_history:
            peak_capital = capital_history[0][0]
            max_drawdown = 0
            for capital, change in capital_history:
                if capital > peak_capital:
                    peak_capital = capital
                drawdown = peak_capital - capital
                if drawdown > max_drawdown:
                    max_drawdown = drawdown
            
            metrics['max_drawdown'] = max_drawdown
            if peak_capital > 0:
                metrics['max_drawdown_percent'] = (max_drawdown / peak_capital) * 100
            else:
                metrics['max_drawdown_percent'] = 0
        else:
            metrics['max_drawdown'] = 0
            metrics['max_drawdown_percent'] = 0
        
        # Add reason-based metrics if the column exists
        cursor.execute("PRAGMA table_info(trade_history)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'reason' in columns:
            # Get trade counts by reason
            cursor.execute('SELECT reason, COUNT(*) FROM trade_history GROUP BY reason')
            reason_counts = cursor.fetchall()
            
            for reason, count in reason_counts:
                if reason:
                    metrics[f'trades_by_{reason.replace(" ", "_").lower()}'] = count
        
        # Add order type metrics if the column exists
        if 'order_type' in columns:
            # Get trade counts by order type
            cursor.execute('SELECT order_type, COUNT(*) FROM trade_history GROUP BY order_type')
            order_type_counts = cursor.fetchall()
            
            for order_type, count in order_type_counts:
                if order_type:
                    metrics[f'trades_by_{order_type.lower()}_order'] = count
                    
            # Get win rate by order type
            cursor.execute('SELECT order_type, COUNT(*) FROM trade_history WHERE pnl > 0 GROUP BY order_type')
            order_type_win_counts = dict(cursor.fetchall())
            
            for order_type, count in order_type_counts:
                if order_type:
                    win_count = order_type_win_counts.get(order_type, 0)
                    win_rate = (win_count / count * 100) if count > 0 else 0
                    metrics[f'{order_type.lower()}_order_win_rate'] = win_rate
        
        conn.close()
        return metrics
    except Exception as e:
        logger.error(f"Error calculating performance metrics: {str(e)}")
        return {}

# IMPROVEMENT: User Notifications via Telegram
def send_telegram_message(message):
    """Send message via Telegram bot"""
    try:
        telegram_token = config.get('NOTIFICATIONS', 'telegram_token', fallback=None)
        telegram_chat_id = config.get('NOTIFICATIONS', 'telegram_chat_id', fallback=None)
        
        if not telegram_token or not telegram_chat_id:
            logger.warning("Telegram notification credentials not configured")
            return False
            
        url = f"https://api.telegram.org/bot{telegram_token}/sendMessage"
        data = {
            "chat_id": telegram_chat_id,
            "text": message,
            "parse_mode": "HTML"
        }
        
        response = requests.post(url, data=data)
        if response.status_code == 200:
            logger.info(f"Telegram notification sent: {message}")
            return True
        else:
            logger.error(f"Failed to send Telegram notification: {response.text}")
            return False
    except Exception as e:
        logger.error(f"Error sending Telegram notification: {str(e)}")
        return False

# Initialize Bybit client with global reuse
def initialize_bybit_client(force_new=False):
    """Initialize exchange client with option to reuse existing client.
    Supports both Bybit and Hyperliquid via bridge."""
    global global_client
    
    # Return existing client if we have one and not forcing new
    if global_client is not None and not force_new:
        return global_client
    
    # Determine which exchange to use
    active_exchange = config.get('EXCHANGE', 'active_exchange', fallback='bybit').lower()
    
    if active_exchange == 'hyperliquid':
        # Use Hyperliquid via bridge
        if not HYPERLIQUID_AVAILABLE:
            logger.error("Hyperliquid selected but bridge not available. "
                        "Install: pip install hyperliquid-python-sdk eth-account")
            raise ImportError("Hyperliquid bridge not available")
        
        logger.info("Initializing Hyperliquid client via bridge...")
        
        try:
            global_client = get_hyperliquid_client()
            logger.info("Successfully connected to Hyperliquid via bridge")
            return global_client
        except Exception as e:
            logger.error(f"Failed to initialize Hyperliquid client: {e}")
            raise
    
    else:
        # Use original Bybit client
        logger.info("Initializing Bybit client...")
        
        api_key = config['API'].get('api_key') or os.getenv('BYBIT_API_KEY')
        api_secret = config['API'].get('api_secret') or os.getenv('BYBIT_API_SECRET')
        testnet = config['API'].getboolean('testnet', False)
        
        if not api_key or not api_secret:
            logger.error("API credentials missing. Set in config.ini or environment variables")
            raise ValueError("API credentials not configured")
        
        global_client = pybit.unified_trading.HTTP(
            testnet=testnet,
            api_key=api_key,
            api_secret=api_secret
        )
        
        logger.info(f"Successfully connected to Bybit {'Testnet' if testnet else 'Mainnet'}")
        return global_client

def get_trading_fees():
    """Get trading fees for the active exchange"""
    active_exchange = config.get('EXCHANGE', 'active_exchange', fallback='bybit').lower()
    
    if active_exchange == 'hyperliquid':
        taker_fee = float(config.get('HYPERLIQUID', 'trading_fee_percent', fallback='0.0432'))
        maker_fee = float(config.get('HYPERLIQUID', 'maker_fee_percent', fallback='0.0144'))
    else:
        taker_fee = float(config.get('TRADING', 'trading_fee_percent', fallback='0.055'))
        maker_fee = float(config.get('TRADING', 'maker_fee_percent', fallback='0.02'))
    
    return taker_fee, maker_fee

# Reset the global client (called on critical errors)
def reset_global_client():
    """Reset the global client, forcing a new initialization on next call"""
    global global_client
    global_client = None
    logger.info("Global Bybit client reset")

# IMPROVEMENT: Retry Logic for API Calls
@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=1, max=10),
    retry=retry_if_exception_type((ConnectionError, TimeoutError)),
    before_sleep=lambda retry_state: logger.warning(f"Retrying API call (attempt {retry_state.attempt_number}) after error")
)
def execute_trade_with_retry(client, action, symbol, entry_price, stop_loss=None, position_size=None, take_profit=None, use_limit_order=None, use_tp_limit_order=None, closeOnTrigger=False):
    """Wrapper function with retry logic for execute_trade"""
    try:
        global api_calls
        api_calls += 1
        result = execute_trade(client, action, symbol, entry_price, stop_loss, position_size, take_profit, use_limit_order, use_tp_limit_order, closeOnTrigger)
        return result
    except Exception as e:
        global api_errors
        api_errors += 1
        logger.error(f"Error in execute_trade_with_retry: {str(e)}")
        raise

@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=1, max=10),
    retry=retry_if_exception_type((ConnectionError, TimeoutError)),
    before_sleep=lambda retry_state: logger.warning(f"Retrying get_symbol_info (attempt {retry_state.attempt_number}) after error")
)
def get_symbol_info_with_retry(client, symbol):
    """Wrapper function with retry logic for get_symbol_info"""
    try:
        global api_calls
        api_calls += 1
        result = get_symbol_info(client, symbol)
        return result
    except Exception as e:
        global api_errors
        api_errors += 1
        logger.error(f"Error in get_symbol_info_with_retry: {str(e)}")
        raise

# Add new function with retry logic for setting leverage
@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=1, max=10),
    retry=retry_if_exception_type((ConnectionError, TimeoutError)),
    before_sleep=lambda retry_state: logger.warning(f"Retrying set_leverage (attempt {retry_state.attempt_number}) after error")
)
def set_appropriate_leverage(client, symbol, entry_price, position_size):
    """Calculate and set the appropriate leverage based on position size and capital"""
    global api_calls
    api_calls += 1
    
    try:
        # Get current capital and apply same buffer as position sizing
        current_capital = get_current_capital()
        balance_buffer_percent = float(config['TRADING'].get('balance_buffer_percent', 5)) / 100
        buffer_amount = current_capital * balance_buffer_percent
        available_capital = current_capital - buffer_amount  # Use available capital!
        
        # Calculate leverage using available_capital
        required_margin = entry_price * position_size
        required_leverage = math.ceil(required_margin / available_capital * 1.05)
        
        # Get symbol info to check max allowed leverage
        symbol_info = get_symbol_info_with_retry(client, symbol)
        if not symbol_info:
            logger.error(f"Cannot set leverage: Symbol information not found for {symbol}")
            return False
        
        # Get max allowed leverage for this symbol
        leverage_filter = symbol_info.get('leverageFilter', {})
        max_leverage = float(leverage_filter.get('maxLeverage', 100))
        
        # Ensure leverage is within bounds
        final_leverage = min(required_leverage, max_leverage)
        final_leverage = max(final_leverage, 1)  # Minimum leverage is 1x
        
        logger.info(f"Setting leverage for {symbol}: Required: {required_leverage}x, Max allowed: {max_leverage}x, Final: {final_leverage}x")
        
        # Set the leverage using Bybit API
        response = client.set_leverage(
            category="linear",
            symbol=symbol,
            buyLeverage=str(final_leverage),
            sellLeverage=str(final_leverage)
        )
        
        if response.get('retCode') == 0:
            logger.info(f"Successfully set leverage to {final_leverage}x for {symbol}")
            return True
        else:
            # Special handling for "leverage not modified" (10001) which is OK
            if response.get('retCode') == 10001 and "leverage not modified" in str(response.get('retMsg', '')).lower():
                logger.info(f"Leverage already set to {final_leverage}x for {symbol}")
                return True
                
            logger.error(f"Failed to set leverage: {response}")
            return False
            
    except Exception as e:
        global api_errors
        api_errors += 1
        logger.error(f"Error setting leverage: {str(e)}")
        return False

# Add new function with retry logic for fetching account balance
@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=1, max=10),
    retry=retry_if_exception_type((ConnectionError, TimeoutError)),
    before_sleep=lambda retry_state: logger.warning(f"Retrying get balance (attempt {retry_state.attempt_number}) after error")
)
def get_bybit_balance():
    """Get balance from Bybit with retry logic"""
    global api_calls
    api_calls += 1
    client = initialize_bybit_client()  # Will reuse existing client
    account_info = client.get_wallet_balance(accountType="UNIFIED")
    if account_info.get('retCode') == 0:
        balance = float(account_info.get('result', {}).get('list', [{}])[0].get('totalEquity', 0))
        logger.info(f"Retrieved account balance from Bybit: ${balance}")
        return balance
    else:
        logger.error(f"Failed to get account information: {account_info}")
        # Consider resetting the client on severe errors
        if account_info.get('retCode') in [10002, 10003, 10004]:  # Auth error codes
            reset_global_client()
        raise ConnectionError(f"Failed to get account information: {account_info}")

# Updated function to get current capital directly from Bybit
def get_current_capital():
    """Get the current capital value directly from Bybit account"""
    try:
        # If using testnet, just return the tracked capital
        if config['API'].getboolean('testnet', False):
            # Read from a file to persist capital between bot restarts
            capital_file = 'capital.txt'
            if os.path.exists(capital_file):
                with open(capital_file, 'r') as f:
                    return float(f.read().strip())
            else:
                # Use initial capital from config if file doesn't exist
                initial_capital = float(config['TRADING'].get('initial_capital', 10000))
                with open(capital_file, 'w') as f:
                    f.write(str(initial_capital))
                return initial_capital
        else:
            # For live trading, always get actual balance from Bybit
            try:
                return get_bybit_balance()
            except Exception as api_error:
                logger.error(f"Error getting balance from Bybit: {str(api_error)}")
                logger.warning("Falling back to local capital tracking")
                
                # Fall back to local tracking if API fails
                capital_file = 'capital.txt'
                if os.path.exists(capital_file):
                    with open(capital_file, 'r') as f:
                        return float(f.read().strip())
                else:
                    initial_capital = float(config['TRADING'].get('initial_capital', 10000))
                    with open(capital_file, 'w') as f:
                        f.write(str(initial_capital))
                    return initial_capital
    except Exception as e:
        logger.error(f"Error getting current capital: {str(e)}")
        # Fallback to initial capital
        return float(config['TRADING'].get('initial_capital', 10000))

# Updated function to update local capital tracker
def update_capital(pnl):
    """Update the local capital tracker after a trade with the P&L"""
    try:
        # Get current capital from Bybit for logging purposes
        current = get_current_capital()
        new_capital = current + pnl
        
        # Only update the local file as a backup, don't rely on it for primary capital tracking
        with open('capital.txt', 'w') as f:
            f.write(str(new_capital))
        
        logger.info(f"Trade P&L: ${pnl}, New estimated capital: ${new_capital}")
        return new_capital
    except Exception as e:
        logger.error(f"Error updating capital: {str(e)}")
        return get_current_capital()






def update_balance_history():
    """Save current balance snapshot to database"""
    try:
        current_balance = get_bybit_balance()
        
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Check if we already have a recent entry (within last few minutes)
        cutoff_time = (datetime.now() - timedelta(minutes=5)).isoformat()
        cursor.execute('''
        SELECT COUNT(*) FROM balance_history
        WHERE timestamp > ?
        ''', (cutoff_time,))
        
        recent_count = cursor.fetchone()[0]
        
        # Only insert if no recent entry
        if recent_count == 0:
            cursor.execute('''
            INSERT INTO balance_history (timestamp, balance, source)
            VALUES (?, ?, ?)
            ''', (
                datetime.now().isoformat(),
                current_balance,
                'reconciliation'
            ))
            
            # NEW: Also insert a "reference_point" record every hour to ensure we have good reference points
            # Check if we already have a reference point in the last hour
            hourly_cutoff = (datetime.now() - timedelta(hours=1)).isoformat()
            cursor.execute('''
            SELECT COUNT(*) FROM balance_history
            WHERE timestamp > ? AND source = 'reference_point'
            ''', (hourly_cutoff,))
            
            hourly_count = cursor.fetchone()[0]
            
            if hourly_count == 0:
                # Insert a special reference point record that will be used for max loss calculations
                cursor.execute('''
                INSERT INTO balance_history (timestamp, balance, source)
                VALUES (?, ?, ?)
                ''', (
                    datetime.now().isoformat(),
                    current_balance,
                    'reference_point'
                ))
                logger.info(f"Hourly reference point recorded: ${current_balance}")
            
            conn.commit()
            logger.info(f"Balance snapshot recorded: ${current_balance}")
        
        conn.close()
        return current_balance
    except Exception as e:
        logger.error(f"Error updating balance history: {str(e)}")
        # If error, try to get current capital using fallback method
        return get_current_capital()

def check_max_daily_loss():
    """Check if max daily loss has been exceeded"""
    try:
        # Get configuration
        max_loss_percent = float(config['RISK_MANAGEMENT'].get('max_daily_loss_percent', 2.0))
        period_hours = int(config['RISK_MANAGEMENT'].get('max_daily_loss_period_hours', 24))
        
        # Get current balance
        current_balance = get_bybit_balance()
        
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Calculate the current reference time (24 hours ago)
        reference_time = datetime.now() - timedelta(hours=period_hours)
        
        # FIRST: Try to find a recent manual reset record (less than 24 hours old)
        cursor.execute('''
        SELECT balance, timestamp FROM balance_history
        WHERE source = 'manual_reset' AND timestamp > ?
        ORDER BY timestamp DESC LIMIT 1
        ''', (reference_time.isoformat(),))
        
        reset_record = cursor.fetchone()
        
        if reset_record:
            # If we found a recent manual reset record, use it
            reference_balance = reset_record[0]
            reset_timestamp = reset_record[1]
            logger.info(f"Using manual reset reference balance from {reset_timestamp}")
        else:
            # Otherwise, use standard 24-hour window logic with improvements
            
            # IMPROVED: First look for reference_point records close to reference time
            # Use a time window of +/- 2 hours around reference time for flexibility
            window_start = (reference_time - timedelta(hours=2)).isoformat()
            window_end = (reference_time + timedelta(hours=2)).isoformat()
            
            cursor.execute('''
            SELECT balance, timestamp FROM balance_history
            WHERE timestamp BETWEEN ? AND ? AND (source = 'reference_point' OR source = 'reconciliation')
            ORDER BY ABS(JULIANDAY(timestamp) - JULIANDAY(?)) ASC
            LIMIT 1
            ''', (window_start, window_end, reference_time.isoformat()))
            
            result = cursor.fetchone()
            
            # If no reference found in window, try finding the closest record before reference time
            if not result:
                cursor.execute('''
                SELECT balance, timestamp FROM balance_history
                WHERE timestamp <= ?
                ORDER BY timestamp DESC LIMIT 1
                ''', (reference_time.isoformat(),))
                
                result = cursor.fetchone()
            
            # If still no reference, try finding highest balance in last period_hours*2
            if not result:
                cursor.execute('''
                SELECT MAX(balance), timestamp FROM balance_history
                WHERE timestamp >= ?
                ''', ((datetime.now() - timedelta(hours=period_hours*2)).isoformat(),))
                
                result = cursor.fetchone()
            
            if not result or not result[0]:
                logger.warning(f"No balance history found for {period_hours}h ago, using initial capital")
                reference_balance = float(config['TRADING'].get('initial_capital', 10000))
            else:
                reference_balance = result[0]
                ref_timestamp = result[1] if len(result) > 1 and result[1] else "unknown time"
                logger.info(f"Using standard reference balance from {ref_timestamp}")
        
        conn.close()
        
        # Calculate loss percentage
        percent_change = ((current_balance - reference_balance) / reference_balance) * 100
        loss_percent = -percent_change if percent_change < 0 else 0
        
        logger.info(f"Loss check: Current: ${current_balance}, Reference: ${reference_balance}, Loss: {loss_percent:.2f}%")
        
        # Return if threshold exceeded
        return loss_percent >= max_loss_percent, loss_percent, current_balance, reference_balance
    except Exception as e:
        logger.error(f"Error checking max daily loss: {str(e)}")
        return False, 0, 0, 0





def close_all_positions_max_loss():
    """Close all active positions due to max daily loss trigger"""
    try:
        client = initialize_bybit_client()
        positions_closed = 0
        total_pnl = 0
        
        # Get all active positions from both local tracking and direct API
        local_positions = list(active_positions.items())
        
        # Also get positions directly from Bybit to ensure we catch everything
        try:
            response = client.get_positions(
                category="linear",
                settleCoin="USDT"
            )
            
            if response.get('retCode') == 0:
                positions_list = response.get('result', {}).get('list', [])
                
                # Add positions from API that aren't in local tracking
                for pos in positions_list:
                    symbol = pos.get('symbol')
                    if float(pos.get('size', 0)) > 0 and symbol not in active_positions:
                        # Create a local representation of this position
                        direction = 'long' if pos.get('side') == 'Buy' else 'short'
                        logger.warning(f"Found untracked {direction} position for {symbol} during max loss closure")
                        local_positions.append((
                            symbol, 
                            {
                                'direction': direction,
                                'position_size': float(pos.get('size', 0)),
                                'entry_price': float(pos.get('avgPrice', 0))
                            }
                        ))
        except Exception as e:
            logger.error(f"Error getting positions directly from Bybit during max loss closure: {str(e)}")
            # Continue with locally tracked positions
        
        # Close each position
        for symbol, position in local_positions:
            direction = position['direction']
            exit_action = "EXIT LONG" if direction == 'long' else "EXIT SHORT"
            
            # Get current price
            try:
                ticker_response = client.get_tickers(category="linear", symbol=symbol)
                if ticker_response.get('retCode') == 0:
                    current_price = float(ticker_response.get('result', {}).get('list', [{}])[0].get('lastPrice', 0))
                    
                    # Execute exit trade
                    result = execute_trade_with_retry(client, exit_action, symbol, current_price, closeOnTrigger=True)
                    
                    if result:
                        logger.info(f"Max daily loss: Closed {direction} position for {symbol}")
                        positions_closed += 1
                        
                        # Calculate realized PnL
                        entry_price = position['entry_price']
                        position_size = position['position_size']
                        
                        if direction == 'long':
                            pnl = (current_price - entry_price) * position_size
                        else:
                            pnl = (entry_price - current_price) * position_size
                        
                        total_pnl += pnl
                    else:
                        logger.error(f"Failed to close {direction} position for {symbol} during max loss closure")
                else:
                    logger.error(f"Failed to get ticker for {symbol} during max loss closure")
            except Exception as e:
                logger.error(f"Error closing position for {symbol}: {str(e)}")
        
        return positions_closed, total_pnl
    except Exception as e:
        logger.error(f"Error in close_all_positions_max_loss: {str(e)}")
        return 0, 0


def close_losing_positions():
    """Close only positions with negative unrealized PnL"""
    try:
        client = initialize_bybit_client()
        positions_closed = 0
        total_pnl = 0
        
        # Get all positions from Bybit API to get unrealized PnL
        try:
            response = client.get_positions(category="linear")
            if response.get('retCode') == 0:
                positions_list = response.get('result', {}).get('list', [])
                
                # Filter to losing positions
                losing_positions = []
                for pos in positions_list:
                    symbol = pos.get('symbol')
                    unrealized_pnl = float(pos.get('unrealisedPnl', 0))
                    
                    if float(pos.get('size', 0)) > 0 and unrealized_pnl < 0:
                        direction = 'long' if pos.get('side') == 'Buy' else 'short'
                        losing_positions.append({
                            'symbol': symbol,
                            'direction': direction,
                            'position_size': float(pos.get('size', 0)),
                            'entry_price': float(pos.get('avgPrice', 0)),
                            'unrealized_pnl': unrealized_pnl
                        })
                
                # Sort by largest losers first
                losing_positions.sort(key=lambda x: x['unrealized_pnl'])
                
                # Close each losing position
                for position in losing_positions:
                    symbol = position['symbol']
                    direction = position['direction']
                    exit_action = "EXIT LONG" if direction == 'long' else "EXIT SHORT"
                    
                    # Get current price
                    ticker_response = client.get_tickers(category="linear", symbol=symbol)
                    if ticker_response.get('retCode') == 0:
                        current_price = float(ticker_response.get('result', {}).get('list', [{}])[0].get('lastPrice', 0))
                        
                        # Execute exit trade
                        result = execute_trade_with_retry(client, exit_action, symbol, current_price, closeOnTrigger=True)
                        
                        if result:
                            logger.info(f"Closed losing {direction} position for {symbol}, Unrealized PnL: {position['unrealized_pnl']}")
                            positions_closed += 1
                            total_pnl += position['unrealized_pnl']
                        else:
                            logger.error(f"Failed to close losing {direction} position for {symbol}")
                
                logger.info(f"Closed {positions_closed} losing positions, Total PnL: {total_pnl}")
                
                # Send notification
                if positions_closed > 0 and config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                    send_telegram_message(f"🔴 <b>LOSING POSITIONS CLOSED</b>\n"
                                        f"Positions closed: {positions_closed}\n"
                                        f"Total realized PnL: ${total_pnl:.2f}")
            else:
                logger.error(f"Failed to get positions from Bybit: {response}")
        except Exception as e:
            logger.error(f"Error getting positions from Bybit: {str(e)}")
        
        return positions_closed, total_pnl
    except Exception as e:
        logger.error(f"Error in close_losing_positions: {str(e)}")
        return 0, 0


def adjust_position_size_for_risk_level(original_size):
    """Adjust position size based on current risk level"""
    try:
        # Skip if feature is disabled
        if not config['RISK_MANAGEMENT'].getboolean('max_daily_loss_enabled', True):
            return original_size
        
        # Skip if tiered response is disabled
        if not config['RISK_MANAGEMENT'].getboolean('use_tiered_response', True):
            return original_size
        
        # Get current loss percentage
        max_loss_percent = float(config['RISK_MANAGEMENT'].get('max_daily_loss_percent', 2.0))
        _, loss_percent, _, _ = check_max_daily_loss()
        
        # Define tier thresholds
        tier1_threshold = max_loss_percent * 0.5  # 50% of max loss
        tier2_threshold = max_loss_percent * 0.75  # 75% of max loss
        
        # Apply adjustments based on tiers
        if loss_percent < tier1_threshold:
            # No adjustment needed
            return original_size
        elif loss_percent >= tier1_threshold and loss_percent < tier2_threshold:
            # Tier 1: Reduce by 25%
            adjusted_size = original_size * 0.75
            logger.info(f"Tier 1 risk adjustment: Position size reduced by 25% ({original_size} → {adjusted_size})")
            return adjusted_size
        elif loss_percent >= tier2_threshold:
            # Tier 2: Reduce by 50%
            adjusted_size = original_size * 0.5
            logger.info(f"Tier 2 risk adjustment: Position size reduced by 50% ({original_size} → {adjusted_size})")
            return adjusted_size
    except Exception as e:
        logger.error(f"Error adjusting position size for risk level: {str(e)}")
        # In case of error, return original size
        return original_size




def handle_max_daily_loss():
    """Main function to manage max daily loss checks and actions"""
    global last_max_loss_notification_time, last_warning_notification_time, max_loss_first_triggered_time, skipped_signals_count
    
    try:
        # Skip if feature is disabled
        if not config['RISK_MANAGEMENT'].getboolean('max_daily_loss_enabled', True):
            return False
        
        # Get configurations
        max_loss_percent = float(config['RISK_MANAGEMENT'].get('max_daily_loss_percent', 2.0))
        warning_percent = float(config['RISK_MANAGEMENT'].get('max_daily_loss_warning_percent', 1.5))
        period_hours = int(config['RISK_MANAGEMENT'].get('max_daily_loss_period_hours', 24))
        use_tiered_response = config['RISK_MANAGEMENT'].getboolean('use_tiered_response', True)
        
        # Check if max loss has been exceeded
        exceeded, loss_percent, current_balance, reference_balance = check_max_daily_loss()
        
        if exceeded:
            # Close positions only on first detection or if positions exist
            positions_to_close = False
            
            # If this is first trigger or we've never triggered before, close positions
            if max_loss_first_triggered_time is None:
                max_loss_first_triggered_time = datetime.now()
                positions_to_close = True
                
            logger.warning(f"Max daily loss of {loss_percent:.2f}% exceeded (threshold: {max_loss_percent}%). Taking action.")
            
            # Close all positions if needed
            positions_closed = 0
            total_pnl = 0
            if positions_to_close:
                positions_closed, total_pnl = close_all_positions_max_loss()
            
                # Record the max loss event
                conn = sqlite3.connect('trading_bot.db')
                cursor = conn.cursor()
                cursor.execute('''
                INSERT INTO max_loss_events 
                (timestamp, current_balance, reference_balance, loss_percent, positions_closed, total_pnl)
                VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    datetime.now().isoformat(),
                    current_balance,
                    reference_balance,
                    loss_percent,
                    positions_closed,
                    total_pnl
                ))
                conn.commit()
                conn.close()
            
            # Determine if we should send notification based on time passed
            current_time = datetime.now()
            should_notify = False
            
            # Send notification immediately when first triggered
            if last_max_loss_notification_time is None:
                should_notify = True
            # Or if it's been more than 30 minutes since last notification
            elif (current_time - last_max_loss_notification_time).total_seconds() > 1800:  # 30 minutes
                should_notify = True
                
            # Send notification if conditions are met
            if should_notify and config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                # Calculate time since first trigger
                if max_loss_first_triggered_time:
                    time_since_trigger = current_time - max_loss_first_triggered_time
                    hours, remainder = divmod(time_since_trigger.total_seconds(), 3600)
                    minutes, seconds = divmod(remainder, 60)
                    trigger_duration = f"{int(hours)}h {int(minutes)}m ago"
                else:
                    trigger_duration = "just now"
                
                # Construct and send message
                send_telegram_message(f"🛑 <b>MAX DAILY LOSS TRIGGERED</b>\n"
                                    f"Loss: {loss_percent:.2f}% (Threshold: {max_loss_percent}%)\n"
                                    f"Current balance: ${current_balance:.2f}\n"
                                    f"Reference balance (from {period_hours}h ago): ${reference_balance:.2f}\n" +
                                    (f"Positions closed: {positions_closed}\n"
                                    f"Total realized PnL: ${total_pnl:.2f}\n" if positions_to_close else "") +
                                    f"First triggered: {trigger_duration}\n"
                                    f"Signals skipped: {skipped_signals_count}")
                
                # Update last notification time
                last_max_loss_notification_time = current_time
            
            return True
        else:
            # Reset max loss triggered time if loss is no longer exceeding threshold
            max_loss_first_triggered_time = None
            skipped_signals_count = 0
            last_max_loss_notification_time = None
        
        # If using tiered response, check tier thresholds
        if use_tiered_response:
            tier2_threshold = max_loss_percent * 0.75  # 75% of max loss
            
            # Tier 2: Close losing positions if loss is between 75% and 100% of max_loss_percent
            if loss_percent >= tier2_threshold and loss_percent < max_loss_percent:
                logger.warning(f"Tier 2 risk level reached: {loss_percent:.2f}% / {max_loss_percent}%. Closing losing positions.")
                close_losing_positions()
        
        # Check if approaching warning threshold
        if loss_percent >= warning_percent and loss_percent < max_loss_percent:
            logger.warning(f"Approaching max daily loss: {loss_percent:.2f}% / {max_loss_percent}%")
            
            # Determine if warning notification should be sent
            current_time = datetime.now()
            should_notify = False
            
            # Send notification for first warning or every 15 minutes
            if last_warning_notification_time is None:
                should_notify = True
            elif (current_time - last_warning_notification_time).total_seconds() > 900:  # 15 minutes
                should_notify = True
            
            # Send warning notification if needed
            if should_notify and config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                send_telegram_message(f"⚠️ <b>APPROACHING MAX DAILY LOSS</b>\n"
                                    f"Current loss: {loss_percent:.2f}%\n"
                                    f"Warning threshold: {warning_percent}%\n"
                                    f"Max threshold: {max_loss_percent}%\n"
                                    f"Current balance: ${current_balance:.2f}")
                
                # Update last warning time
                last_warning_notification_time = current_time
        else:
            # Reset warning notification time if not approaching
            last_warning_notification_time = None
        
        return False
    except Exception as e:
        logger.error(f"Error in handle_max_daily_loss: {str(e)}")
        return False



# Calculate position size based on risk management
def calculate_position_size(entry_price, stop_loss_price, symbol_info, custom_risk_percentage=None, use_limit_entry=None):
    """
    Calculate position size based on risk management with integrated fee calculation
    
    Args:
        entry_price (float): Entry price
        stop_loss_price (float): Stop loss price
        symbol_info (dict): Symbol information from exchange
        custom_risk_percentage (float, optional): Custom risk percentage (0-1). Defaults to None.
        use_limit_entry (bool, optional): Whether to use limit order for entry. Defaults to None (use config).
        
    Returns:
        float: Calculated position size
    """
    try:
        # Use custom risk percentage if provided, otherwise use from config
        if custom_risk_percentage is not None:
            risk_percentage = custom_risk_percentage
        else:
            risk_percentage = float(config['TRADING'].get('risk_percentage', 5)) / 100
            
        max_leverage = float(config['TRADING'].get('max_leverage', 5))
        use_risk_only = config['TRADING'].getboolean('use_risk_only', False)
        
        # Get the new parameters
        taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100  # For market orders
        maker_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02)) / 100     # For limit orders
        balance_buffer_percent = float(config['TRADING'].get('balance_buffer_percent', 5)) / 100
        
        # Determine if we're using limit orders for entry - use parameter if provided, otherwise config
        if use_limit_entry is None:
            use_limit_entry = config['TRADING'].getboolean('use_limit_orders_entry', False)
        
        # Get the current capital (will be updated after each trade)
        current_capital = get_current_capital()
        logger.info(f"Current capital: ${current_capital}")
        
        # Apply buffer to available capital
        buffer_amount = current_capital * balance_buffer_percent
        available_capital = current_capital - buffer_amount
        logger.info(f"Buffer amount: ${buffer_amount} ({balance_buffer_percent*100}%), Available capital: ${available_capital}")
        
        # Calculate risk amount in dollars based on available capital
        risk_amount = available_capital * risk_percentage
        logger.info(f"Risk amount (${risk_percentage*100}% of available capital): ${risk_amount:.2f}")
        
        # Calculate price risk per unit (loss from price movement)
        price_risk_per_unit = abs(float(entry_price) - float(stop_loss_price))
        
        if price_risk_per_unit <= 0.0001:
            logger.warning("Risk per unit too small, cannot calculate position size")
            return 0
        
        # Determine which fee percentages to use based on order type
        entry_fee_percent = maker_fee_percent if use_limit_entry else taker_fee_percent
        exit_fee_percent = taker_fee_percent  # Stop loss is always a market order
        
        # Calculate fee cost per unit (using appropriate fee percentages for entry and exit)
        entry_fee_cost = entry_fee_percent * float(entry_price)
        exit_fee_cost = exit_fee_percent * float(stop_loss_price)
        fee_cost_per_unit = entry_fee_cost + exit_fee_cost
        
        # Total risk per unit includes both price movement and fees
        total_risk_per_unit = price_risk_per_unit + fee_cost_per_unit
        
        # Calculate position size based on total risk per unit
        position_size = risk_amount / total_risk_per_unit
        
        if not use_risk_only:
            # Consider max leverage based on available capital (after buffer)
            max_units_based_on_capital = (available_capital * max_leverage) / float(entry_price)
            position_size = min(position_size, max_units_based_on_capital)
        
        # Log detailed breakdown of risk components
        estimated_price_risk = position_size * price_risk_per_unit
        estimated_entry_fee = position_size * entry_fee_cost
        estimated_exit_fee = position_size * exit_fee_cost
        estimated_fee_cost = estimated_entry_fee + estimated_exit_fee
        total_estimated_risk = estimated_price_risk + estimated_fee_cost
        
        logger.info(f"Position size: {position_size}")
        logger.info(f"Price risk per unit: ${price_risk_per_unit}")
        logger.info(f"Entry fee per unit ({entry_fee_percent*100:.3f}% {'maker' if use_limit_entry else 'taker'} fee): ${entry_fee_cost}")
        logger.info(f"Exit fee per unit ({exit_fee_percent*100:.3f}% taker fee): ${exit_fee_cost}")
        logger.info(f"Total fee cost per unit: ${fee_cost_per_unit}")
        logger.info(f"Total risk per unit: ${total_risk_per_unit}")
        logger.info(f"Estimated price risk: ${estimated_price_risk:.2f}")
        logger.info(f"Estimated entry fee: ${estimated_entry_fee:.2f}")
        logger.info(f"Estimated exit fee: ${estimated_exit_fee:.2f}")
        logger.info(f"Estimated total fee cost: ${estimated_fee_cost:.2f}")
        logger.info(f"Total estimated risk: ${total_estimated_risk:.2f} ({(total_estimated_risk/available_capital)*100:.2f}% of available capital)")
        logger.info(f"Target risk: ${risk_amount:.2f}")
        
        # Apply risk-level based position size adjustment
        if config.get('RISK_MANAGEMENT', 'max_daily_loss_enabled', fallback='True').lower() == 'true':
            position_size = adjust_position_size_for_risk_level(position_size)
        
        # Apply lot size restrictions based on symbol
        min_qty = float(symbol_info.get('lotSizeFilter', {}).get('minOrderQty', symbol_info.get('minOrderQty', 0.001)))
        qty_step = float(symbol_info.get('lotSizeFilter', {}).get('qtyStep', 0.001))
        
        # Round down to the nearest step size
        position_size = (position_size // qty_step) * qty_step
        
        # Ensure minimum size
        if position_size < min_qty:
            position_size = 0
            logger.warning(f"Calculated position size {position_size} is less than minimum {min_qty}")
        
        logger.info(f"Final position size with integrated fee risk: {position_size}")
        return position_size
    
    except Exception as e:
        logger.error(f"Error calculating position size: {str(e)}")
        return 0


def calculate_position_size_fixed_risk(entry_price, stop_loss_price, symbol_info, fixed_risk_amount, use_limit_entry=None):
    """
    Calculate position size based on fixed dollar risk amount
    
    Args:
        entry_price (float): Entry price
        stop_loss_price (float): Stop loss price
        symbol_info (dict): Symbol information from exchange
        fixed_risk_amount (float): Fixed risk amount in dollars
        use_limit_entry (bool, optional): Whether to use limit order for entry. Defaults to None (use config).
        
    Returns:
        float: Calculated position size
    """
    try:
        # Get fee percentages
        taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100  # For market orders
        maker_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02)) / 100     # For limit orders
        
        # Determine if we're using limit orders for entry - use parameter if provided, otherwise config
        if use_limit_entry is None:
            use_limit_entry = config['TRADING'].getboolean('use_limit_orders_entry', False)
        
        # Calculate price risk per unit (loss from price movement)
        price_risk_per_unit = abs(float(entry_price) - float(stop_loss_price))
        
        if price_risk_per_unit <= 0.0001:
            logger.warning("Risk per unit too small, cannot calculate position size")
            return 0
        
        # Determine which fee percentages to use based on order type
        entry_fee_percent = maker_fee_percent if use_limit_entry else taker_fee_percent
        exit_fee_percent = taker_fee_percent  # Stop loss is always a market order
        
        # Calculate fee cost per unit (using appropriate fee percentages for entry and exit)
        entry_fee_cost = entry_fee_percent * float(entry_price)
        exit_fee_cost = exit_fee_percent * float(stop_loss_price)
        fee_cost_per_unit = entry_fee_cost + exit_fee_cost
        
        # Total risk per unit includes both price movement and fees
        total_risk_per_unit = price_risk_per_unit + fee_cost_per_unit
        
        # Calculate position size based on fixed risk amount
        position_size = fixed_risk_amount / total_risk_per_unit
        
        # Log detailed breakdown of risk components
        estimated_price_risk = position_size * price_risk_per_unit
        estimated_entry_fee = position_size * entry_fee_cost
        estimated_exit_fee = position_size * exit_fee_cost
        estimated_fee_cost = estimated_entry_fee + estimated_exit_fee
        total_estimated_risk = estimated_price_risk + estimated_fee_cost
        
        logger.info(f"[Fixed Risk] Position size: {position_size}")
        logger.info(f"[Fixed Risk] Fixed risk amount: ${fixed_risk_amount}")
        logger.info(f"[Fixed Risk] Price risk per unit: ${price_risk_per_unit}")
        logger.info(f"[Fixed Risk] Entry fee per unit ({entry_fee_percent*100:.3f}% {'maker' if use_limit_entry else 'taker'} fee): ${entry_fee_cost}")
        logger.info(f"[Fixed Risk] Exit fee per unit ({exit_fee_percent*100:.3f}% taker fee): ${exit_fee_cost}")
        logger.info(f"[Fixed Risk] Total fee cost per unit: ${fee_cost_per_unit}")
        logger.info(f"[Fixed Risk] Total risk per unit: ${total_risk_per_unit}")
        logger.info(f"[Fixed Risk] Estimated price risk: ${estimated_price_risk:.2f}")
        logger.info(f"[Fixed Risk] Estimated entry fee: ${estimated_entry_fee:.2f}")
        logger.info(f"[Fixed Risk] Estimated exit fee: ${estimated_exit_fee:.2f}")
        logger.info(f"[Fixed Risk] Estimated total fee cost: ${estimated_fee_cost:.2f}")
        logger.info(f"[Fixed Risk] Total estimated risk: ${total_estimated_risk:.2f}")
        
        # Apply risk-level based position size adjustment if needed
        if config.get('RISK_MANAGEMENT', 'max_daily_loss_enabled', fallback='True').lower() == 'true':
            position_size = adjust_position_size_for_risk_level(position_size)
        
        # Apply lot size restrictions based on symbol
        min_qty = float(symbol_info.get('lotSizeFilter', {}).get('minOrderQty', symbol_info.get('minOrderQty', 0.001)))
        qty_step = float(symbol_info.get('lotSizeFilter', {}).get('qtyStep', 0.001))
        
        # Round down to the nearest step size
        position_size = (position_size // qty_step) * qty_step
        
        # Ensure minimum size
        if position_size < min_qty:
            position_size = 0
            logger.warning(f"Calculated position size {position_size} is less than minimum {min_qty}")
        
        logger.info(f"[Fixed Risk] Final position size with integrated fee risk: {position_size}")
        return position_size
    
    except Exception as e:
        logger.error(f"Error calculating position size for fixed risk: {str(e)}")
        return 0



# Get symbol information from Bybit
def get_symbol_info(client, symbol):
    try:
        response = client.get_instruments_info(
            category="linear",
            symbol=symbol
        )
        
        if response.get('retCode') == 0:
            symbols = response.get('result', {}).get('list', [])
            for sym_info in symbols:
                if sym_info.get('symbol') == symbol:
                    return sym_info
        
        logger.error(f"Symbol information not found for {symbol}")
        return None
    except Exception as e:
        logger.error(f"Error getting symbol information: {str(e)}")
        return None



# qty format function
def format_position_quantity(qty, symbol_info):
    """
    Format position quantity according to exchange requirements for the specific symbol
    
    Args:
        qty (float): Raw quantity value
        symbol_info (dict): Symbol information from the exchange
        
    Returns:
        str: Properly formatted quantity string
    """
    try:
        # Get the quantity step size for this symbol
        qty_step = float(symbol_info.get('lotSizeFilter', {}).get('qtyStep', 0.001))
        
        # Calculate the number of decimal places needed
        # FIX: Handle scientific notation by calculating from the actual value
        if qty_step > 0 and qty_step < 1:
            # Use logarithm to find decimal places (handles scientific notation)
            decimal_places = max(0, -int(math.floor(math.log10(qty_step))))
        else:
            decimal_places = 0
            step_str = f"{qty_step:.10f}".rstrip('0')  # Force decimal notation
            if '.' in step_str:
                decimal_places = len(step_str.split('.')[1])
        
        # Round down to the nearest step size
        qty = math.floor(qty / qty_step) * qty_step
        
        # Format to the appropriate number of decimal places
        formatted_qty = f"{qty:.{decimal_places}f}"
        
        # Remove trailing zeros and decimal point if not needed
        if '.' in formatted_qty:
            formatted_qty = formatted_qty.rstrip('0').rstrip('.')
            if not formatted_qty:  # Handle the case where it's all zeros
                formatted_qty = '0'
                
        logger.info(f"Formatted position quantity from {qty} to {formatted_qty} with {decimal_places} decimal places")
        return formatted_qty
        
    except Exception as e:
        logger.error(f"Error formatting position quantity: {str(e)}")
        # Fallback to 5 decimal places (safe for most crypto)
        return f"{qty:.5f}"


# Function to price a limit order close to the current price
def format_price(price, symbol_info, is_buy=True, limit_buffer_percent=0.0):
    """
    Format price according to exchange requirements and add appropriate buffer for limit orders
    
    Args:
        price (float): Base price (usually current market price)
        symbol_info (dict): Symbol information from the exchange
        is_buy (bool): True if this is a buy order, False for sell order
        limit_buffer_percent (float): Percentage buffer for limit price (default: 0.0%)
        
    Returns:
        str: Properly formatted price string for limit orders
    """
    try:
        # Get the price step (tick size) for this symbol
        tick_size = float(symbol_info.get('priceFilter', {}).get('tickSize', 0.00001))
        
        # Calculate buffer amount based on direction
        buffer_amount = price * (limit_buffer_percent / 100)
        
        # Adjust price based on order side:
        # - Buy limit orders should be below current price (subtract buffer)
        # - Sell limit orders should be above current price (add buffer)
        if is_buy:
            adjusted_price = price - buffer_amount
        else:
            adjusted_price = price + buffer_amount
        
        # Round to the nearest tick size
        adjusted_price = round(adjusted_price / tick_size) * tick_size
        
        # Calculate the number of decimal places needed
        decimal_places = 0
        tick_str = str(tick_size)
        if '.' in tick_str:
            decimal_places = len(tick_str.split('.')[1].rstrip('0'))
            
        # For cryptocurrencies, ensure we have at least 5 decimal places for low-value coins
        if 'USDT' in symbol_info.get('symbol', '') and price < 1.0:
            decimal_places = max(decimal_places, 5)
        
        # Format to the appropriate number of decimal places
        formatted_price = f"{adjusted_price:.{decimal_places}f}"
        
        # Log the price formatting details
        logger.info(f"Formatted price from {price} to {formatted_price} with {decimal_places} decimal places ({'+' if not is_buy else '-'}{limit_buffer_percent}% buffer)")
        return formatted_price
        
    except Exception as e:
        logger.error(f"Error formatting price: {str(e)}")
        # For USDT pairs, use more decimal places as fallback
        if price and str(price).replace('.', '').isdigit():
            symbol = symbol_info.get('symbol', '')
            if 'USDT' in symbol:
                # Use the original price directly from the alert
                if price < 1.0:
                    # For low-value coins like DOGE, use more decimal places
                    return f"{price:.5f}"
                else:
                    return f"{price:.2f}"
        # Fallback to original price with 5 decimal places for safety
        return f"{price:.5f}"


# Helper function to check pending order status
def check_pending_order_status(client, order_id, order_data=None):
    """
    Check the status of a pending order, including partial fills and proper order type detection
    
    Args:
        client: Bybit client instance
        order_id (str): Order ID to check
        
    Returns:
        dict: Order details if found, None otherwise
    """
    try:
        # First try to get active orders
        active_response = client.get_open_orders(
            category="linear",
            orderId=order_id
        )
        
        if active_response.get('retCode') == 0:
            orders = active_response.get('result', {}).get('list', [])
            if orders:
                order = orders[0]
                
                # Check for cumulative executed quantity
                cum_exec_qty = float(order.get('cumExecQty', 0))
                orig_qty = float(order.get('qty', 0))
                
                # Determine if partially filled
                is_partial = cum_exec_qty > 0 and cum_exec_qty < orig_qty
                
                # Detect if this is a stop loss order by checking order type and stop order type
                is_stop_loss = False
                order_type = order.get('orderType', '')
                stop_order_type = order.get('stopOrderType', '')
                trigger_price = order.get('triggerPrice', 0)
                
                if order_type == 'Market' and stop_order_type in ['StopLoss', 'StopMarket']:
                    is_stop_loss = True
                # Reset not_found_count since we found the order successfully
                if order_data is not None and order_data.get('not_found_count', 0) > 0:
                    order_data['not_found_count'] = 0
                    save_pending_order(order_data)
                return {
                    'order_id': order.get('orderId'),
                    'symbol': order.get('symbol'),
                    'status': order.get('orderStatus'),
                    'price': float(order.get('price', 0)),
                    'qty': float(order.get('qty', 0)),
                    'side': order.get('side'),
                    'filled_qty': cum_exec_qty,
                    'remaining_qty': orig_qty - cum_exec_qty,
                    'is_partial': is_partial,
                    'order_type': order.get('orderType'),
                    'stop_order_type': stop_order_type,
                    'is_stop_loss': is_stop_loss,
                    'trigger_price': float(trigger_price) if trigger_price else 0,
                    'create_time': order.get('createdTime'),
                    'update_time': order.get('updatedTime')
                }
        
        # If not found in active orders, try order history
        history_response = client.get_order_history(
            category="linear",
            orderId=order_id
        )
        
        if history_response.get('retCode') == 0:
            orders = history_response.get('result', {}).get('list', [])
            if orders:
                order = orders[0]
                
                # Check for cumulative executed quantity
                cum_exec_qty = float(order.get('cumExecQty', 0))
                orig_qty = float(order.get('qty', 0))
                
                # Determine if partially filled
                is_partial = cum_exec_qty > 0 and cum_exec_qty < orig_qty
                
                # Detect if this is a stop loss order
                is_stop_loss = False
                order_type = order.get('orderType', '')
                stop_order_type = order.get('stopOrderType', '')
                trigger_price = order.get('triggerPrice', 0)
                
                if order_type == 'Market' and stop_order_type in ['StopLoss', 'StopMarket']:
                    is_stop_loss = True
                
                return {
                    'order_id': order.get('orderId'),
                    'symbol': order.get('symbol'),
                    'status': order.get('orderStatus'),
                    'price': float(order.get('price', 0)),
                    'qty': float(order.get('qty', 0)),
                    'side': order.get('side'),
                    'filled_qty': cum_exec_qty,
                    'remaining_qty': orig_qty - cum_exec_qty,
                    'is_partial': is_partial,
                    'order_type': order.get('orderType'),
                    'stop_order_type': stop_order_type,
                    'is_stop_loss': is_stop_loss,
                    'trigger_price': float(trigger_price) if trigger_price else 0,
                    'create_time': order.get('createdTime'),
                    'update_time': order.get('updatedTime')
                }
            else:
                # Order not found in history - could be connectivity issue or genuinely gone
                # Use a counter: only treat as Filled after 3 consecutive misses with good connectivity
                not_found_count = int(order_data.get('not_found_count', 0)) + 1
                order_data['not_found_count'] = not_found_count
                save_pending_order(order_data)
                
                if bot_status == "running" and not_found_count >= 3:
                    logger.info(f"Order {order_id} not found in history for {not_found_count} consecutive checks - treating as completed")
                    return {
                        'order_id': order_id,
                        'status': 'Filled',
                        'symbol': 'Unknown',
                        'price': 0,
                        'qty': 0,
                        'is_partial': False,
                        'is_stop_loss': order_data.get('is_stop_loss', False)
                    }
                else:
                    logger.warning(f"Order {order_id} not found (miss #{not_found_count}) - connectivity may be unstable, skipping for now")
                    return None
        
        # Check if both responses are OK but no order was found
        if active_response.get('retCode') == 0 and history_response.get('retCode') == 0:
            # Order not found anywhere - could be connectivity hiccup or genuinely gone
            # Use a counter: only treat as Filled after 3 consecutive misses with good connectivity
            not_found_count = int(order_data.get('not_found_count', 0)) + 1 if order_data else 1
            if order_data:
                order_data['not_found_count'] = not_found_count
                save_pending_order(order_data)
            
            if bot_status == "running" and not_found_count >= 3:
                logger.info(f"Order {order_id} not found anywhere for {not_found_count} consecutive checks - treating as completed")
                return {
                    'order_id': order_id,
                    'status': 'Filled',
                    'symbol': 'Unknown',
                    'price': 0,
                    'qty': 0,
                    'is_partial': False,
                    'is_stop_loss': order_data.get('is_stop_loss', False) if order_data else False
                }
            else:
                logger.warning(f"Order {order_id} not found anywhere (miss #{not_found_count}) - waiting for confirmation before acting")
                return None
                
        # If both checks fail with non-zero retCode
        if active_response.get('retCode') != 0:
            logger.warning(f"Failed to get active order status for {order_id}: {active_response.get('retMsg', 'Unknown error')}")
        
        if history_response.get('retCode') != 0:
            logger.warning(f"Failed to get order history for {order_id}: {history_response.get('retMsg', 'Unknown error')}")
        
        return None
    except Exception as e:
        logger.error(f"Error checking order status: {str(e)}")
        return None


# Function to process pending orders when filled
# Function to process pending orders when filled
def process_filled_order(order_data):
    """
    Process a filled order with transaction support and improved order type detection
    
    Args:
        order_data (dict): Order data including purpose and related position info
    """
    try:
        global active_positions, pending_orders
        
        # Get order details
        order_id = order_data['order_id']
        symbol = order_data['symbol']
        purpose = order_data.get('purpose', 'entry')
        
        # Start database connection for transaction
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        try:
            # Begin transaction
            cursor.execute('BEGIN TRANSACTION')
            
            # Handle based on purpose
            if purpose == 'entry':
                # This was an entry order that got filled - need to record the position
                side = order_data['side']
                direction = 'long' if side == 'Buy' else 'short'
                
                # Get stop loss from the original order data (may have been stored with the order)
                stop_loss = order_data.get('stop_loss', 0)
                take_profit = order_data.get('take_profit', 0)
                use_tp_limit = order_data.get('use_tp_limit_order', False)
                
                # Create position data
                # Try to get the actual fill price from the exchange (limit orders can fill at a better price)
                actual_entry_price = float(order_data['price'])  # Default fallback = limit price
                try:
                    fill_response = initialize_bybit_client().get_order_history(
                        category="linear",
                        symbol=symbol,
                        limit=10
                    )
                    if fill_response.get('retCode') == 0:
                        fill_list = fill_response.get('result', {}).get('list', [])
                        for fill in fill_list:
                            if str(fill.get('orderId', '')) == str(order_id):
                                avg_px = fill.get('avgPrice', '')
                                if avg_px and float(avg_px) > 0:
                                    actual_entry_price = float(avg_px)
                                    logger.info(f"Actual fill price for {symbol}: {actual_entry_price} (limit was {order_data['price']})")
                                break
                except Exception as price_err:
                    logger.warning(f"Could not fetch actual fill price for {order_id}, using limit price. Error: {price_err}")

                # Create position data
                position_data = {
                    'direction': direction,
                    'entry_price': actual_entry_price,
                    'position_size': float(order_data['qty']),
                    'stop_loss': stop_loss,
                    'take_profit': take_profit,
                    'order_id': order_id,
                    'entry_time': order_data['created_time'],
                    'last_check_time': datetime.now().isoformat(),
                    'account_balance': get_current_capital(),
                    'order_type': 'Limit',
                    'tp_order_type': 'Limit' if use_tp_limit else 'Market'
                }
                
                # Save to active positions
                active_positions[symbol] = position_data
                
                # Save position to database with proper column handling
                cursor.execute("PRAGMA table_info(positions)")
                columns = [col[1] for col in cursor.fetchall()]
                
                # Build dynamic insert query based on available columns
                column_names = []
                placeholders = []
                values = []
                
                for field, value in position_data.items():
                    if field in columns:
                        column_names.append(field)
                        placeholders.append('?')
                        values.append(value)
                
                # Make sure symbol is always included as it's the primary key
                if 'symbol' not in column_names:
                    column_names.append('symbol')
                    placeholders.append('?')
                    values.append(symbol)
                
                # Create dynamic query
                pos_query = f'''
                INSERT OR REPLACE INTO positions 
                (symbol, {', '.join(column_names)})
                VALUES (?, {', '.join(placeholders)})
                '''
                
                cursor.execute(pos_query, [symbol] + values)
                
                # Set stop loss and take profit in Bybit
                client = initialize_bybit_client()
                
                # First set standard stop loss and take profit (if not using limit TP)
                if stop_loss or (take_profit and not use_tp_limit):
                    # Prepare parameters
                    params = {
                        "category": "linear",
                        "symbol": symbol,
                        "positionIdx": 0  # For one-way position mode
                    }
                    
                    if stop_loss:
                        params["stopLoss"] = str(stop_loss)
                    
                    if take_profit and not use_tp_limit:
                        params["takeProfit"] = str(take_profit)
                    
                    sl_tp_response = client.set_trading_stop(**params)
                    
                    if sl_tp_response.get('retCode') == 0:
                        logger.info(f"Stop loss/take profit set for {direction} position on {symbol}")
                        # Track the SL order in pending_orders
                        sl_oid = sl_tp_response.get('result', {}).get('sl_order_id')
                        if sl_oid:
                            sl_order_data = {
                                'order_id': sl_oid,
                                'symbol': symbol,
                                'order_type': 'Market',
                                'stop_order_type': 'StopLoss',
                                'side': 'Sell' if direction == 'long' else 'Buy',
                                'price': float(stop_loss),
                                'trigger_price': float(stop_loss),
                                'qty': float(active_positions[symbol]['position_size']),
                                'status': 'New',
                                'created_time': datetime.now().isoformat(),
                                'purpose': 'exit',
                                'related_position_symbol': symbol,
                                'is_stop_loss': True
                            }
                            pending_orders[sl_oid] = sl_order_data
                            save_pending_order(sl_order_data)
                            logger.info(f"SL order {sl_oid} added to pending orders tracking for {symbol}")

                        # Track the native TP trigger order in pending_orders (if not using limit TP)
                        tp_oid = sl_tp_response.get('result', {}).get('tp_order_id')
                        if tp_oid and take_profit and not use_tp_limit:
                            tp_order_data = {
                                'order_id': tp_oid,
                                'symbol': symbol,
                                'order_type': 'Market',
                                'stop_order_type': 'TakeProfit',
                                'side': 'Sell' if direction == 'long' else 'Buy',
                                'price': float(take_profit),
                                'trigger_price': float(take_profit),
                                'qty': float(active_positions[symbol]['position_size']),
                                'status': 'New',
                                'created_time': datetime.now().isoformat(),
                                'purpose': 'exit',
                                'related_position_symbol': symbol,
                                'is_stop_loss': False
                            }
                            pending_orders[tp_oid] = tp_order_data
                            save_pending_order(tp_order_data)
                            active_positions[symbol]['tp_order_id'] = tp_oid
                            logger.info(f"Native TP order {tp_oid} added to pending orders tracking for {symbol}")
                    else:
                        logger.error(f"Failed to set stop loss/take profit: {sl_tp_response}")
                
                # Place limit take profit order if requested
                if take_profit and use_tp_limit:
                    # Check for existing TP orders before placing new ones
                    if has_existing_tp_order(symbol):
                        logger.info(f"Skipping TP order placement - existing TP order found for {symbol}")
                    else:
                        try:
                            # Get symbol info for proper formatting
                            symbol_info = get_symbol_info_with_retry(client, symbol)
                            
                            if not symbol_info:
                                logger.error(f"Could not get symbol info for {symbol} when setting limit TP")
                                logger.info(f"Falling back to standard take profit for {symbol}")
                                
                                # Fallback to standard TP
                                try:
                                    params = {
                                        "category": "linear",
                                        "symbol": symbol,
                                        "positionIdx": 0
                                    }
                                    params["takeProfit"] = str(take_profit)
                                    
                                    fallback_response = client.set_trading_stop(**params)
                                    if fallback_response.get('retCode') == 0:
                                        logger.info(f"Successfully set fallback standard take profit for {symbol} at {take_profit}")
                                    else:
                                        logger.error(f"Failed to set fallback take profit: {fallback_response}")
                                except Exception as fallback_error:
                                    logger.error(f"Error setting fallback take profit: {str(fallback_error)}")
                            else:
                                # Format quantity for TP order
                                tp_formatted_qty = str(float(order_data['qty']))
                                
                                # Determine the correct side for TP order (opposite of entry)
                                tp_side = "Sell" if direction == "long" else "Buy"
                                
                                # Place limit order for take profit
                                tp_response = client.place_order(
                                    category="linear",
                                    symbol=symbol,
                                    side=tp_side,
                                    orderType="Limit",
                                    qty=tp_formatted_qty,
                                    price=str(take_profit),
                                    timeInForce="GTC",
                                    reduceOnly=True
                                )
                                
                                if tp_response.get('retCode') == 0:
                                    tp_order_id = tp_response.get('result', {}).get('orderId')
                                    logger.info(f"Take profit limit order placed for {symbol} at {take_profit}")
                                    
                                    # Create pending order data for TP
                                    tp_order_data = {
                                        'order_id': tp_order_id,
                                        'symbol': symbol,
                                        'order_type': 'Limit',
                                        'side': tp_side,
                                        'price': float(take_profit),
                                        'qty': float(tp_formatted_qty),
                                        'status': 'New',
                                        'created_time': datetime.now().isoformat(),
                                        'purpose': 'exit',
                                        'related_position_symbol': symbol,
                                        'is_stop_loss': False
                                    }
                                    
                                    # Save to pending orders database
                                    save_pending_order(tp_order_data)
                                    
                                    # Update position data with TP order ID
                                    active_positions[symbol]['tp_order_id'] = tp_order_id
                                    cursor.execute('''
                                    UPDATE positions SET tp_order_id = ? WHERE symbol = ?
                                    ''', (tp_order_id, symbol))
                                    
                                    # Force save the tp_order_id through the standard save path too
                                    save_position(symbol, active_positions[symbol])
                                    
                                    # Start a thread to check and adjust TP size after a delay
                                    tp_adjustment_check_seconds = int(config['TRADING'].get('tp_adjustment_check_seconds', 10))
                                    if tp_adjustment_check_seconds > 0:
                                        threading.Thread(target=check_and_adjust_tp_size, args=(symbol, tp_adjustment_check_seconds), daemon=True).start()
                                        logger.info(f"Scheduled TP size adjustment check for {symbol} in {tp_adjustment_check_seconds} seconds")
                                else:
                                    logger.error(f"Failed to place take profit limit order: {tp_response}")
                        except Exception as tp_error:
                            logger.error(f"Error placing take profit limit order: {str(tp_error)}")
                
                # Send notification
                if config.getboolean('NOTIFICATIONS', 'notify_entries', fallback=True):
                    send_telegram_message(f"✅ <b>LIMIT ORDER FILLED</b>\n"
                                         f"Symbol: {symbol}\n"
                                         f"Direction: {direction.upper()}\n"
                                         f"Price: {position_data['entry_price']}\n"
                                         f"Size: {position_data['position_size']}\n"
                                         f"Stop Loss: {stop_loss if stop_loss else 'None'}\n"
                                         f"Take Profit: {take_profit if take_profit else 'None'}"
                                         + (" (Limit Order)" if take_profit and use_tp_limit else ""))
            
            elif purpose == 'exit':
                # This was an exit order that got filled - need to record the trade
                related_symbol = order_data.get('related_position_symbol', symbol)
                
                # Check if we have the position in active_positions
                if related_symbol in active_positions:
                    position = active_positions[related_symbol]
                    
                    # Determine if this is a stop loss order
                    is_stop_loss = order_data.get('is_stop_loss', False)
                    order_type = order_data.get('order_type', 'Limit')
                    
                    # Determine exit reason based on order type
                    if is_stop_loss:
                        exit_reason = "Stop loss triggered"
                        order_type = 'StopLoss'
                    else:
                        exit_reason = "Limit exit order filled"
                    
                    # Get the filled quantity from order_data
                    filled_qty = float(order_data.get('qty', 0))
                    position_size = float(position['position_size'])
                    
                    # Check if this is a partial fill (filled_qty < position_size)
                    is_partial_fill = filled_qty < position_size and filled_qty > 0
                    
                    if is_partial_fill:
                        logger.info(f"Partial TP fill detected for {related_symbol}: {filled_qty}/{position_size} units")
                        
                        # Calculate the remaining position size
                        remaining_size = position_size - filled_qty
                        
                        # Record a trade for the filled portion only
                        record_completed_trade(
                            symbol=related_symbol,
                            direction=position['direction'],
                            entry_price=position['entry_price'],
                            exit_price=float(order_data['price']),
                            position_size=filled_qty,  # Only the filled portion
                            entry_time=position['entry_time'],
                            stop_loss=position.get('stop_loss'),
                            stopped_out=is_stop_loss,
                            reason=f"Partial {exit_reason.lower()} ({filled_qty}/{position_size} units)",
                            order_id=order_id,
                            order_type=order_type,
                            is_partial_fill=True
                        )
                        
                        # Update the position with the remaining size
                        position['position_size'] = remaining_size
                        
                        # Update the position in the database
                        cursor.execute('''
                        UPDATE positions SET position_size = ? WHERE symbol = ?
                        ''', (remaining_size, related_symbol))
                        
                        # Send notification for partial exit
                        if config.getboolean('NOTIFICATIONS', 'notify_exits', fallback=True):
                            # Determine notification title based on order type
                            notification_title = "PARTIAL LIMIT EXIT"
                            if is_stop_loss:
                                notification_title = "PARTIAL STOP LOSS"
                            
                            send_telegram_message(f"✅ <b>{notification_title}</b>\n"
                                                 f"Symbol: {related_symbol}\n"
                                                 f"Direction: {position['direction'].upper()}\n"
                                                 f"Exit Price: {order_data['price']}\n"
                                                 f"Filled: {filled_qty}/{position_size} units\n"
                                                 f"Remaining: {remaining_size} units")
                        
                        # Start a thread to check and adjust TP size for the remaining position
                        # This is needed because the TP order may now be out of sync with the position size
                        tp_adjustment_check_seconds = int(config['TRADING'].get('tp_adjustment_check_seconds', 10))
                        if tp_adjustment_check_seconds > 0:
                            threading.Thread(target=check_and_adjust_tp_size, args=(related_symbol, 5), daemon=True).start()
                            logger.info(f"Scheduled TP size adjustment check for {related_symbol} after partial fill")
                    else:
                        # FIX 3: Before closing position, verify it's actually gone from the exchange
                        # This prevents false closures caused by connectivity hiccups
                        try:
                            verification_client = initialize_bybit_client()
                            exchange_position = get_bybit_position_details(verification_client, related_symbol)
                            if exchange_position is not None and float(exchange_position.get('position_size', 0)) > 0:
                                logger.warning(
                                    f"FIX3 ABORT: Order {order_id} for {related_symbol} was marked Filled, "
                                    f"but position still exists on exchange with size {exchange_position['position_size']}. "
                                    f"Skipping position close - likely a false fill detection."
                                )
                                # Reset the not_found_count so this order gets re-evaluated fresh
                                if order_id in pending_orders:
                                    pending_orders[order_id]['not_found_count'] = 0
                                    pending_orders[order_id]['status'] = 'New'
                                    save_pending_order(pending_orders[order_id])
                                return
                        except Exception as verify_err:
                            logger.warning(f"FIX3: Could not verify position on exchange: {verify_err} — proceeding with close")

                        # This is a full fill - record the trade and remove the position
                        record_completed_trade(
                            symbol=related_symbol,
                            direction=position['direction'],
                            entry_price=position['entry_price'],
                            exit_price=float(order_data['price']),
                            position_size=position_size,
                            entry_time=position['entry_time'],
                            stop_loss=position.get('stop_loss'),
                            stopped_out=is_stop_loss,
                            reason=exit_reason,
                            order_id=order_id,
                            order_type=order_type
                        )
                        
                        # Remove position from database
                        cursor.execute('DELETE FROM positions WHERE symbol = ?', (related_symbol,))
                        
                        # Remove from active positions (memory)
                        if related_symbol in active_positions:
                            del active_positions[related_symbol]
                        
                        # Send notification
                        if config.getboolean('NOTIFICATIONS', 'notify_exits', fallback=True):
                            # Determine notification title based on order type
                            notification_title = "LIMIT EXIT ORDER FILLED"
                            if is_stop_loss:
                                notification_title = "STOP LOSS TRIGGERED"
                            
                            send_telegram_message(f"✅ <b>{notification_title}</b>\n"
                                                 f"Symbol: {related_symbol}\n"
                                                 f"Direction: {position['direction'].upper()}\n"
                                                 f"Exit Price: {order_data['price']}\n"
                                                 f"Position Size: {position_size}")
                else:
                    logger.warning(f"Position {related_symbol} not found in active positions when processing filled exit order")
            
            # Remove the processed order from pending orders database
            cursor.execute('DELETE FROM pending_orders WHERE order_id = ?', (order_id,))
            
            # Commit the transaction
            conn.commit()
            
            # Only remove from memory after successful commit
            if order_id in pending_orders:
                del pending_orders[order_id]
            
            logger.info(f"Successfully processed order {order_id} for {symbol}")
        
        except Exception as e:
            # Rollback transaction on error
            conn.rollback()
            logger.error(f"Error in process_filled_order transaction: {str(e)}")
            raise
        
        finally:
            conn.close()
        
    except Exception as e:
        logger.error(f"Error processing filled order: {str(e)}")


# Function to check and update pending limit orders
# Function to check and update pending limit orders
def check_pending_orders():
    """Check status of all pending limit orders and process filled ones"""
    try:
        global pending_orders
        
        # Skip if no pending orders
        if not pending_orders:
            return
        
        # NEW SAFETY CHECK: Skip if connectivity is not fully confirmed
        if bot_status != "running":
            logger.debug("Skipping pending orders check - connectivity not confirmed (bot_status: " + bot_status + ")")
            return
        
        # Only log the count at debug level
        if len(pending_orders) > 0:
            logger.debug(f"Checking status of {len(pending_orders)} pending orders")
        
        client = initialize_bybit_client()
        current_time = datetime.now()
        
        # Check each pending order
        for order_id, order_data in list(pending_orders.items()):
            # Check if order has an expiry time
            if 'expiry_time' in order_data and order_data['expiry_time']:
                try:
                    expiry_time = datetime.fromisoformat(order_data['expiry_time'])
                    
                    # Cancel if expired
                    if current_time > expiry_time:
                        logger.info(f"Order {order_id} for {order_data['symbol']} has expired, canceling")
                        
                        try:
                            cancel_response = client.cancel_order(
                                category="linear",
                                symbol=order_data['symbol'],
                                orderId=order_id
                            )
                            
                            if cancel_response.get('retCode') == 0:
                                logger.info(f"Successfully canceled expired order {order_id}")
                                
                                # Send notification
                                if config.getboolean('NOTIFICATIONS', 'notify_order_status', fallback=True):
                                    send_telegram_message(f"⏱️ <b>LIMIT ORDER EXPIRED</b>\n"
                                                        f"Symbol: {order_data['symbol']}\n"
                                                        f"Side: {order_data['side']}\n"
                                                        f"Price: {order_data['price']}\n"
                                                        f"Quantity: {order_data['qty']}")
                                
                                # Remove from pending orders
                                remove_pending_order(order_id)
                                continue
                            else:
                                logger.warning(f"Failed to cancel expired order {order_id}: {cancel_response}")
                        except Exception as cancel_error:
                            logger.error(f"Error canceling expired order: {str(cancel_error)}")
                except ValueError:
                    logger.warning(f"Invalid expiry_time format for order {order_id}")
            
            # Check current status with retry mechanism
            order_status = None
            max_retries = 3
            retry_count = 0
            
            while retry_count < max_retries and order_status is None:
                try:
                    order_status = check_pending_order_status(client, order_id, order_data)
                    break  # Success, exit the retry loop
                except Exception as e:
                    retry_count += 1
                    if retry_count >= max_retries:
                        logger.warning(f"Failed to get status for order {order_id} after {max_retries} attempts: {str(e)}")
                    else:
                        logger.debug(f"Retrying order status check, attempt {retry_count}/{max_retries}")
                        time.sleep(2)  # Wait before retrying
            
            if not order_status:
                # Don't remove orders if status check fails - might be connection issue
                logger.debug(f"Could not get status for order {order_id}, will try again next cycle")
                continue  # Skip to next order without removing this one
            
            # Check if status has changed
            status_changed = order_status['status'] != order_data.get('status', 'Unknown')
            
            if status_changed:
                # Only log and update database if status has changed
                logger.info(f"Order {order_id} for {order_data['symbol']} status changed: {order_data.get('status', 'Unknown')} → {order_status['status']}")
                
                # Update order status in memory
                order_data['status'] = order_status['status']
                
                # Check if this is a stop loss order and update order_data
                # IMPORTANT: Only upgrade to True from API status, never downgrade.
                # The original tracking data is the authoritative source.
                if order_status.get('is_stop_loss', False) and not order_data.get('is_stop_loss', False):
                    order_data['is_stop_loss'] = True
                    order_data['stop_order_type'] = order_status.get('stop_order_type', '')
                    if order_status.get('trigger_price', 0) > 0:
                        order_data['trigger_price'] = order_status['trigger_price']
                # Always update stop_order_type if API provides it and we don't have one
                elif order_status.get('stop_order_type', '') and not order_data.get('stop_order_type', ''):
                    order_data['stop_order_type'] = order_status.get('stop_order_type', '')
                
                # Only save to database if status changed
                save_pending_order(order_data)
            else:
                # Just log at debug level, don't update database
                logger.debug(f"Order {order_id} for {order_data['symbol']} has status {order_status['status']}, still pending")
            
            # Process based on status
            if order_status['status'] == 'Filled' or (order_status.get('is_partial', False) and order_status.get('filled_qty', 0) > 0):
                if order_status.get('is_partial', False):
                    logger.info(f"Order {order_id} for {order_data['symbol']} has been partially filled ({order_status['filled_qty']}/{order_status['qty']} units)")
                    
                    # Create a copy for processing the filled portion
                    filled_order_data = order_data.copy()
                    filled_order_data['qty'] = order_status['filled_qty']
                    filled_order_data['is_stop_loss'] = order_status.get('is_stop_loss', False)
                    
                    # Get the original order's purpose
                    purpose = order_data.get('purpose', 'entry')
                    
                    # Handle based on purpose
                    if purpose == 'exit':
                        # For exit (TP) orders, update the pending order with the remaining quantity
                        # but keep it in pending_orders so it continues to be monitored
                        remaining_order_data = order_data.copy()
                        remaining_order_data['qty'] = order_status['remaining_qty']
                        save_pending_order(remaining_order_data)
                    else:
                        # For entry orders, use standard partial fill processing
                        # (This case is less common since entry orders typically execute all-or-none)
                        remaining_order_data = order_data.copy()
                        remaining_order_data['qty'] = order_status['remaining_qty']
                        save_pending_order(remaining_order_data)
                    
                    # Process the filled portion
                    process_filled_order(filled_order_data)
                    
                    # If fully filled now, remove from pending orders
                    if order_status['remaining_qty'] <= 0:
                        remove_pending_order(order_id)
                else:
                    logger.info(f"Order {order_id} for {order_data['symbol']} has been filled")
                    
                    # Add stop loss info to order_data if applicable
                    order_data['is_stop_loss'] = order_status.get('is_stop_loss', False)
                    
                    process_filled_order(order_data)
                    
                    # If fully filled, remove from pending orders
                    if not order_status.get('is_partial', False):
                        remove_pending_order(order_id)
            elif order_status['status'] in ['Cancelled', 'Rejected', 'Failed']:
                logger.info(f"Order {order_id} for {order_data['symbol']} has status {order_status['status']}, removing")
                
                # Send notification
                if config.getboolean('NOTIFICATIONS', 'notify_order_status', fallback=True):
                    send_telegram_message(f"❌ <b>LIMIT ORDER {order_status['status'].upper()}</b>\n"
                                        f"Symbol: {order_data['symbol']}\n"
                                        f"Side: {order_data['side']}\n"
                                        f"Price: {order_data['price']}\n"
                                        f"Quantity: {order_data['qty']}")
                
                # Remove from pending orders
                remove_pending_order(order_id)
            
    except Exception as e:
        logger.error(f"Error checking pending orders: {str(e)}")


# Start a background thread to monitor pending orders
def start_pending_orders_monitor():
    """Start a background thread to monitor pending limit orders"""
    def monitor_loop():
        while True:
            try:
                # Sleep for the configured interval
                check_interval = int(config.get('MONITORING', 'pending_orders_check_seconds', fallback=30))
                time.sleep(check_interval)
                
                # Check pending orders
                check_pending_orders()
            except Exception as e:
                logger.error(f"Error in pending orders monitor: {str(e)}")
                time.sleep(60)  # Sleep for 60 seconds on error before retrying
    
    # Start the monitoring thread
    monitor_thread = threading.Thread(target=monitor_loop, daemon=True)
    monitor_thread.start()
    logger.info(f"Started pending orders monitor thread with interval: {config.get('MONITORING', 'pending_orders_check_seconds', fallback=30)} seconds")



def start_cancellation_cleanup_thread():
    """Start a background thread to periodically clean up failed cancellations"""
    def cleanup_loop():
        while True:
            try:
                # Sleep first to let the bot initialize
                time.sleep(120)  # Run every 2 minutes
                
                # Process the cleanup queue
                process_cancellation_cleanup_queue()
            except Exception as e:
                logger.error(f"Error in cancellation cleanup loop: {str(e)}")
                time.sleep(60)  # Sleep for 1 minute on error before retrying
    
    # Start the cleanup thread
    cleanup_thread = threading.Thread(target=cleanup_loop, daemon=True)
    cleanup_thread.start()
    logger.info("Started cancellation cleanup thread")


# Execute trade on Bybit
# Execute trade on Bybit
# Execute trade on Bybit
# Execute trade on Bybit
def execute_trade(client, action, symbol, entry_price, stop_loss=None, position_size=None, take_profit=None, use_limit_order=None, use_tp_limit_order=None, closeOnTrigger=False):
    try:
        # Determine if we should use limit orders from parameters or config
        use_limit_orders_entry = use_limit_order if use_limit_order is not None else config['TRADING'].getboolean('use_limit_orders_entry', False)
        use_limit_orders_exit = use_tp_limit_order if use_tp_limit_order is not None else config['TRADING'].getboolean('use_limit_orders_exit', False)
        
        # Get limit order timeout in minutes
        limit_order_timeout_mins = int(config['TRADING'].get('limit_order_timeout_minutes', 5))
        
        # Get symbol information
        symbol_info = get_symbol_info_with_retry(client, symbol)
        if not symbol_info:
            logger.error(f"Symbol information not found for {symbol}")
            
            # Send notification for error
            if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                send_telegram_message(f"❌ <b>ERROR</b>\n"
                                      f"Failed to get symbol information for {symbol}")
            return False
        
        if action.startswith("LONG ENTRY"):
            # If position size is not provided, calculate it
            if not position_size:
                position_size = calculate_position_size(entry_price, stop_loss, symbol_info, None, use_limit_orders_entry)
                
            if position_size <= 0:
                logger.warning("Position size calculation resulted in zero or negative value, skipping trade")
                
                # Send notification for error
                if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                    send_telegram_message(f"⚠️ <b>WARNING</b>\n"
                                        f"Skipped {symbol} LONG ENTRY - Position size too small")
                return False
                
            # Set leverage for the position
            if position_size > 0:
                leverage_success = set_appropriate_leverage(client, symbol, entry_price, position_size)
                if not leverage_success:
                    logger.warning(f"Could not set appropriate leverage for {symbol}, proceeding with exchange default")
                    
                    # Send notification for warning
                    if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                        send_telegram_message(f"⚠️ <b>WARNING</b>\n"
                                            f"Could not set leverage for {symbol} LONG position\n"
                                            f"Proceeding with exchange default leverage")
            
            # Format the position size according to exchange requirements
            formatted_qty = format_position_quantity(position_size, symbol_info)
            logger.info(f"Formatted position size for {symbol}: {position_size} → {formatted_qty}")
            
            # Determine order type and parameters
            if use_limit_orders_entry:
                # Format limit price (slightly below market price for buys)
                formatted_price = format_price(entry_price, symbol_info, is_buy=True)
                
                # Calculate expiry time
                expiry_time = (datetime.now() + timedelta(minutes=limit_order_timeout_mins)).isoformat()
                
                # Place limit order to enter long position
                response = client.place_order(
                    category="linear",
                    symbol=symbol,
                    side="Buy",
                    orderType="Limit",
                    qty=formatted_qty,
                    price=formatted_price,
                    timeInForce="GTC",  # Good Till Canceled
                    reduceOnly=False
                )
                
                if response.get('retCode') == 0:
                    order_id = response.get('result', {}).get('orderId')
                    
                    # Create pending order data
                    order_data = {
                        'order_id': order_id,
                        'symbol': symbol,
                        'order_type': 'Limit',
                        'side': 'Buy',
                        'price': float(formatted_price),
                        'qty': float(formatted_qty),
                        'status': 'New',
                        'created_time': datetime.now().isoformat(),
                        'purpose': 'entry',
                        'stop_loss': stop_loss,
                        'take_profit': take_profit,
                        'use_tp_limit_order': use_limit_orders_exit,
                        'expiry_time': expiry_time
                    }
                    
                    # Save to pending orders database
                    save_pending_order(order_data)
                    
                    logger.info(f"Limit order placed for LONG entry on {symbol}: {formatted_qty} units at {formatted_price}")
                    
                    # Send notification for limit order placement
                    if config.getboolean('NOTIFICATIONS', 'notify_entries', fallback=True):
                        if stop_loss:
                            maker_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02)) / 100
                            taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100
                            price_risk = (float(formatted_price) - stop_loss) * position_size
                            entry_fee = maker_fee_percent * float(formatted_price) * position_size
                            exit_fee = taker_fee_percent * stop_loss * position_size
                            fee_cost = entry_fee + exit_fee
                            total_risk = price_risk + fee_cost
                            risk_message = f"Risk $: ${total_risk:.2f} (Price: ${price_risk:.2f}, Fees: ${fee_cost:.2f})"
                        else:
                            risk_message = "Risk: N/A (No stop loss set)"
                        
                        # Add take profit to message if provided
                        take_profit_message = f"Take Profit: {take_profit}\n" if take_profit else ""
                        
                        send_telegram_message(f"🟢 <b>LONG ENTRY LIMIT ORDER PLACED</b>\n"
                                            f"Symbol: {symbol}\n"
                                            f"Size: {position_size}\n"
                                            f"Limit Price: {formatted_price}\n"
                                            f"Stop Loss: {stop_loss if stop_loss else 'None'}\n"
                                            f"{take_profit_message}"
                                            f"{risk_message}\n"
                                            f"Expires: in {limit_order_timeout_mins} mins")
                    
                    return True
                else:
                    logger.error(f"Failed to place limit order for LONG entry: {response}")
                    
                    # Send notification for error
                    if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                        send_telegram_message(f"❌ <b>ERROR</b>\n"
                                            f"Failed to place limit order for LONG entry on {symbol}:\n"
                                            f"{response.get('retMsg', 'Unknown error')}")
                    
                    return False
            else:
                # Place market order to enter long position
                response = client.place_order(
                    category="linear",
                    symbol=symbol,
                    side="Buy",
                    orderType="Market",
                    qty=formatted_qty,
                    reduceOnly=False
                )
                
                if response.get('retCode') == 0:
                    order_id = response.get('result', {}).get('orderId')
                    
                    # Get the actual execution price from Bybit
                    time.sleep(1)  # Small delay to ensure order is processed
                    order_detail = client.get_order_history(
                        category="linear",
                        symbol=symbol,
                        orderId=order_id
                    )
                    
                    actual_entry_price = entry_price  # Default fallback to estimated price
                    if order_detail.get('retCode') == 0:
                        # Get the executed price from order details
                        order_result = order_detail.get('result', {}).get('list', [{}])[0]
                        if float(order_result.get('execQty', 0)) > 0:
                            actual_entry_price = float(order_result.get('avgPrice', entry_price))
                    
                    logger.info(f"Long position entered for {symbol}: {position_size} units at {actual_entry_price} (Bybit execution price)")
                    
                    # Get current account balance
                    current_balance = get_current_capital()

                    # Store position information with actual entry price
                    position_data = {
                        'direction': 'long',
                        'entry_price': actual_entry_price,  # Use actual price from Bybit
                        'position_size': position_size,
                        'stop_loss': stop_loss,
                        'take_profit': take_profit,  # Add take profit to position data
                        'order_id': order_id,
                        'entry_time': datetime.now().isoformat(),
                        'last_check_time': datetime.now().isoformat(),
                        'account_balance': current_balance,
                        'order_type': 'Market',
                        'tp_order_type': 'Limit' if use_limit_orders_exit else 'Market'
                    }

                    active_positions[symbol] = position_data
                    
                    # Save position to database for persistence
                    save_position(symbol, position_data)
                    
                    # Set stop loss and take profit using set_trading_stop API if provided
                    if stop_loss or (take_profit and not use_limit_orders_exit):
                        # Add a small delay to ensure position is registered
                        time.sleep(1)
                        
                        # Prepare parameters
                        params = {
                            "category": "linear",
                            "symbol": symbol,
                            "positionIdx": 0  # For one-way position mode
                        }
                        
                        if stop_loss:
                            params["stopLoss"] = str(stop_loss)
                        
                        if take_profit and not use_limit_orders_exit:
                            params["takeProfit"] = str(take_profit)
                        
                        sl_tp_response = client.set_trading_stop(**params)
                        
                        if sl_tp_response.get('retCode') == 0:
                            logger.info(f"Stop loss/take profit set for long position on {symbol}")
                            if stop_loss:
                                logger.info(f"Stop loss: {stop_loss}")
                            if take_profit and not use_limit_orders_exit:
                                logger.info(f"Take profit: {take_profit}")
                            # Track the SL order in pending_orders
                            sl_oid = sl_tp_response.get('result', {}).get('sl_order_id')
                            if sl_oid:
                                sl_order_data = {
                                    'order_id': sl_oid,
                                    'symbol': symbol,
                                    'order_type': 'Market',
                                    'stop_order_type': 'StopLoss',
                                    'side': 'Sell',
                                    'price': float(stop_loss),
                                    'trigger_price': float(stop_loss),
                                    'qty': float(active_positions[symbol]['position_size']),
                                    'status': 'New',
                                    'created_time': datetime.now().isoformat(),
                                    'purpose': 'exit',
                                    'related_position_symbol': symbol,
                                    'is_stop_loss': True
                                }
                                pending_orders[sl_oid] = sl_order_data
                                save_pending_order(sl_order_data)
                                logger.info(f"SL order {sl_oid} added to pending orders tracking for {symbol}")

                            # Track the native TP trigger order in pending_orders (if not using limit TP)
                            tp_oid = sl_tp_response.get('result', {}).get('tp_order_id')
                            if tp_oid and take_profit and not use_limit_orders_exit:
                                tp_order_data = {
                                    'order_id': tp_oid,
                                    'symbol': symbol,
                                    'order_type': 'Market',
                                    'stop_order_type': 'TakeProfit',
                                    'side': 'Sell',
                                    'price': float(take_profit),
                                    'trigger_price': float(take_profit),
                                    'qty': float(active_positions[symbol]['position_size']),
                                    'status': 'New',
                                    'created_time': datetime.now().isoformat(),
                                    'purpose': 'exit',
                                    'related_position_symbol': symbol,
                                    'is_stop_loss': False
                                }
                                pending_orders[tp_oid] = tp_order_data
                                save_pending_order(tp_order_data)
                                active_positions[symbol]['tp_order_id'] = tp_oid
                                logger.info(f"Native TP order {tp_oid} added to pending orders tracking for {symbol}")

                            # Update position in database
                            save_position(symbol, active_positions[symbol])
                        else:
                            logger.error(f"Failed to set stop loss/take profit using trading stop API: {sl_tp_response}")
                            
                            # Send notification for error
                            if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                                message = f"Failed to set "
                                if stop_loss and take_profit:
                                    message += "stop loss and take profit"
                                elif stop_loss:
                                    message += "stop loss"
                                else:
                                    message += "take profit"
                                
                                send_telegram_message(f"⚠️ <b>WARNING</b>\n"
                                                    f"{message} for {symbol} long position")
                    
                    # Place take profit limit order if requested
                    if take_profit and use_limit_orders_exit:
                        # Check for existing TP orders before placing new ones
                        if has_existing_tp_order(symbol):
                            logger.info(f"Skipping TP order placement - existing TP order found for {symbol}")
                        else:
                            try:
                                # Get exact position size from Bybit
                                exact_position_size = get_exact_position_size_from_exchange(client, symbol)
                                if exact_position_size is not None:
                                    position_size = exact_position_size
                                    logger.info(f"Using exact position size from Bybit for TP: {position_size}")
                                
                                # Use exact position size for TP orders to avoid rounding issues
                                tp_formatted_qty = str(position_size)
                                
                                # Place limit order for take profit
                                tp_response = client.place_order(
                                    category="linear",
                                    symbol=symbol,
                                    side="Sell",  # Sell to close long position
                                    orderType="Limit",
                                    qty=tp_formatted_qty,
                                    price=str(take_profit),
                                    timeInForce="GTC",
                                    reduceOnly=True,
                                    closeOnTrigger=True  # Added closeOnTrigger parameter
                                )
                                
                                if tp_response.get('retCode') == 0:
                                    tp_order_id = tp_response.get('result', {}).get('orderId')
                                    logger.info(f"Take profit limit order placed for {symbol} at {take_profit}")
                                    
                                    # Create pending order data for TP
                                    tp_order_data = {
                                        'order_id': tp_order_id,
                                        'symbol': symbol,
                                        'order_type': 'Limit',
                                        'side': 'Sell',
                                        'price': float(take_profit),
                                        'qty': float(position_size),
                                        'status': 'New',
                                        'created_time': datetime.now().isoformat(),
                                        'purpose': 'exit',
                                        'related_position_symbol': symbol,
                                        'is_stop_loss': False
                                    }
                                    
                                    # Save to pending orders database
                                    save_pending_order(tp_order_data)
                                    
                                    # Update position data with TP order ID
                                    active_positions[symbol]['tp_order_id'] = tp_order_id
                                    save_position(symbol, active_positions[symbol])
                                    
                                    # Force save the tp_order_id through the standard save path too
                                    save_position(symbol, active_positions[symbol])
                                    
                                    # Start a thread to check and adjust TP size after a delay
                                    tp_adjustment_check_seconds = int(config['TRADING'].get('tp_adjustment_check_seconds', 10))
                                    if tp_adjustment_check_seconds > 0:
                                        threading.Thread(target=check_and_adjust_tp_size, args=(symbol, tp_adjustment_check_seconds), daemon=True).start()
                                        logger.info(f"Scheduled TP size adjustment check for {symbol} in {tp_adjustment_check_seconds} seconds")
                                else:
                                    logger.error(f"Failed to place take profit limit order: {tp_response}")
                            except Exception as tp_error:
                                logger.error(f"Error placing take profit limit order: {str(tp_error)}")
                    
                    # Send notification for entry
                    if config.getboolean('NOTIFICATIONS', 'notify_entries', fallback=True):
                        if stop_loss:
                            taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100
                            price_risk = (actual_entry_price - stop_loss) * position_size
                            entry_fee = taker_fee_percent * actual_entry_price * position_size
                            exit_fee = taker_fee_percent * stop_loss * position_size
                            fee_cost = entry_fee + exit_fee
                            total_risk = price_risk + fee_cost
                            risk_message = f"Risk: ${total_risk:.2f} (Price: ${price_risk:.2f}, Fees: ${fee_cost:.2f})"
                        else:
                            risk_message = "Risk: N/A (No stop loss set)"
                        
                        # Add take profit to message if provided
                        take_profit_message = f"Take Profit: {take_profit}" if take_profit else "Take Profit: None"
                        if take_profit and use_limit_orders_exit:
                            take_profit_message += " (Limit Order)"
                        take_profit_message += "\n"
                        
                        send_telegram_message(f"🟢 <b>LONG ENTRY</b>\n"
                                            f"Symbol: {symbol}\n"
                                            f"Size: {position_size}\n"
                                            f"Price: {actual_entry_price}\n"
                                            f"Stop Loss: {stop_loss if stop_loss else 'None'}\n"
                                            f"{take_profit_message}"
                                            f"{risk_message}")
                    
                    return True
                else:
                    if response.get('retCode') == 10001 and "qty invalid" in str(response.get('retMsg', '')).lower():
                        # Specific error handling for quantity formatting issues
                        logger.error(f"Order quantity rejected by Bybit: {response.get('retMsg')}. " 
                                    f"Original quantity: {position_size}, Formatted quantity: {formatted_qty}. "
                                    f"Symbol info qty_step: {symbol_info.get('lotSizeFilter', {}).get('qtyStep', 'unknown')}")
                        
                        # Send more detailed notification for troubleshooting
                        if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                            send_telegram_message(f"❌ <b>ERROR</b>\n"
                                                f"Order quantity rejected for {symbol}:\n"
                                                f"Issue: {response.get('retMsg', 'Unknown error')}\n"
                                                f"Original qty: {position_size}\n"
                                                f"Formatted qty: {formatted_qty}")
                    else:
                        logger.error(f"Failed to enter long position: {response}")
                        
                        # Send notification for error
                        if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                            send_telegram_message(f"❌ <b>ERROR</b>\n"
                                                f"Failed to enter long position for {symbol}:\n"
                                                f"{response.get('retMsg', 'Unknown error')}")
                    
                    return False
                
        elif action.startswith("SHORT ENTRY"):
            # If position size is not provided, calculate it
            if not position_size:
                position_size = calculate_position_size(entry_price, stop_loss, symbol_info, None, use_limit_orders_entry)
                
            if position_size <= 0:
                logger.warning("Position size calculation resulted in zero or negative value, skipping trade")
                
                # Send notification for error
                if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                    send_telegram_message(f"⚠️ <b>WARNING</b>\n"
                                        f"Skipped {symbol} SHORT ENTRY - Position size too small")
                
                return False
                
            # Set leverage for the position
            if position_size > 0:
                leverage_success = set_appropriate_leverage(client, symbol, entry_price, position_size)
                if not leverage_success:
                    logger.warning(f"Could not set appropriate leverage for {symbol}, proceeding with exchange default")
                    
                    # Send notification for warning
                    if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                        send_telegram_message(f"⚠️ <b>WARNING</b>\n"
                                            f"Could not set leverage for {symbol} SHORT position\n"
                                            f"Proceeding with exchange default leverage")
            
            # Format the position size according to exchange requirements
            formatted_qty = format_position_quantity(position_size, symbol_info)
            logger.info(f"Formatted position size for {symbol}: {position_size} → {formatted_qty}")
            
            # Determine order type and parameters
            if use_limit_orders_entry:
                # Format limit price (slightly above market price for sells)
                formatted_price = format_price(entry_price, symbol_info, is_buy=False)
                
                # Calculate expiry time
                expiry_time = (datetime.now() + timedelta(minutes=limit_order_timeout_mins)).isoformat()
                
                # Place limit order to enter short position
                response = client.place_order(
                    category="linear",
                    symbol=symbol,
                    side="Sell",
                    orderType="Limit",
                    qty=formatted_qty,
                    price=formatted_price,
                    timeInForce="GTC",  # Good Till Canceled
                    reduceOnly=False
                )
                
                if response.get('retCode') == 0:
                    order_id = response.get('result', {}).get('orderId')
                    
                    # Create pending order data
                    order_data = {
                        'order_id': order_id,
                        'symbol': symbol,
                        'order_type': 'Limit',
                        'side': 'Sell',
                        'price': float(formatted_price),
                        'qty': float(formatted_qty),
                        'status': 'New',
                        'created_time': datetime.now().isoformat(),
                        'purpose': 'entry',
                        'stop_loss': stop_loss,
                        'take_profit': take_profit,
                        'use_tp_limit_order': use_limit_orders_exit,
                        'expiry_time': expiry_time
                    }
                    
                    # Save to pending orders database
                    save_pending_order(order_data)
                    
                    logger.info(f"Limit order placed for SHORT entry on {symbol}: {formatted_qty} units at {formatted_price}")
                    
                    # Send notification for limit order placement
                    if config.getboolean('NOTIFICATIONS', 'notify_entries', fallback=True):
                        if stop_loss:
                            maker_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02)) / 100
                            taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100
                            price_risk = (stop_loss - float(formatted_price)) * position_size
                            entry_fee = maker_fee_percent * float(formatted_price) * position_size
                            exit_fee = taker_fee_percent * stop_loss * position_size
                            fee_cost = entry_fee + exit_fee
                            total_risk = price_risk + fee_cost
                            risk_message = f"Risk $: ${total_risk:.2f} (Price: ${price_risk:.2f}, Fees: ${fee_cost:.2f})"
                        else:
                            risk_message = "Risk: N/A (No stop loss set)"
                        
                        # Add take profit to message if provided
                        take_profit_message = f"Take Profit: {take_profit}\n" if take_profit else ""
                        
                        send_telegram_message(f"🔴 <b>SHORT ENTRY LIMIT ORDER PLACED</b>\n"
                                            f"Symbol: {symbol}\n"
                                            f"Size: {position_size}\n"
                                            f"Limit Price: {formatted_price}\n"
                                            f"Stop Loss: {stop_loss if stop_loss else 'None'}\n"
                                            f"{take_profit_message}"
                                            f"{risk_message}\n"
                                            f"Expires: in {limit_order_timeout_mins} mins")
                    
                    return True
                else:
                    logger.error(f"Failed to place limit order for SHORT entry: {response}")
                    
                    # Send notification for error
                    if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                        send_telegram_message(f"❌ <b>ERROR</b>\n"
                                            f"Failed to place limit order for SHORT entry on {symbol}:\n"
                                            f"{response.get('retMsg', 'Unknown error')}")
                    
                    return False
            else:
                # Place market order to enter short position
                response = client.place_order(
                    category="linear",
                    symbol=symbol,
                    side="Sell",
                    orderType="Market",
                    qty=formatted_qty,
                    reduceOnly=False
                )
                
                if response.get('retCode') == 0:
                    order_id = response.get('result', {}).get('orderId')

                    # Get the actual execution price from Bybit
                    time.sleep(1)  # Small delay to ensure order is processed
                    order_detail = client.get_order_history(
                        category="linear",
                        symbol=symbol,
                        orderId=order_id
                    )
                    
                    actual_entry_price = entry_price  # Default fallback to estimated price
                    if order_detail.get('retCode') == 0:
                        # Get the executed price from order details
                        order_result = order_detail.get('result', {}).get('list', [{}])[0]
                        if float(order_result.get('execQty', 0)) > 0:
                            actual_entry_price = float(order_result.get('avgPrice', entry_price))
                    
                    logger.info(f"Short position entered for {symbol}: {position_size} units at {actual_entry_price} (Bybit execution price)")
                    
                    # Get current account balance
                    current_balance = get_current_capital()
                    
                    # Store position information with actual entry price
                    position_data = {
                        'direction': 'short',
                        'entry_price': actual_entry_price,  # Use actual price from Bybit
                        'position_size': position_size,
                        'stop_loss': stop_loss,
                        'take_profit': take_profit,  # Add take profit to position data
                        'order_id': order_id,
                        'entry_time': datetime.now().isoformat(),
                        'last_check_time': datetime.now().isoformat(),
                        'account_balance': current_balance,
                        'order_type': 'Market',
                        'tp_order_type': 'Limit' if use_limit_orders_exit else 'Market'
                    }
                    
                    active_positions[symbol] = position_data
                    
                    # Save position to database for persistence
                    save_position(symbol, position_data)
                    
                    # Set stop loss and take profit using set_trading_stop API if provided
                    if stop_loss or (take_profit and not use_limit_orders_exit):
                        # Add a small delay to ensure position is registered
                        time.sleep(1)
                        
                        # Prepare parameters
                        params = {
                            "category": "linear",
                            "symbol": symbol,
                            "positionIdx": 0  # For one-way position mode
                        }
                        
                        if stop_loss:
                            params["stopLoss"] = str(stop_loss)
                        
                        if take_profit and not use_limit_orders_exit:
                            params["takeProfit"] = str(take_profit)
                        
                        sl_tp_response = client.set_trading_stop(**params)
                        
                        if sl_tp_response.get('retCode') == 0:
                            logger.info(f"Stop loss/take profit set for short position on {symbol}")
                            if stop_loss:
                                logger.info(f"Stop loss: {stop_loss}")
                            if take_profit and not use_limit_orders_exit:
                                logger.info(f"Take profit: {take_profit}")
                            # Track the SL order in pending_orders
                            sl_oid = sl_tp_response.get('result', {}).get('sl_order_id')
                            if sl_oid:
                                sl_order_data = {
                                    'order_id': sl_oid,
                                    'symbol': symbol,
                                    'order_type': 'Market',
                                    'stop_order_type': 'StopLoss',
                                    'side': 'Buy',
                                    'price': float(stop_loss),
                                    'trigger_price': float(stop_loss),
                                    'qty': float(active_positions[symbol]['position_size']),
                                    'status': 'New',
                                    'created_time': datetime.now().isoformat(),
                                    'purpose': 'exit',
                                    'related_position_symbol': symbol,
                                    'is_stop_loss': True
                                }
                                pending_orders[sl_oid] = sl_order_data
                                save_pending_order(sl_order_data)
                                logger.info(f"SL order {sl_oid} added to pending orders tracking for {symbol}")

                            # Track the native TP trigger order in pending_orders (if not using limit TP)
                            tp_oid = sl_tp_response.get('result', {}).get('tp_order_id')
                            if tp_oid and take_profit and not use_limit_orders_exit:
                                tp_order_data = {
                                    'order_id': tp_oid,
                                    'symbol': symbol,
                                    'order_type': 'Market',
                                    'stop_order_type': 'TakeProfit',
                                    'side': 'Buy',
                                    'price': float(take_profit),
                                    'trigger_price': float(take_profit),
                                    'qty': float(active_positions[symbol]['position_size']),
                                    'status': 'New',
                                    'created_time': datetime.now().isoformat(),
                                    'purpose': 'exit',
                                    'related_position_symbol': symbol,
                                    'is_stop_loss': False
                                }
                                pending_orders[tp_oid] = tp_order_data
                                save_pending_order(tp_order_data)
                                active_positions[symbol]['tp_order_id'] = tp_oid
                                logger.info(f"Native TP order {tp_oid} added to pending orders tracking for {symbol}")

                            # Update position in database
                            save_position(symbol, active_positions[symbol])
                        else:
                            logger.error(f"Failed to set stop loss/take profit using trading stop API: {sl_tp_response}")
                            
                            # Send notification for error
                            if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                                message = f"Failed to set "
                                if stop_loss and take_profit:
                                    message += "stop loss and take profit"
                                elif stop_loss:
                                    message += "stop loss"
                                else:
                                    message += "take profit"
                                
                                send_telegram_message(f"⚠️ <b>WARNING</b>\n"
                                                    f"{message} for {symbol} short position")
                    
                    # Place take profit limit order if requested
                    if take_profit and use_limit_orders_exit:
                        # Check for existing TP orders before placing new ones
                        if has_existing_tp_order(symbol):
                            logger.info(f"Skipping TP order placement - existing TP order found for {symbol}")
                        else:
                            try:
                                # Get exact position size from Bybit
                                exact_position_size = get_exact_position_size_from_exchange(client, symbol)
                                if exact_position_size is not None:
                                    position_size = exact_position_size
                                    logger.info(f"Using exact position size from Bybit for TP: {position_size}")
                                
                                # Use exact position size for TP orders to avoid rounding issues
                                tp_formatted_qty = str(position_size)
                                
                                # Place limit order for take profit
                                tp_response = client.place_order(
                                    category="linear",
                                    symbol=symbol,
                                    side="Buy",  # Buy to close short position
                                    orderType="Limit",
                                    qty=tp_formatted_qty,
                                    price=str(take_profit),
                                    timeInForce="GTC",
                                    reduceOnly=True,
                                    closeOnTrigger=True  # Added closeOnTrigger parameter
                                )
                                
                                if tp_response.get('retCode') == 0:
                                    tp_order_id = tp_response.get('result', {}).get('orderId')
                                    logger.info(f"Take profit limit order placed for {symbol} at {take_profit}")
                                    
                                    # Create pending order data for TP
                                    tp_order_data = {
                                        'order_id': tp_order_id,
                                        'symbol': symbol,
                                        'order_type': 'Limit',
                                        'side': 'Buy',
                                        'price': float(take_profit),
                                        'qty': float(position_size),
                                        'status': 'New',
                                        'created_time': datetime.now().isoformat(),
                                        'purpose': 'exit',
                                        'related_position_symbol': symbol,
                                        'is_stop_loss': False
                                    }
                                    
                                    # Save to pending orders database
                                    save_pending_order(tp_order_data)
                                    
                                    # Update position data with TP order ID
                                    active_positions[symbol]['tp_order_id'] = tp_order_id
                                    save_position(symbol, active_positions[symbol])
                                    
                                    # Force save the tp_order_id through the standard save path too
                                    save_position(symbol, active_positions[symbol])
                                    
                                    # Start a thread to check and adjust TP size after a delay
                                    tp_adjustment_check_seconds = int(config['TRADING'].get('tp_adjustment_check_seconds', 10))
                                    if tp_adjustment_check_seconds > 0:
                                        threading.Thread(target=check_and_adjust_tp_size, args=(symbol, tp_adjustment_check_seconds), daemon=True).start()
                                        logger.info(f"Scheduled TP size adjustment check for {symbol} in {tp_adjustment_check_seconds} seconds")
                                else:
                                    logger.error(f"Failed to place take profit limit order: {tp_response}")
                            except Exception as tp_error:
                                logger.error(f"Error placing take profit limit order: {str(tp_error)}")
                    
                    # Send notification for entry
                    if config.getboolean('NOTIFICATIONS', 'notify_entries', fallback=True):
                        if stop_loss:
                            taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100
                            price_risk = (stop_loss - actual_entry_price) * position_size
                            entry_fee = taker_fee_percent * actual_entry_price * position_size
                            exit_fee = taker_fee_percent * stop_loss * position_size
                            fee_cost = entry_fee + exit_fee
                            total_risk = price_risk + fee_cost
                            risk_message = f"Risk: ${total_risk:.2f} (Price: ${price_risk:.2f}, Fees: ${fee_cost:.2f})"
                        else:
                            risk_message = "Risk: N/A (No stop loss set)"
                        
                        # Add take profit to message if provided
                        take_profit_message = f"Take Profit: {take_profit}" if take_profit else "Take Profit: None"
                        if take_profit and use_limit_orders_exit:
                            take_profit_message += " (Limit Order)"
                        take_profit_message += "\n"
                        
                        send_telegram_message(f"🔴 <b>SHORT ENTRY</b>\n"
                                            f"Symbol: {symbol}\n"
                                            f"Size: {position_size}\n"
                                            f"Price: {actual_entry_price}\n"
                                            f"Stop Loss: {stop_loss if stop_loss else 'None'}\n"
                                            f"{take_profit_message}"
                                            f"{risk_message}")
                    
                    return True
                else:
                    if response.get('retCode') == 10001 and "qty invalid" in str(response.get('retMsg', '')).lower():
                        # Specific error handling for quantity formatting issues
                        logger.error(f"Order quantity rejected by Bybit: {response.get('retMsg')}. " 
                                    f"Original quantity: {position_size}, Formatted quantity: {formatted_qty}. "
                                    f"Symbol info qty_step: {symbol_info.get('lotSizeFilter', {}).get('qtyStep', 'unknown')}")
                        
                        # Send more detailed notification for troubleshooting
                        if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                            send_telegram_message(f"❌ <b>ERROR</b>\n"
                                                f"Order quantity rejected for {symbol}:\n"
                                                f"Issue: {response.get('retMsg', 'Unknown error')}\n"
                                                f"Original qty: {position_size}\n"
                                                f"Formatted qty: {formatted_qty}")
                    else:
                        logger.error(f"Failed to enter short position: {response}")
                        
                        # Send notification for error
                        if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                            send_telegram_message(f"❌ <b>ERROR</b>\n"
                                                f"Failed to enter short position for {symbol}:\n"
                                                f"{response.get('retMsg', 'Unknown error')}")
                    
                    return False
                
        elif action.startswith("EXIT LONG"):
            if symbol in active_positions and active_positions[symbol]['direction'] == 'long':
                # Cancel any related TP orders before exiting
                cancel_related_tp_orders(client, symbol)
                
                # Get exact position size directly from Bybit for accuracy
                exact_position_size = get_exact_position_size_from_exchange(client, symbol)
                
                if exact_position_size is None:
                    logger.warning(f"Could not get exact position size from Bybit for {symbol}, using local tracking")
                    position_size = active_positions[symbol]['position_size']
                else:
                    position_size = exact_position_size
                    logger.info(f"Using exact position size from Bybit: {position_size}")
                
                # Determine if we should use limit order for exit
                if use_limit_orders_exit:
                    # Format limit price (slightly above market price for sells)
                    formatted_price = format_price(entry_price, symbol_info, is_buy=False)
                    
                    # Calculate expiry time
                    expiry_time = (datetime.now() + timedelta(minutes=limit_order_timeout_mins)).isoformat()
                    
                    # Place limit order to exit long position using exact size
                    response = client.place_order(
                        category="linear",
                        symbol=symbol,
                        side="Sell",
                        orderType="Limit",
                        qty=str(position_size),
                        price=formatted_price,
                        timeInForce="GTC",
                        reduceOnly=True,
                        closeOnTrigger=True
                    )
                    
                    if response.get('retCode') == 0:
                        order_id = response.get('result', {}).get('orderId')
                        
                        # Create pending order data
                        order_data = {
                            'order_id': order_id,
                            'symbol': symbol,
                            'order_type': 'Limit',
                            'side': 'Sell',
                            'price': float(formatted_price),
                            'qty': float(position_size),
                            'status': 'New',
                            'created_time': datetime.now().isoformat(),
                            'purpose': 'exit',
                            'related_position_symbol': symbol,
                            'expiry_time': expiry_time,
                            'is_take_profit': True  # Flag this as a TP order
                        }
                        
                        # Save to pending orders database
                        save_pending_order(order_data)
                        
                        logger.info(f"Limit order placed to exit LONG position on {symbol}: {position_size} units at {formatted_price}")
                        
                        # Send notification for limit order placement
                        if config.getboolean('NOTIFICATIONS', 'notify_exits', fallback=True):
                            # Get position details for notification
                            entry_price = active_positions[symbol]['entry_price']
                            stop_loss = active_positions[symbol].get('stop_loss')
                            
                            # Calculate estimated P&L
                            estimated_pnl = (float(formatted_price) - entry_price) * position_size
                            
                            send_telegram_message(f"🟢 <b>LONG EXIT LIMIT ORDER PLACED</b>\n"
                                                f"Symbol: {symbol}\n"
                                                f"Size: {position_size}\n"
                                                f"Limit Price: {formatted_price}\n"
                                                f"Entry Price: {entry_price}\n"
                                                f"Estimated P&L: ${estimated_pnl:.2f}\n"
                                                f"Expires: in {limit_order_timeout_mins} mins")
                        
                        return True
                    else:
                        logger.error(f"Failed to place limit order to exit long position: {response}")
                        
                        # Send notification for error
                        if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                            send_telegram_message(f"❌ <b>ERROR</b>\n"
                                                f"Failed to place limit order to exit long position for {symbol}:\n"
                                                f"{response.get('retMsg', 'Unknown error')}")
                        
                        return False
                else:
                    # Place market order to exit long position using exact size
                    response = client.place_order(
                        category="linear",
                        symbol=symbol,
                        side="Sell",
                        orderType="Market",
                        qty=str(position_size),
                        reduceOnly=True,
                        closeOnTrigger=True
                    )
                    
                    if response.get('retCode') == 0:
                        exit_order_id = response.get('result', {}).get('orderId')
                        
                        # Get the actual execution price from Bybit
                        time.sleep(1)  # Small delay to ensure order is processed
                        order_detail = client.get_order_history(
                            category="linear",
                            symbol=symbol,
                            orderId=exit_order_id
                        )
                        
                        actual_exit_price = entry_price  # Default fallback to estimated price
                        if order_detail.get('retCode') == 0:
                            # Get the executed price from order details
                            order_result = order_detail.get('result', {}).get('list', [{}])[0]
                            if float(order_result.get('execQty', 0)) > 0:
                                actual_exit_price = float(order_result.get('avgPrice', entry_price))
                        
                        logger.info(f"Long position exited for {symbol}: {position_size} units at {actual_exit_price} (Bybit execution price)")
                        
                        # Get position details before removing from active positions
                        entry = float(active_positions[symbol]['entry_price'])
                        stop_loss = active_positions[symbol].get('stop_loss')
                        entry_time = active_positions[symbol]['entry_time']
                        order_type = active_positions[symbol].get('order_type', 'Market')
                        
                        # Calculate P&L with actual exit price
                        exit_price = actual_exit_price
                        size = float(position_size)
                        pnl = (exit_price - entry) * size
                        
                        # Record completed trade in history
                        record_completed_trade(
                            symbol=symbol,
                            direction='long',
                            entry_price=entry,
                            exit_price=exit_price,
                            position_size=size,
                            entry_time=entry_time,
                            stop_loss=stop_loss,
                            stopped_out=False,
                            reason="Strategy exit signal",
                            order_id=exit_order_id,
                            order_type=order_type
                        )
                        
                        # Update capital with the P&L
                        update_capital(pnl)
                        
                        logger.info(f"Trade P&L: {pnl}, New capital: {get_current_capital()}")
                        
                        # Remove from active positions
                        del active_positions[symbol]
                        
                        # Remove from database
                        remove_position(symbol)
                        
                        return True
                    else:
                        if response.get('retCode') == 10001 and "qty invalid" in str(response.get('retMsg', '')).lower():
                            # Specific error handling for quantity formatting issues
                            logger.error(f"Order quantity rejected by Bybit: {response.get('retMsg')}. " 
                                        f"Original quantity: {position_size}, Symbol: {symbol}")
                            
                            # Send more detailed notification for troubleshooting
                            if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                                send_telegram_message(f"❌ <b>ERROR</b>\n"
                                                    f"Order quantity rejected for {symbol} exit:\n"
                                                    f"Issue: {response.get('retMsg', 'Unknown error')}\n"
                                                    f"Quantity: {position_size}")
                        else:
                            logger.error(f"Failed to exit long position: {response}")
                            
                            # Send notification for error
                            if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                                send_telegram_message(f"❌ <b>ERROR</b>\n"
                                                    f"Failed to exit long position for {symbol}:\n"
                                                    f"{response.get('retMsg', 'Unknown error')}")
                        
                        return False
            else:
                logger.warning(f"No active long position found for {symbol}")
                
                # Send notification for error
                if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                    send_telegram_message(f"⚠️ <b>WARNING</b>\n"
                                        f"Attempted to exit non-existent long position for {symbol}")
                
                return False
                
        elif action.startswith("EXIT SHORT"):
            if symbol in active_positions and active_positions[symbol]['direction'] == 'short':
                # Cancel any related TP orders before exiting
                cancel_related_tp_orders(client, symbol)
                
                # Get exact position size directly from Bybit for accuracy
                exact_position_size = get_exact_position_size_from_exchange(client, symbol)
                
                if exact_position_size is None:
                    logger.warning(f"Could not get exact position size from Bybit for {symbol}, using local tracking")
                    position_size = active_positions[symbol]['position_size']
                else:
                    position_size = exact_position_size
                    logger.info(f"Using exact position size from Bybit: {position_size}")
                
                # Determine if we should use limit order for exit
                if use_limit_orders_exit:
                    # Format limit price (slightly below market price for buys)
                    formatted_price = format_price(entry_price, symbol_info, is_buy=True)
                    
                    # Calculate expiry time
                    expiry_time = (datetime.now() + timedelta(minutes=limit_order_timeout_mins)).isoformat()
                    
                    # Place limit order to exit short position using exact size
                    response = client.place_order(
                        category="linear",
                        symbol=symbol,
                        side="Buy",
                        orderType="Limit",
                        qty=str(position_size),
                        price=formatted_price,
                        timeInForce="GTC",
                        reduceOnly=True,
                        closeOnTrigger=True
                    )
                    
                    if response.get('retCode') == 0:
                        order_id = response.get('result', {}).get('orderId')
                        
                        # Create pending order data
                        order_data = {
                            'order_id': order_id,
                            'symbol': symbol,
                            'order_type': 'Limit',
                            'side': 'Buy',
                            'price': float(formatted_price),
                            'qty': float(position_size),
                            'status': 'New',
                            'created_time': datetime.now().isoformat(),
                            'purpose': 'exit',
                            'related_position_symbol': symbol,
                            'expiry_time': expiry_time,
                            'is_take_profit': True  # Flag this as a TP order
                        }
                        
                        # Save to pending orders database
                        save_pending_order(order_data)
                        
                        logger.info(f"Limit order placed to exit SHORT position on {symbol}: {position_size} units at {formatted_price}")
                        
                        # Send notification for limit order placement
                        if config.getboolean('NOTIFICATIONS', 'notify_exits', fallback=True):
                            # Get position details for notification
                            entry_price = active_positions[symbol]['entry_price']
                            stop_loss = active_positions[symbol].get('stop_loss')
                            
                            # Calculate estimated P&L
                            estimated_pnl = (entry_price - float(formatted_price)) * position_size
                            
                            send_telegram_message(f"🔴 <b>SHORT EXIT LIMIT ORDER PLACED</b>\n"
                                                f"Symbol: {symbol}\n"
                                                f"Size: {position_size}\n"
                                                f"Limit Price: {formatted_price}\n"
                                                f"Entry Price: {entry_price}\n"
                                                f"Estimated P&L: ${estimated_pnl:.2f}\n"
                                                f"Expires: in {limit_order_timeout_mins} mins")
                        
                        return True
                    else:
                        logger.error(f"Failed to place limit order to exit short position: {response}")
                        
                        # Send notification for error
                        if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                            send_telegram_message(f"❌ <b>ERROR</b>\n"
                                                f"Failed to place limit order to exit short position for {symbol}:\n"
                                                f"{response.get('retMsg', 'Unknown error')}")
                        
                        return False
                else:
                    # Place market order to exit short position using exact size
                    response = client.place_order(
                        category="linear",
                        symbol=symbol,
                        side="Buy",
                        orderType="Market",
                        qty=str(position_size),
                        reduceOnly=True,
                        closeOnTrigger=True
                    )
                    
                    if response.get('retCode') == 0:
                        exit_order_id = response.get('result', {}).get('orderId')
                        
                        # Get the actual execution price from Bybit
                        time.sleep(1)  # Small delay to ensure order is processed
                        order_detail = client.get_order_history(
                            category="linear",
                            symbol=symbol,
                            orderId=exit_order_id
                        )
                        
                        actual_exit_price = entry_price  # Default fallback to estimated price
                        if order_detail.get('retCode') == 0:
                            # Get the executed price from order details
                            order_result = order_detail.get('result', {}).get('list', [{}])[0]
                            if float(order_result.get('execQty', 0)) > 0:
                                actual_exit_price = float(order_result.get('avgPrice', entry_price))
                        
                        logger.info(f"Short position exited for {symbol}: {position_size} units at {actual_exit_price} (Bybit execution price)")
                        
                        # Get position details before removing from active positions
                        entry = float(active_positions[symbol]['entry_price'])
                        stop_loss = active_positions[symbol].get('stop_loss')
                        entry_time = active_positions[symbol]['entry_time']
                        order_type = active_positions[symbol].get('order_type', 'Market')
                        
                        # Calculate P&L with actual exit price
                        exit_price = actual_exit_price
                        size = float(position_size)
                        pnl = (entry - exit_price) * size
                        
                        # Record completed trade in history
                        record_completed_trade(
                            symbol=symbol,
                            direction='short',
                            entry_price=entry,
                            exit_price=exit_price,
                            position_size=size,
                            entry_time=entry_time,
                            stop_loss=stop_loss,
                            stopped_out=False,
                            reason="Strategy exit signal",
                            order_id=exit_order_id,
                            order_type=order_type
                        )
                        
                        # Update capital with the P&L
                        update_capital(pnl)
                        
                        logger.info(f"Trade P&L: {pnl}, New capital: {get_current_capital()}")
                        
                        # Remove from active positions
                        del active_positions[symbol]
                        
                        # Remove from database
                        remove_position(symbol)
                        
                        return True
                    else:
                        if response.get('retCode') == 10001 and "qty invalid" in str(response.get('retMsg', '')).lower():
                            # Specific error handling for quantity formatting issues
                            logger.error(f"Order quantity rejected by Bybit: {response.get('retMsg')}. " 
                                        f"Original quantity: {position_size}, Symbol: {symbol}")
                            
                            # Send more detailed notification for troubleshooting
                            if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                                send_telegram_message(f"❌ <b>ERROR</b>\n"
                                                    f"Order quantity rejected for {symbol} exit:\n"
                                                    f"Issue: {response.get('retMsg', 'Unknown error')}\n"
                                                    f"Quantity: {position_size}")
                        else:
                            logger.error(f"Failed to exit short position: {response}")
                            
                            # Send notification for error
                            if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                                send_telegram_message(f"❌ <b>ERROR</b>\n"
                                                    f"Failed to exit short position for {symbol}:\n"
                                                    f"{response.get('retMsg', 'Unknown error')}")
                        
                        return False
            else:
                logger.warning(f"No active short position found for {symbol}")
                
                # Send notification for error
                if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                    send_telegram_message(f"⚠️ <b>WARNING</b>\n"
                                        f"Attempted to exit non-existent short position for {symbol}")
                
                return False
        
        else:
            logger.warning(f"Unknown action: {action}")
            
            # Send notification for error
            if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                send_telegram_message(f"❌ <b>ERROR</b>\n"
                                    f"Unknown action received: {action}")
            
            return False
            
    except Exception as e:
        logger.error(f"Error executing trade: {str(e)}")
        
        # Send notification for error
        if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
            send_telegram_message(f"❌ <b>ERROR</b>\n"
                                f"Error executing trade: {str(e)}")
        
        return False

# Helper function to cancel order by ID
def cancel_order(client, order_id, symbol):
    """Cancel an order by ID"""
    try:
        response = client.cancel_order(
            category="linear",
            symbol=symbol,
            orderId=order_id
        )
        
        if response.get('retCode') == 0:
            logger.info(f"Successfully canceled order {order_id} for {symbol}")
            return True
        else:
            logger.error(f"Failed to cancel order {order_id}: {response}")
            return False
    except Exception as e:
        logger.error(f"Error canceling order: {str(e)}")
        return False


def cancel_related_tp_orders(client, symbol, max_retries=3):
    """Cancel any limit TP orders related to the given symbol with improved checking"""
    try:
        orders_cancelled = 0
        failed_cancellations = []
        
        # Get all related orders for this symbol
        related_orders = []
        for order_id, order_data in list(pending_orders.items()):
            if (order_data.get('purpose') == 'exit' and 
                order_data.get('related_position_symbol') == symbol):
                related_orders.append((order_id, order_data))
        
        if not related_orders:
            logger.info(f"No related TP orders found for {symbol}")
            return 0
            
        logger.info(f"Found {len(related_orders)} related orders for {symbol}, checking status before cancellation")
        
        # Check status of each order before attempting to cancel
        for order_id, order_data in related_orders:
            # Check current order status
            order_status = check_pending_order_status(client, order_id, order_data)
            
            # Skip if order doesn't exist or is already completed/cancelled
            if not order_status:
                logger.info(f"Order {order_id} not found, skipping cancellation")
                # Remove from pending orders since it doesn't exist
                if order_id in pending_orders:
                    remove_pending_order(order_id)
                continue
                
            if order_status['status'] in ['Filled', 'Cancelled', 'Rejected']:
                logger.info(f"Order {order_id} already has status {order_status['status']}, skipping cancellation")
                # Remove from pending orders since it's already completed
                if order_id in pending_orders:
                    remove_pending_order(order_id)
                continue
            
            # IMPORTANT: Skip stop loss orders - only cancel take profit orders
            if order_status.get('is_stop_loss', False) or order_data.get('is_stop_loss', False):
                logger.info(f"Order {order_id} is a stop loss order, skipping cancellation")
                continue
                
            # Check if this is a take profit order — covers both:
            # 1. Limit TP orders (limit exit orders)
            # 2. Native TP trigger orders (Market type with stop_order_type = 'TakeProfit')
            is_take_profit = False

            # Check from API status first
            api_order_type = order_status.get('order_type', '')
            api_stop_order_type = order_status.get('stop_order_type', '')

            if api_order_type == 'Limit' and api_stop_order_type == '' and bool(order_status.get('reduceOnly', False)):
                is_take_profit = True  # Standard limit TP order
            elif api_order_type == 'Market' and api_stop_order_type == 'TakeProfit':
                is_take_profit = True  # Native TP trigger order

            # If not identified from API status, check the stored order data
            if not is_take_profit:
                stored_order_type = order_data.get('order_type', '')
                stored_stop_order_type = order_data.get('stop_order_type', '')
                if (not order_data.get('is_stop_loss', False) and
                        order_data.get('purpose') == 'exit'):
                    if stored_order_type == 'Limit':
                        is_take_profit = True  # Standard limit TP order
                    elif stored_order_type == 'Market' and stored_stop_order_type == 'TakeProfit':
                        is_take_profit = True  # Native TP trigger order

            # Skip if not a take profit order
            if not is_take_profit:
                logger.info(f"Order {order_id} is not a take profit order, skipping cancellation")
                continue
            
            # If we get here, the order is a take profit limit order that's still active and needs cancellation
            logger.info(f"Found active related TP order {order_id} for {symbol}, cancelling")
            
            # Try to cancel with retries
            cancel_success = False
            retry_count = 0
            
            while not cancel_success and retry_count < max_retries:
                try:
                    # Cancel the order on Bybit
                    result = cancel_order(client, order_id, symbol)
                    
                    if result:
                        # Remove from pending orders
                        conn = sqlite3.connect('trading_bot.db')
                        cursor = conn.cursor()
                        
                        try:
                            # Start a transaction
                            cursor.execute('BEGIN TRANSACTION')
                            
                            # Remove from database first
                            cursor.execute('DELETE FROM pending_orders WHERE order_id = ?', (order_id,))
                            
                            # Remove from memory only if database update succeeds
                            if order_id in pending_orders:
                                del pending_orders[order_id]
                            
                            # Commit the transaction
                            conn.commit()
                            cancel_success = True
                            orders_cancelled += 1
                            logger.info(f"Successfully cancelled TP order {order_id} for {symbol}")
                            
                        except Exception as db_error:
                            # Rollback in case of database error
                            conn.rollback()
                            logger.error(f"Database error during order cancellation: {str(db_error)}")
                            # Don't increment retry count for database errors
                            # We'll try again with the same cancellation result
                        
                        finally:
                            conn.close()
                    else:
                        retry_count += 1
                        if retry_count < max_retries:
                            logger.warning(f"Failed to cancel TP order {order_id}, retrying ({retry_count}/{max_retries})...")
                            time.sleep(1)  # Small delay before retry
                        else:
                            logger.error(f"Failed to cancel TP order {order_id} after {max_retries} attempts")
                            failed_cancellations.append(order_id)
                except Exception as e:
                    retry_count += 1
                    if retry_count < max_retries:
                        logger.warning(f"Error cancelling TP order {order_id}: {str(e)}, retrying ({retry_count}/{max_retries})...")
                        time.sleep(1)  # Small delay before retry
                    else:
                        logger.error(f"Error cancelling TP order {order_id} after {max_retries} attempts: {str(e)}")
                        failed_cancellations.append(order_id)
        
        if orders_cancelled > 0:
            logger.info(f"Cancelled {orders_cancelled} related TP orders for {symbol}")
            
        # Handle failed cancellations by adding to cleanup queue
        if failed_cancellations:
            add_to_cancellation_cleanup_queue(failed_cancellations, symbol)
            
        return orders_cancelled
    except Exception as e:
        logger.error(f"Error in cancel_related_tp_orders: {str(e)}")
        return 0


def remove_related_stop_loss_orders_from_database(symbol):
    """
    Remove any STOP LOSS orders for the given symbol from LOCAL DATABASE ONLY.
    IMPORTANT: 
    - This function does NOT cancel orders on the exchange
    - This function ONLY handles stop loss orders, not limit orders
    - It only cleans up the local tracking database
    """
    try:
        global pending_orders
        orders_removed = 0
        
        # Find all stop loss orders for this symbol
        for order_id, order_data in list(pending_orders.items()):
            # ONLY match stop loss orders for the specified symbol, NOT limit orders
            is_stop_loss = (order_data.get('is_stop_loss', False) or 
                           (order_data.get('order_type', '') == 'Market' and 
                           order_data.get('stop_order_type', '') in ['StopLoss', 'StopMarket']))
            
            if order_data.get('symbol') == symbol and is_stop_loss:
                logger.info(f"LOCAL DATABASE CLEANUP ONLY: Removing stop loss order {order_id} for {symbol} from tracking database")
                
                # Remove from memory
                if order_id in pending_orders:
                    del pending_orders[order_id]
                
                # Remove from database (direct query, no API calls)
                conn = sqlite3.connect('trading_bot.db')
                cursor = conn.cursor()
                try:
                    cursor.execute('DELETE FROM pending_orders WHERE order_id = ?', (order_id,))
                    conn.commit()
                    orders_removed += 1
                    logger.info(f"LOCAL DATABASE CLEANUP ONLY: Successfully removed stop loss order {order_id} from database")
                except Exception as db_error:
                    logger.error(f"Database error when removing stop loss order from database: {str(db_error)}")
                finally:
                    conn.close()
        
        if orders_removed > 0:
            logger.info(f"LOCAL DATABASE CLEANUP: Removed {orders_removed} stop loss orders for {symbol} from local database (NO exchange operations performed)")
        
        return orders_removed
    except Exception as e:
        logger.error(f"Error removing related stop loss orders from database: {str(e)}")
        return 0

def add_to_cancellation_cleanup_queue(order_ids, symbol):
    """Add orders to a cleanup queue for later cancellation attempts"""
    global orders_pending_cancellation
    
    current_time = datetime.now().isoformat()
    for order_id in order_ids:
        if order_id in orders_pending_cancellation:
            # Increment retry count if already in queue
            orders_pending_cancellation[order_id]['retry_count'] += 1
            orders_pending_cancellation[order_id]['last_attempt'] = current_time
        else:
            # Add new entry to queue
            orders_pending_cancellation[order_id] = {
                'symbol': symbol,
                'retry_count': 1,
                'last_attempt': current_time
            }
    
    logger.info(f"Added {len(order_ids)} orders to cancellation cleanup queue for {symbol}")


def process_cancellation_cleanup_queue():
    """Process orders in the cancellation cleanup queue with exponential backoff"""
    global orders_pending_cancellation
    
    if not orders_pending_cancellation:
        return
        
    logger.info(f"Processing cancellation cleanup queue, {len(orders_pending_cancellation)} orders pending")
    current_time = datetime.now()
    client = initialize_bybit_client()
    
    for order_id, data in list(orders_pending_cancellation.items()):
        try:
            # Calculate backoff time based on retry count
            # 5 seconds for first retry, then 15, 45, 2 minutes, 6 minutes, etc.
            backoff_seconds = 5 * (3 ** (min(data['retry_count'] - 1, 5)))
            
            # Check if enough time has passed for retry
            last_attempt = datetime.fromisoformat(data['last_attempt'])
            if (current_time - last_attempt).total_seconds() < backoff_seconds:
                continue  # Skip if not enough time has passed
                
            # Check if order still exists
            order_status = check_pending_order_status(client, order_id)
            
            # If order no longer exists or is already cancelled, remove from queue
            if not order_status or order_status['status'] in ['Cancelled', 'Rejected', 'Filled']:
                logger.info(f"Order {order_id} already {order_status['status'] if order_status else 'gone'}, removing from cleanup queue")
                del orders_pending_cancellation[order_id]
                
                # Also make sure it's removed from pending_orders
                if order_id in pending_orders:
                    remove_pending_order(order_id)
                    
                continue
                
            # Try to cancel again
            symbol = data['symbol']
            result = cancel_order(client, order_id, symbol)
            
            if result:
                logger.info(f"Successfully cancelled order {order_id} from cleanup queue")
                # Remove from pending orders
                if order_id in pending_orders:
                    remove_pending_order(order_id)
                # Remove from cleanup queue
                del orders_pending_cancellation[order_id]
            else:
                # Update retry count and timestamp
                data['retry_count'] += 1
                data['last_attempt'] = current_time.isoformat()
                logger.warning(f"Failed to cancel order {order_id} from cleanup queue, retry count: {data['retry_count']}")
                
        except Exception as e:
            logger.error(f"Error processing cancellation for order {order_id}: {str(e)}")
            # Update retry data even on error
            data['retry_count'] += 1
            data['last_attempt'] = current_time.isoformat()


def has_existing_tp_order(symbol):
    """Check if there's already a take profit order for this symbol"""
    try:
        for order_id, order_data in pending_orders.items():
            if (order_data.get('purpose') == 'exit' and 
                order_data.get('related_position_symbol') == symbol and
                not order_data.get('is_stop_loss', False)):
                return True
        return False
    except Exception as e:
        logger.error(f"Error checking for existing TP orders: {str(e)}")
        return False  # Default to false on error to allow TP placement





def check_and_adjust_tp_size(symbol, delay_seconds=10):
    """
    Check and adjust TP limit order size to match position size after a delay
    
    Args:
        symbol (str): The symbol to check
        delay_seconds (int): Delay in seconds before checking
    """
    try:
        # Sleep for the specified delay
        time.sleep(delay_seconds)
        
        logger.info(f"Running TP size adjustment check for {symbol} after {delay_seconds} seconds delay")
        
        # Check if position still exists
        if symbol not in active_positions:
            logger.info(f"Position for {symbol} no longer exists, skipping TP adjustment")
            return
            
        # Get position data
        position = active_positions[symbol]
        direction = position['direction']
        take_profit = position.get('take_profit')
        tp_order_id = position.get('tp_order_id')
        use_limit_tp = position.get('tp_order_type', 'Market') == 'Limit'
        
        # Skip if no take profit price at all
        if not take_profit:
            logger.info(f"No take profit set for {symbol}, skipping adjustment")
            return

        # If tp_order_type is not 'Limit', scan pending_orders as a fallback
        # (tp_order_type can be lost in a race condition between reconciliation and process_filled_order)
        if not use_limit_tp or not tp_order_id:
            for oid, odata in list(pending_orders.items()):
                if (odata.get('purpose') == 'exit' and
                        odata.get('related_position_symbol') == symbol and
                        not odata.get('is_stop_loss', False)):
                    use_limit_tp = True
                    tp_order_id = oid
                    logger.info(f"TP order found via pending_orders scan: {oid} (tp_order_type was '{position.get('tp_order_type', 'missing')}')")
                    break

        # Skip if still no limit TP found
        if not use_limit_tp:
            logger.info(f"No limit TP order found for {symbol}, skipping adjustment")
            return

        # Skip if no TP order ID after scan
        if not tp_order_id:
            logger.info(f"No TP order ID found for {symbol}, skipping adjustment")
            return
        
        # Initialize Bybit client
        client = initialize_bybit_client()
        
        # Get exact position size from Bybit
        exact_position_size = get_exact_position_size_from_exchange(client, symbol)
        
        if exact_position_size is None:
            logger.warning(f"Could not get exact position size from Bybit for {symbol}, using local size")
            exact_position_size = float(position['position_size'])
        else:
            # Update our local tracking with the real position size
            if abs(exact_position_size - float(position['position_size'])) > 0.0001:
                logger.info(f"Updating local position size for {symbol}: {position['position_size']} → {exact_position_size}")
                position['position_size'] = exact_position_size
                save_position(symbol, position)
        
        # Get TP order size
        tp_order_size = 0
        tp_order_found = False
        
        # First check in pending_orders
        for order_id, order_data in pending_orders.items():
            if order_id == tp_order_id:
                tp_order_size = float(order_data.get('qty', 0))
                tp_order_found = True
                break
                
        # If not found in pending orders, check with Bybit API
        if not tp_order_found:
            try:
                order_status = check_pending_order_status(client, tp_order_id)
                if order_status:
                    tp_order_size = float(order_status.get('qty', 0))
                    tp_order_found = True
                else:
                    # TP order doesn't exist anymore
                    logger.warning(f"TP order {tp_order_id} not found for {symbol}, may need to place a new one")
                    
                    # Check if there are any other exit orders for this symbol
                    other_tp_found = False
                    for oid, order_data in pending_orders.items():
                        if (order_data.get('purpose') == 'exit' and 
                            order_data.get('related_position_symbol') == symbol and
                            not order_data.get('is_stop_loss', False)):
                            logger.info(f"Found another TP order {oid} for {symbol}, using that instead")
                            tp_order_id = oid
                            tp_order_size = float(order_data.get('qty', 0))
                            tp_order_found = True
                            other_tp_found = True
                            
                            # Update position with new TP order ID
                            position['tp_order_id'] = tp_order_id
                            save_position(symbol, position)
                            break
                    
                    # If no TP orders found, may need to create a new one
                    if not other_tp_found:
                        logger.warning(f"No TP orders found for {symbol}, will place a new one")
                        
                        # Get symbol info for proper formatting
                        symbol_info = get_symbol_info_with_retry(client, symbol)
                        
                        if not symbol_info:
                            logger.error(f"Could not get symbol info for {symbol} when placing new TP")
                            return
                            
                        # Use exact position size directly for TP
                        tp_formatted_qty = str(exact_position_size)
                        
                        # Determine the correct side for TP order
                        tp_side = "Sell" if direction == "long" else "Buy"
                        
                        # Place new limit order for take profit
                        tp_response = client.place_order(
                            category="linear",
                            symbol=symbol,
                            side=tp_side,
                            orderType="Limit",
                            qty=tp_formatted_qty,
                            price=str(take_profit),
                            timeInForce="GTC",
                            reduceOnly=True,
                            closeOnTrigger=True
                        )
                        
                        if tp_response.get('retCode') == 0:
                            new_tp_order_id = tp_response.get('result', {}).get('orderId')
                            logger.info(f"New TP limit order placed for {symbol} with size {exact_position_size} at {take_profit}")
                            
                            # Create pending order data for new TP
                            tp_order_data = {
                                'order_id': new_tp_order_id,
                                'symbol': symbol,
                                'order_type': 'Limit',
                                'side': tp_side,
                                'price': float(take_profit),
                                'qty': float(exact_position_size),
                                'status': 'New',
                                'created_time': datetime.now().isoformat(),
                                'purpose': 'exit',
                                'related_position_symbol': symbol,
                                'is_stop_loss': False
                            }
                            
                            # Save to pending orders database
                            save_pending_order(tp_order_data)
                            
                            # Update position data with new TP order ID
                            position['tp_order_id'] = new_tp_order_id
                            save_position(symbol, position)
                            
                            # Send notification about new TP
                            if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                                send_telegram_message(f"🔄 <b>NEW TP ORDER PLACED</b>\n"
                                                   f"Symbol: {symbol}\n"
                                                   f"Direction: {direction.upper()}\n"
                                                   f"Size: {exact_position_size}\n"
                                                   f"Take Profit: {take_profit}")
                                                   
                            # No need to continue with adjustment since we just created a new TP with the correct size
                            return
                        else:
                            logger.error(f"Failed to place new TP limit order: {tp_response}")
                            return
            except Exception as e:
                logger.error(f"Error checking TP order status: {str(e)}")
                return
        
        # If TP order not found and couldn't create a new one, exit
        if not tp_order_found:
            logger.warning(f"TP order not found and couldn't create a new one for {symbol}")
            return
                
        # Compare sizes and adjust if needed
        if abs(exact_position_size - tp_order_size) > 0.0001:  # Small threshold to account for floating point precision
            logger.warning(f"TP order size mismatch for {symbol}: TP size {tp_order_size}, Position size {exact_position_size} - adjusting")
            
            # Cancel existing TP order
            cancel_success = cancel_order(client, tp_order_id, symbol)
            if not cancel_success:
                logger.error(f"Failed to cancel existing TP order {tp_order_id} for {symbol}")
                return
                
            # Remove from pending orders
            if tp_order_id in pending_orders:
                remove_pending_order(tp_order_id)
                
            # Get symbol info for proper formatting
            symbol_info = get_symbol_info_with_retry(client, symbol)
            
            if not symbol_info:
                logger.error(f"Could not get symbol info for {symbol} when adjusting TP")
                return
                
            # Use exact position size directly for TP
            tp_formatted_qty = str(exact_position_size)
            
            # Determine the correct side for TP order
            tp_side = "Sell" if direction == "long" else "Buy"
            
            # Place new limit order for take profit
            tp_response = client.place_order(
                category="linear",
                symbol=symbol,
                side=tp_side,
                orderType="Limit",
                qty=tp_formatted_qty,
                price=str(take_profit),
                timeInForce="GTC",
                reduceOnly=True,
                closeOnTrigger=True
            )
            
            if tp_response.get('retCode') == 0:
                new_tp_order_id = tp_response.get('result', {}).get('orderId')
                logger.info(f"New TP limit order placed for {symbol} with adjusted size {exact_position_size} at {take_profit}")
                
                # Create pending order data for new TP
                tp_order_data = {
                    'order_id': new_tp_order_id,
                    'symbol': symbol,
                    'order_type': 'Limit',
                    'side': tp_side,
                    'price': float(take_profit),
                    'qty': float(exact_position_size),
                    'status': 'New',
                    'created_time': datetime.now().isoformat(),
                    'purpose': 'exit',
                    'related_position_symbol': symbol,
                    'is_stop_loss': False
                }
                
                # Save to pending orders database
                save_pending_order(tp_order_data)
                
                # Update position data with new TP order ID
                position['tp_order_id'] = new_tp_order_id
                save_position(symbol, position)
                
                # Send notification about adjusted TP
                if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                    send_telegram_message(f"🔄 <b>TP SIZE ADJUSTED</b>\n"
                                       f"Symbol: {symbol}\n"
                                       f"Direction: {direction.upper()}\n"
                                       f"Old Size: {tp_order_size}\n"
                                       f"New Size: {exact_position_size}\n"
                                       f"Take Profit: {take_profit}")
            else:
                logger.error(f"Failed to place adjusted TP limit order: {tp_response}")
        else:
            logger.info(f"TP size check for {symbol}: TP size {tp_order_size} matches position size {exact_position_size} - no adjustment needed")
            
    except Exception as e:
        logger.error(f"Error in check_and_adjust_tp_size for {symbol}: {str(e)}")



# IMPROVEMENT: Enhanced Position Reconciliation with minutely checks
# IMPROVEMENT: Enhanced Position Reconciliation with minutely checks
def enhanced_reconcile_positions():
    """Enhanced version of reconcile_positions with better handling of position changes"""
    global previous_positions, active_positions
    
    logger.info("Starting enhanced position reconciliation...")
    try:
        client = initialize_bybit_client()  # Will reuse existing client
        
        # Store current positions for comparison before updating
        if not previous_positions:
            previous_positions = active_positions.copy()
            
        # Get positions from database
        local_positions = load_positions()
        
        # Update active_positions from database
        active_positions = local_positions.copy()
        
        # Capture initial capital for comparison
        initial_capital = get_current_capital()
        
        # Get actual positions from Bybit
        bybit_positions = {}
        try:
            # Try different methods to get positions with proper error handling
            try:
                # First attempt - with settleCoin parameter
                response = client.get_positions(category="linear", settleCoin="USDT")
                if response.get('retCode') != 0:
                    logger.warning(f"Failed to get positions with settleCoin parameter: {response}")
                    raise Exception("Failed with settleCoin parameter")
            except Exception as e1:
                logger.warning(f"Trying alternative method to get positions: {str(e1)}")
                try:
                    # Second attempt - get position info without specific parameters
                    response = client.get_position_info(category="linear")
                except Exception as e2:
                    logger.warning(f"Second attempt failed: {str(e2)}")
                    # Third attempt - try with specific symbols if you know them
                    default_symbol = config['TRADING'].get('default_symbol', 'BTCUSDT')
                    response = client.get_positions(category="linear", symbol=default_symbol)
            
            # Process positions regardless of which method worked
            if response.get('retCode') == 0:
                positions_list = response.get('result', {}).get('list', [])
                for pos in positions_list:
                    if float(pos.get('size', 0)) > 0:
                        symbol = pos.get('symbol')
                        bybit_positions[symbol] = {
                            'direction': 'long' if pos.get('side') == 'Buy' else 'short',
                            'position_size': float(pos.get('size', 0)),
                            'entry_price': float(pos.get('avgPrice', 0)),
                            'unrealized_pnl': float(pos.get('unrealisedPnl', 0)),
                            'stop_loss': float(pos.get('stopLoss', 0)) if pos.get('stopLoss') else None,
                            'take_profit': float(pos.get('takeProfit', 0)) if pos.get('takeProfit') else None
                        }
            else:
                logger.error(f"Failed to get positions from Bybit: {response}")
                
                # Consider resetting client on severe errors
                if response.get('retCode') in [10002, 10003, 10004]:  # Auth error codes
                    reset_global_client()
                
                # Send notification for error
                if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                    send_telegram_message(f"❌ <b>SYSTEM ERROR</b>\n"
                                        f"Failed to get positions from Bybit during reconciliation")
                return
                
        except Exception as e:
            logger.error(f"Error retrieving positions from Bybit: {str(e)}")
            
            # Consider resetting client on connection errors
            if isinstance(e, ConnectionError) or isinstance(e, TimeoutError):
                reset_global_client()
            
            # Send notification for error
            if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                send_telegram_message(f"❌ <b>SYSTEM ERROR</b>\n"
                                    f"Error retrieving positions from Bybit: {str(e)}")
            return
        
        # Get current prices for all symbols for P&L calculations
        current_prices = {}
        for symbol in set(list(local_positions.keys()) + list(bybit_positions.keys())):
            try:
                ticker_response = client.get_tickers(category="linear", symbol=symbol)
                if ticker_response.get('retCode') == 0:
                    current_prices[symbol] = float(ticker_response.get('result', {}).get('list', [{}])[0].get('lastPrice', 0))
            except Exception as e:
                logger.warning(f"Failed to get current price for {symbol}: {str(e)}")
        
        # PART 1: Reconcile differences - find positions on Bybit not tracked locally
        for symbol, bybit_pos in bybit_positions.items():
            if symbol not in local_positions:
                logger.warning(f"Found position on Bybit not tracked locally: {symbol}")
                
                # Add to local tracking with estimated entry time
                bybit_pos['entry_time'] = datetime.now().isoformat()
                bybit_pos['order_id'] = 'reconciled'
                bybit_pos['last_check_time'] = datetime.now().isoformat()
                bybit_pos['order_type'] = 'Market'  # Assume market order for reconciled positions
                save_position(symbol, bybit_pos)
                active_positions[symbol] = bybit_pos
                
                # Log the reconciliation event
                log_reconciliation_event(
                    "NEW_POSITION_FOUND", 
                    symbol, 
                    f"Found untracked {bybit_pos['direction'].upper()} position",
                    bybit_pos
                )
                
                # Send notification for reconciliation
                if config.getboolean('NOTIFICATIONS', 'notify_reconciliation', fallback=False):
                    send_telegram_message(f"🔄 <b>POSITION RECONCILIATION</b>\n"
                                        f"Found untracked {bybit_pos['direction'].upper()} position for {symbol}\n"
                                        f"Size: {bybit_pos['position_size']}\n"
                                        f"Entry: {bybit_pos['entry_price']}")
            else:
                # Position exists both locally and on Bybit
                
                # Update last check time
                update_position_check_time(symbol)
                
                # Update position size and other details if different
                local_pos = local_positions[symbol]
                size_changed = abs(bybit_pos['position_size'] - local_pos['position_size']) > 0.0001
                sl_changed = (bybit_pos['stop_loss'] != local_pos.get('stop_loss'))
                
                # PART 1.5: Check if position size has changed (partial fill completed)
                if size_changed:
                    if bybit_pos['position_size'] > local_pos['position_size']:
                        # Position size has increased
                        logger.warning(f"Position size increased for {symbol}: {local_pos['position_size']} → {bybit_pos['position_size']} - updating TP orders")
                        
                        # Update the local position size
                        local_pos['position_size'] = bybit_pos['position_size']
                        
                        # First, cancel any existing TP orders
                        tp_orders_canceled = cancel_related_tp_orders(client, symbol)
                        if tp_orders_canceled > 0:
                            logger.info(f"Canceled {tp_orders_canceled} existing TP orders for {symbol}")
                        
                        # Check if we should use limit TP orders
                        use_limit_tp = local_pos.get('tp_order_type', 'Market') == 'Limit'
                        take_profit = local_pos.get('take_profit')
                        
                        if take_profit and use_limit_tp:
                            # Place new TP limit order with correct full size
                            try:
                                # Get symbol info for proper formatting
                                symbol_info = get_symbol_info_with_retry(client, symbol)
                                
                                if symbol_info:
                                    # Format quantity for TP order using the new full size
                                    tp_formatted_qty = format_position_quantity(bybit_pos['position_size'], symbol_info)
                                    
                                    # Determine the correct side for TP order
                                    direction = local_pos['direction']
                                    tp_side = "Sell" if direction == "long" else "Buy"
                                    
                                    # Place limit order for take profit
                                    tp_response = client.place_order(
                                        category="linear",
                                        symbol=symbol,
                                        side=tp_side,
                                        orderType="Limit",
                                        qty=tp_formatted_qty,
                                        price=str(take_profit),
                                        timeInForce="GTC",
                                        reduceOnly=True,
                                        closeOnTrigger=True
                                    )
                                    
                                    if tp_response.get('retCode') == 0:
                                        tp_order_id = tp_response.get('result', {}).get('orderId')
                                        logger.info(f"New take profit limit order placed for {symbol} with full size {bybit_pos['position_size']} at {take_profit}")
                                        
                                        # Create pending order data for TP
                                        tp_order_data = {
                                            'order_id': tp_order_id,
                                            'symbol': symbol,
                                            'order_type': 'Limit',
                                            'side': tp_side,
                                            'price': float(take_profit),
                                            'qty': float(tp_formatted_qty),
                                            'status': 'New',
                                            'created_time': datetime.now().isoformat(),
                                            'purpose': 'exit',
                                            'related_position_symbol': symbol,
                                            'is_stop_loss': False
                                        }
                                        
                                        # Save to pending orders database
                                        save_pending_order(tp_order_data)
                                        
                                        # Update position data with new TP order ID
                                        local_pos['tp_order_id'] = tp_order_id
                                        
                                        # Send notification about updated TP
                                        if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                                            send_telegram_message(f"🔄 <b>POSITION TP UPDATED</b>\n"
                                                                f"Symbol: {symbol}\n"
                                                                f"Direction: {direction.upper()}\n"
                                                                f"New Size: {bybit_pos['position_size']}\n"
                                                                f"Take Profit: {take_profit}")
                                    else:
                                        logger.error(f"Failed to place updated take profit limit order: {tp_response}")
                                else:
                                    logger.error(f"Could not get symbol info for {symbol} when updating TP")
                            except Exception as e:
                                logger.error(f"Error updating TP limit order: {str(e)}")
                    
                    elif bybit_pos['position_size'] < local_pos['position_size']:
                        # Position size has decreased - likely a partial TP fill
                        logger.warning(f"Position size decreased for {symbol}: {local_pos['position_size']} → {bybit_pos['position_size']} - possible partial TP fill")
                        
                        # Update the local position size
                        local_pos['position_size'] = bybit_pos['position_size']
                        
                        # Schedule a TP size check to ensure TP order size matches the new position size
                        tp_adjustment_check_seconds = int(config['TRADING'].get('tp_adjustment_check_seconds', 10))
                        if tp_adjustment_check_seconds > 0:
                            threading.Thread(target=check_and_adjust_tp_size, args=(symbol, 5), daemon=True).start()
                            logger.info(f"Scheduled TP size adjustment check for {symbol} after reconciliation detected reduced position size")
                        
                        # Log the reconciliation event
                        log_reconciliation_event(
                            "POSITION_SIZE_DECREASED", 
                            symbol, 
                            f"Position size decreased: {local_pos['position_size']} → {bybit_pos['position_size']}, adjusting TP",
                            local_pos
                        )
                        
                        # Send notification for position size change
                        if config.getboolean('NOTIFICATIONS', 'notify_reconciliation', fallback=False):
                            send_telegram_message(f"🔄 <b>POSITION SIZE REDUCED</b>\n"
                                               f"Symbol: {symbol}\n"
                                               f"Direction: {local_pos['direction'].upper()}\n"
                                               f"Old Size: {local_pos['position_size']}\n"
                                               f"New Size: {bybit_pos['position_size']}\n"
                                               f"Adjusting TP order...")
                
                # Regular position detail updates (for any changes including size)
                if size_changed or sl_changed:
                    changes = []
                    if size_changed:
                        changes.append(f"Size: {local_pos['position_size']} → {bybit_pos['position_size']}")
                        local_pos['position_size'] = bybit_pos['position_size']
                    
                    if sl_changed and bybit_pos['stop_loss']:
                        changes.append(f"Stop Loss: {local_pos.get('stop_loss')} → {bybit_pos['stop_loss']}")
                        local_pos['stop_loss'] = bybit_pos['stop_loss']
                    
                    if bybit_pos['take_profit']:
                        changes.append(f"Take Profit: {local_pos.get('take_profit', 'None')} → {bybit_pos['take_profit']}")
                        local_pos['take_profit'] = bybit_pos['take_profit']
                    
                    # Save updated position
                    local_pos['last_check_time'] = datetime.now().isoformat()
                    save_position(symbol, local_pos)
                    active_positions[symbol] = local_pos
                    
                    change_msg = ", ".join(changes)
                    logger.warning(f"Position details changed for {symbol}: {change_msg}")
                    
                    # Log the reconciliation event
                    log_reconciliation_event(
                        "POSITION_CHANGED", 
                        symbol, 
                        f"Position details updated: {change_msg}",
                        local_pos
                    )
                    
                    # Send notification for significant changes
                    if size_changed and config.getboolean('NOTIFICATIONS', 'notify_reconciliation', fallback=False):
                        send_telegram_message(f"🔄 <b>POSITION RECONCILIATION</b>\n"
                                            f"Updated {symbol} {local_pos['direction'].upper()} position\n"
                                            f"{change_msg}")
        
        # PART 2: Check for closed positions still tracked locally
        positions_closed = False
        for symbol in list(local_positions.keys()):
            if symbol not in bybit_positions:
                # Check if this position was recently closed via webhook
                current_time = datetime.now()
                if (symbol in recently_closed_positions and 
                    (current_time - recently_closed_positions[symbol]).total_seconds() < RECONCILIATION_COOLING_PERIOD):
                    logger.info(f"Skipping reconciliation for recently closed position: {symbol}")
                    continue
                
                positions_closed = True
                logger.warning(f"Position {symbol} tracked locally but not found on Bybit, determining closure reason")
                
                local_pos = local_positions[symbol]
                direction = local_pos['direction']
                entry_price = local_pos['entry_price']
                position_size = local_pos['position_size']
                entry_time = local_pos['entry_time']
                stop_loss = local_pos.get('stop_loss')
                order_type = local_pos.get('order_type', 'Market')  # Default to Market if not specified
                
                # Try to get accurate exit price using enhanced function
                accurate_exit_price = get_accurate_bybit_exit_price(client, symbol, None, direction)
                
                # If we got a price, use it directly
                if accurate_exit_price:
                    exit_price = accurate_exit_price
                    logger.info(f"Using actual exit price from Bybit: {exit_price} for {symbol}")
                else:
                    # Only as an absolute last resort use current price
                    exit_price = current_prices.get(symbol, entry_price)
                    logger.warning(f"Could not determine exit price for {symbol}, using current price: {exit_price}")
                
                # Determine if this was likely a stop loss hit
                was_stopped_out = False
                close_reason = "Unknown"
                
                # If we have a stop loss level, check if likely hit
                if stop_loss:
                    if direction == 'long' and exit_price <= stop_loss:
                        was_stopped_out = True
                        close_reason = "Stop loss hit"
                    elif direction == 'short' and exit_price >= stop_loss:
                        was_stopped_out = True
                        close_reason = "Stop loss hit"
                
                # Check if it was a take profit hit
                take_profit = local_pos.get('take_profit')
                if take_profit and symbol in current_prices:
                    if direction == 'long' and take_profit <= current_prices[symbol]:
                        was_stopped_out = False
                        if not accurate_exit_price:  # Only override if we didn't get an accurate price
                            exit_price = take_profit  # Use take profit as exit price
                        close_reason = "Take profit hit"
                    elif direction == 'short' and take_profit >= current_prices[symbol]:
                        was_stopped_out = False
                        if not accurate_exit_price:  # Only override if we didn't get an accurate price
                            exit_price = take_profit  # Use take profit as exit price
                        close_reason = "Take profit hit"
                
                # Check previous positions to see if we sent an exit signal
                if symbol in previous_positions and not was_stopped_out and close_reason == "Unknown":
                    close_reason = "Manual exit or strategy signal"
                
                # If we still don't know, default to "closed on exchange"
                if close_reason == "Unknown":
                    close_reason = "Closed on exchange (reconciliation)"
                
                # Record the trade with the determined reason
                record_completed_trade(
                    symbol=symbol,
                    direction=direction,
                    entry_price=entry_price,
                    exit_price=exit_price,
                    position_size=position_size,
                    entry_time=entry_time,
                    stop_loss=stop_loss,
                    stopped_out=was_stopped_out,
                    reason=close_reason,
                    order_type=order_type
                )
                
                # Log the reconciliation event
                log_reconciliation_event(
                    "POSITION_CLOSED", 
                    symbol, 
                    f"{direction.upper()} position closed. Reason: {close_reason}",
                    {
                        "direction": direction,
                        "entry_price": entry_price,
                        "exit_price": exit_price,
                        "position_size": position_size,
                        "reason": close_reason
                    }
                )
                
                # Remove from tracking
                remove_position(symbol)
                if symbol in active_positions:
                    del active_positions[symbol]
                    
                # Calculate P&L
                if direction == 'long':
                    pnl = (exit_price - entry_price) * position_size
                else:
                    pnl = (entry_price - exit_price) * position_size
                
                # Update capital - for tracking purposes (will be corrected at end)
                update_capital(pnl)
                
        # PART 3: Check for untracked TP orders
        try:
            # Get all active orders from Bybit - using settleCoin parameter
            open_orders_response = client.get_open_orders(
                category="linear",
                settleCoin="USDT"  # Add this parameter
            )
            
            if open_orders_response.get('retCode') == 0:
                open_orders_list = open_orders_response.get('result', {}).get('list', [])
                
                # Look for TP orders that might not be tracked
                for order in open_orders_list:
                    # IMPORTANT: Convert order_id to string for consistent comparison
                    order_id_str = str(order.get('orderId', ''))
                    
                    # Check if already tracked - compare as strings to avoid type mismatch
                    already_tracked = False
                    for tracked_id in pending_orders.keys():
                        if str(tracked_id) == order_id_str:
                            already_tracked = True
                            break
                    
                    if not already_tracked and order.get('reduceOnly') is True:
                        symbol = order.get('symbol')
                        
                        # Check if we have an active position for this symbol
                        if symbol in active_positions:
                            # Determine if this is a stop loss based on order type
                            order_type = order.get('orderType', 'Limit')
                            stop_order_type = order.get('stopOrderType', '')
                            trigger_price = float(order.get('triggerPrice', 0) or 0)
                            
                            is_stop_loss = (
                                order_type == 'Market' and stop_order_type in ['StopLoss', 'StopMarket']
                            )

                            # This appears to be a TP/SL order that's not being tracked
                            logger.warning(f"Found untracked {'SL' if is_stop_loss else 'TP'} order {order_id_str} for {symbol}, adding to tracking")
                            
                            # Create order data structure with correct type info
                            order_data = {
                                'order_id': order_id_str,
                                'symbol': symbol,
                                'order_type': order_type,
                                'stop_order_type': stop_order_type,
                                'side': order.get('side'),
                                'price': float(order.get('price', 0)),
                                'trigger_price': trigger_price,
                                'qty': float(order.get('qty', 0)),
                                'status': order.get('orderStatus', 'New'),
                                'created_time': datetime.now().isoformat(),
                                'purpose': 'exit',
                                'related_position_symbol': symbol,
                                'is_stop_loss': is_stop_loss
                            }
                            
                            # Save to pending orders using string ID
                            pending_orders[order_id_str] = order_data
                            save_pending_order(order_data)
                            
                            # Also update the position with the TP order ID (only if it's a TP, not SL)
                            if not is_stop_loss:
                                active_positions[symbol]['tp_order_id'] = order_id_str
                                save_position(symbol, active_positions[symbol])
                            
                            # Log the reconciliation event
                            log_reconciliation_event(
                                "UNTRACKED_TP_ORDER_FOUND", 
                                symbol, 
                                f"Found untracked {'SL' if is_stop_loss else 'TP'} order {order_id_str}, added to tracking",
                                order_data
                            )
                            
                            # Send notification if configured
                            if config.getboolean('NOTIFICATIONS', 'notify_reconciliation', fallback=False):
                                direction = active_positions[symbol]['direction']
                                send_telegram_message(f"🔄 <b>RECONCILIATION</b>\n"
                                                   f"Found untracked {'SL' if is_stop_loss else 'TP'} order for {symbol} {direction.upper()}\n"
                                                   f"Price: {order_data['price']}\n"
                                                   f"Added to tracking")
            else:
                logger.warning(f"Failed to get open orders with settleCoin: {open_orders_response.get('retMsg')}")
                
                # Fallback: Query open orders for each active position symbol
                for symbol in active_positions:
                    try:
                        symbol_orders_response = client.get_open_orders(
                            category="linear",
                            symbol=symbol
                        )
                        
                        if symbol_orders_response.get('retCode') == 0:
                            symbol_orders = symbol_orders_response.get('result', {}).get('list', [])
                            
                            # Process orders for this symbol
                            for order in symbol_orders:
                                # IMPORTANT: Convert order_id to string for consistent comparison
                                order_id_str = str(order.get('orderId', ''))
                                
                                # Check if already tracked - compare as strings
                                already_tracked = False
                                for tracked_id in pending_orders.keys():
                                    if str(tracked_id) == order_id_str:
                                        already_tracked = True
                                        break
                                
                                if not already_tracked and order.get('reduceOnly') is True:
                                    # Determine if this is a stop loss
                                    order_type = order.get('orderType', 'Limit')
                                    stop_order_type = order.get('stopOrderType', '')
                                    trigger_price = float(order.get('triggerPrice', 0) or 0)
                                    
                                    is_stop_loss = (
                                        order_type == 'Market' and stop_order_type in ['StopLoss', 'StopMarket']
                                    )

                                    logger.warning(f"Found untracked {'SL' if is_stop_loss else 'TP'} order {order_id_str} for {symbol} (fallback method), adding to tracking")
                                    
                                    # Create order data structure with correct type info
                                    order_data = {
                                        'order_id': order_id_str,
                                        'symbol': symbol,
                                        'order_type': order_type,
                                        'stop_order_type': stop_order_type,
                                        'side': order.get('side'),
                                        'price': float(order.get('price', 0)),
                                        'trigger_price': trigger_price,
                                        'qty': float(order.get('qty', 0)),
                                        'status': order.get('orderStatus', 'New'),
                                        'created_time': datetime.now().isoformat(),
                                        'purpose': 'exit',
                                        'related_position_symbol': symbol,
                                        'is_stop_loss': is_stop_loss
                                    }
                                    
                                    # Save to pending orders using string ID
                                    pending_orders[order_id_str] = order_data
                                    save_pending_order(order_data)
                                    
                                    # Also update the position with the TP order ID (only if it's a TP)
                                    if not is_stop_loss:
                                        active_positions[symbol]['tp_order_id'] = order_id_str
                                        save_position(symbol, active_positions[symbol])
                                    
                                    # Log the reconciliation event
                                    log_reconciliation_event(
                                        "UNTRACKED_TP_ORDER_FOUND", 
                                        symbol, 
                                        f"Found untracked {'SL' if is_stop_loss else 'TP'} order {order_id_str}, added to tracking (fallback method)",
                                        order_data
                                    )
                    except Exception as symbol_error:
                        logger.error(f"Error checking open orders for {symbol}: {str(symbol_error)}")
        except Exception as e:
            logger.error(f"Error checking for untracked TP orders: {str(e)}")
        
        
        # PART 4: Synchronize capital with Bybit after all reconciliation is done
        try:
            # Only refresh capital if we're not in testnet or if positions changed
            if not config['API'].getboolean('testnet', False) or positions_closed:
                current_capital = get_bybit_balance()
                # Store the new capital value to our local tracker
                with open('capital.txt', 'w') as f:
                    f.write(str(current_capital))
                
                # Log capital sync
                logger.info(f"Capital synchronized with Bybit: ${current_capital} (change: ${current_capital - initial_capital:.2f})")
                
                # If capital changed significantly, log an event
                if abs(current_capital - initial_capital) > 1.0:  # More than $1 change
                    log_reconciliation_event(
                        "CAPITAL_SYNC", 
                        "GLOBAL", 
                        f"Capital synchronized with Bybit. Previous: ${initial_capital:.2f}, Current: ${current_capital:.2f}, Change: ${current_capital - initial_capital:.2f}",
                        {
                            "previous_capital": initial_capital,
                            "current_capital": current_capital,
                            "change": current_capital - initial_capital
                        }
                    )
                    
                    # Notify on significant capital changes
                    if abs(current_capital - initial_capital) > float(config['TRADING'].get('risk_percentage', 5)) * initial_capital / 100:
                        if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                            send_telegram_message(f"💰 <b>CAPITAL UPDATE</b>\n"
                                                f"Capital synchronized with Bybit\n"
                                                f"Previous: ${initial_capital:.2f}\n"
                                                f"Current: ${current_capital:.2f}\n"
                                                f"Change: ${current_capital - initial_capital:.2f} ({(current_capital - initial_capital) / initial_capital * 100:.2f}%)")
        except Exception as e:
            logger.error(f"Error synchronizing capital with Bybit: {str(e)}")
        
        # Update previous positions for next comparison
        previous_positions = active_positions.copy()
        
        # Update balance history for max loss tracking
        update_balance_history()
        
        # Check for max daily loss condition
        handle_max_daily_loss()

        # Clean up old entries from recently_closed_positions
        current_time = datetime.now()
        for symbol in list(recently_closed_positions.keys()):
            if (current_time - recently_closed_positions[symbol]).total_seconds() > RECONCILIATION_COOLING_PERIOD * 2:
                del recently_closed_positions[symbol]
                
        logger.info(f"Enhanced position reconciliation complete. Active positions: {len(active_positions)}")
    except Exception as e:
        logger.error(f"Error during enhanced position reconciliation: {str(e)}")
        
        # Send notification for error
        if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
            send_telegram_message(f"❌ <b>SYSTEM ERROR</b>\n"
                                f"Error during position reconciliation: {str(e)}")





def start_reconciliation_thread():
    """Start a background thread to regularly reconcile positions"""
    def reconciliation_loop():
        # Initialize enhanced database on first run
        init_enhanced_database()
        
        # Initial reconciliation on startup
        enhanced_reconcile_positions()
        
        schedule_mode = config.get('MONITORING', 'reconciliation_schedule_mode', fallback='interval').strip().lower()
        
        if schedule_mode == 'fixed_second':
            # Fixed-second mode: run at a specific second of the minute, every N minutes
            fixed_second = int(config.get('MONITORING', 'reconciliation_fixed_second', fallback=30))
            fixed_second = max(0, min(59, fixed_second))  # Clamp to 0-59
            reconciliation_interval = int(config.get('MONITORING', 'reconciliation_interval_seconds', fallback=60))
            interval_minutes = max(1, reconciliation_interval // 60)  # Convert to minutes, minimum 1
            
            logger.info(f"Reconciliation using fixed_second mode: every {interval_minutes} minute(s) at second :{fixed_second:02d}")
            
            while True:
                try:
                    # Calculate seconds until the next target time
                    now = datetime.now()
                    current_minute = now.minute
                    current_second = now.second
                    current_microsecond = now.microsecond
                    
                    # Find the next minute that aligns with our interval
                    # We align to minutes divisible by interval_minutes from the top of the hour
                    if interval_minutes == 1:
                        # Every minute: next target is the next occurrence of :XX seconds
                        if current_second < fixed_second:
                            # Target is later this minute
                            target = now.replace(second=fixed_second, microsecond=0)
                        else:
                            # Target is next minute
                            target = now.replace(second=fixed_second, microsecond=0) + timedelta(minutes=1)
                    else:
                        # Every N minutes: align to minutes divisible by interval_minutes
                        # Find the next aligned minute
                        minutes_since_hour = current_minute % interval_minutes
                        
                        if minutes_since_hour == 0 and current_second < fixed_second:
                            # We're on an aligned minute and haven't passed the target second yet
                            target = now.replace(second=fixed_second, microsecond=0)
                        else:
                            # Jump to the next aligned minute
                            minutes_to_next = interval_minutes - minutes_since_hour
                            if minutes_since_hour == 0:
                                minutes_to_next = interval_minutes  # We already passed it this minute
                            target = now.replace(second=fixed_second, microsecond=0) + timedelta(minutes=minutes_to_next)
                    
                    sleep_seconds = (target - now).total_seconds()
                    if sleep_seconds < 0.5:
                        sleep_seconds += interval_minutes * 60  # Safety: avoid spinning
                    
                    logger.debug(f"Reconciliation sleeping {sleep_seconds:.1f}s until {target.strftime('%H:%M:%S')}")
                    time.sleep(sleep_seconds)
                    
                    # Run the enhanced reconciliation
                    enhanced_reconcile_positions()
                except Exception as e:
                    logger.error(f"Error in reconciliation loop (fixed_second): {str(e)}")
                    time.sleep(60)  # Sleep for 60 seconds on error before retrying
        else:
            # Original interval mode: sleep N seconds between runs
            while True:
                try:
                    reconciliation_interval = int(config.get('MONITORING', 'reconciliation_interval_seconds', fallback=60))
                    time.sleep(reconciliation_interval)
                    
                    # Run the enhanced reconciliation
                    enhanced_reconcile_positions()
                except Exception as e:
                    logger.error(f"Error in reconciliation loop: {str(e)}")
                    time.sleep(60)  # Sleep for 60 seconds on error before retrying
    
    # Start the reconciliation thread
    reconciliation_thread = threading.Thread(target=reconciliation_loop, daemon=True)
    reconciliation_thread.start()
    
    schedule_mode = config.get('MONITORING', 'reconciliation_schedule_mode', fallback='interval').strip().lower()
    if schedule_mode == 'fixed_second':
        fixed_second = int(config.get('MONITORING', 'reconciliation_fixed_second', fallback=30))
        interval_minutes = max(1, int(config.get('MONITORING', 'reconciliation_interval_seconds', fallback=60)) // 60)
        logger.info(f"Started position reconciliation thread (fixed_second mode: every {interval_minutes}min at :{fixed_second:02d}s)")
    else:
        logger.info(f"Started position reconciliation thread with interval: {config.get('MONITORING', 'reconciliation_interval_seconds', fallback=60)} seconds")

# IMPROVEMENT: Connection Resilience
def check_internet():
    """Check basic internet connectivity by reaching google.com"""
    try:
        response = requests.get("https://www.google.com", timeout=5)
        return response.status_code == 200
    except Exception:
        return False


def check_exchange():
    """Check actual exchange connectivity with a real API call"""
    try:
        active_exchange = config.get('EXCHANGE', 'active_exchange', fallback='bybit').lower()
        client = initialize_bybit_client()
        
        if active_exchange == 'hyperliquid':
            # For Hyperliquid: get_server_time() is fake (returns local time),
            # so we use get_tickers() which makes a real API call to Hyperliquid
            response = client.get_tickers(category="linear", symbol="BTCUSDT")
            if response.get('retCode') == 0:
                price = response.get('result', {}).get('list', [{}])[0].get('lastPrice', None)
                if price and float(price) > 0:
                    return True
            return False
        else:
            # For Bybit: get_server_time() works fine as a real connectivity check
            response = client.get_server_time()
            return response.get('retCode') == 0
    except Exception as e:
        logger.debug(f"Exchange connectivity check failed: {e}")
        return False


def check_connectivity():
    """Run both internet and exchange connectivity checks simultaneously"""
    global bot_status, connection_loss_time
    global internet_status, exchange_status, internet_loss_time, exchange_loss_time
    
    # Run both checks at the same time using threads
    internet_result = [None]
    exchange_result = [None]
    
    def run_internet_check():
        internet_result[0] = check_internet()
    
    def run_exchange_check():
        exchange_result[0] = check_exchange()
    
    internet_thread = threading.Thread(target=run_internet_check)
    exchange_thread = threading.Thread(target=run_exchange_check)
    
    internet_thread.start()
    exchange_thread.start()
    
    # Wait for both to finish (with a timeout so we don't hang forever)
    internet_thread.join(timeout=10)
    exchange_thread.join(timeout=10)
    
    # If threads didn't finish in time, treat as failure
    internet_ok = internet_result[0] if internet_result[0] is not None else False
    exchange_ok = exchange_result[0] if exchange_result[0] is not None else False
    
    active_exchange = config.get('EXCHANGE', 'active_exchange', fallback='bybit').lower()
    now = datetime.now()
    
    # ---- Process internet status ----
    if internet_ok:
        if internet_status != "online":
            downtime_info = ""
            if internet_loss_time:
                duration = now - internet_loss_time
                days = duration.days
                hours, remainder = divmod(duration.seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                downtime_info = f" (was down for {days}d {hours}h {minutes}m {seconds}s)"
            logger.info(f"Internet connectivity restored{downtime_info}")
            internet_loss_time = None
        internet_status = "online"
    else:
        if internet_status != "offline":
            internet_loss_time = now
            logger.warning("Internet connectivity lost - cannot reach google.com")
        internet_status = "offline"
    
    # ---- Process exchange status ----
    if exchange_ok:
        if exchange_status != "online":
            downtime_info = ""
            if exchange_loss_time:
                duration = now - exchange_loss_time
                days = duration.days
                hours, remainder = divmod(duration.seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                downtime_info = f" (was down for {days}d {hours}h {minutes}m {seconds}s)"
            logger.info(f"{active_exchange.capitalize()} exchange connectivity restored{downtime_info}")
            exchange_loss_time = None
        exchange_status = "online"
    else:
        if exchange_status != "offline":
            exchange_loss_time = now
            logger.warning(f"{active_exchange.capitalize()} exchange connectivity lost")
            
            # Consider resetting client if exchange is down
            if internet_ok:
                # Internet works but exchange doesn't - might be an auth issue
                reset_global_client()
        exchange_status = "offline"
    
    # ---- Determine overall bot status ----
    previous_bot_status = bot_status
    
    if internet_ok and exchange_ok:
        # Everything is fine
        if previous_bot_status != "running":
            # We recovered from a problem
            disconnection_info = ""
            if connection_loss_time:
                duration = now - connection_loss_time
                days = duration.days
                hours, remainder = divmod(duration.seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                disconnection_info = (f"\nDisconnected at: {connection_loss_time.strftime('%Y-%m-%d %H:%M:%S')}"
                                     f"\nReconnected at: {now.strftime('%Y-%m-%d %H:%M:%S')}"
                                     f"\nDowntime: {days}d {hours}h {minutes}m {seconds}s")
            
            logger.info(f"All connections restored, resuming normal operation{disconnection_info}")
            
            if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                send_telegram_message(f"✅ <b>SYSTEM STATUS</b>\n"
                                     f"All connections restored, bot is running normally{disconnection_info}")
            
            connection_loss_time = None
        
        bot_status = "running"
        return True
    
    else:
        # Something is wrong - build a status message
        problems = []
        if not internet_ok:
            problems.append("Internet DOWN")
        if not exchange_ok:
            problems.append(f"{active_exchange.capitalize()} exchange DOWN")
        problem_str = " | ".join(problems)
        
        if previous_bot_status == "running":
            # First time detecting a problem
            connection_loss_time = now
            
            logger.warning(f"Connectivity issue detected: {problem_str}")
            
            # Play sound alarm if enabled
            if config['MONITORING'].getboolean('sound_alarm_enabled', False) and WINSOUND_AVAILABLE:
                try:
                    frequency = int(config['MONITORING'].get('sound_alarm_frequency', 1000))
                    duration = int(config['MONITORING'].get('sound_alarm_duration', 500))
                    winsound.Beep(frequency, duration)
                except Exception as sound_error:
                    logger.error(f"Error playing sound alarm: {str(sound_error)}")
            
            # Send Telegram notification
            if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                send_telegram_message(f"⚠️ <b>CONNECTIVITY ISSUE</b>\n"
                                     f"Internet: {'✅ Online' if internet_ok else '❌ Offline'}\n"
                                     f"{active_exchange.capitalize()}: {'✅ Online' if exchange_ok else '❌ Offline'}\n"
                                     f"Detected at: {now.strftime('%Y-%m-%d %H:%M:%S')}")
        
        else:
            # Already in degraded state - just play alarm on each check if enabled
            if config['MONITORING'].getboolean('sound_alarm_enabled', False) and WINSOUND_AVAILABLE :
                try:
                    frequency = int(config['MONITORING'].get('sound_alarm_frequency', 1000))
                    duration = int(config['MONITORING'].get('sound_alarm_duration', 500))
                    winsound.Beep(frequency, duration)
                except Exception as sound_error:
                    logger.error(f"Error playing sound alarm: {str(sound_error)}")
        
        bot_status = "degraded"
        return False


def start_connectivity_monitor():
    """Start a background thread to monitor API connectivity"""
    def monitor_loop():
        while True:
            check_connectivity()
            # Get the check interval from config, default to 60 seconds if not specified
            check_interval = int(config.get('MONITORING', 'connectivity_check_seconds', fallback=60))
            time.sleep(check_interval)
    
    monitor_thread = threading.Thread(target=monitor_loop, daemon=True)
    monitor_thread.start()
    logger.info(f"Connection monitoring started with interval: {config.get('MONITORING', 'connectivity_check_seconds', fallback=60)} seconds")

# Validate webhook data
def validate_webhook_data(data):
    try:
        # Check required fields
        required_fields = ['action', 'ticker', 'close']
        if not all(field in data for field in required_fields):
            logger.warning(f"Missing required fields in webhook data: {data}")
            return False
        
        # Validate action type
        valid_actions = ["LONG ENTRY", "SHORT ENTRY", "EXIT LONG", "EXIT SHORT"]
        if not any(data.get('action').startswith(valid_action) for valid_action in valid_actions):
            logger.warning(f"Invalid action in webhook data: {data.get('action')}")
            return False
            
        # Validate numerical fields
        try:
            float(data.get('close', 0))
            if 'stopLoss' in data:
                float(data.get('stopLoss', 0))
        except ValueError:
            logger.warning(f"Invalid numerical values in webhook data: {data}")
            return False
            
        return True
    except Exception as e:
        logger.error(f"Error validating webhook data: {str(e)}")
        return False


# Process alert from TradingView
def process_tradingview_alert(alert_data):
    try:
        # Validate incoming data
        if not validate_webhook_data(alert_data):
            logger.error(f"Invalid webhook data, skipping: {alert_data}")
            
            # Send notification for error
            if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                send_telegram_message(f"❌ <b>ERROR</b>\n"
                                    f"Invalid webhook data received: {json.dumps(alert_data)}")
            
            return
        
        # ========== Check for max daily loss status before processing new trades ==========
        # Only check for entry signals, not exits
        action = alert_data.get('action', '')
        if action.startswith("LONG ENTRY") or action.startswith("SHORT ENTRY"):
            if config.get('RISK_MANAGEMENT', 'max_daily_loss_enabled', fallback='True').lower() == 'true':
                # Check if max loss threshold has been exceeded
                max_loss_exceeded, loss_percent, current_balance, reference_balance = check_max_daily_loss()
                
                if max_loss_exceeded:
                    global skipped_signals_count  # Make sure to use the global counter variable
                    skipped_signals_count += 1  # Increment the counter when signals are skipped
                    
                    logger.warning(f"Max daily loss triggered ({loss_percent:.2f}%), ignoring entry signal: {action}")
                    
                    # Send notification
                    if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                        max_loss_percent = float(config['RISK_MANAGEMENT'].get('max_daily_loss_percent', 2.0))
                        send_telegram_message(f"❌ <b>SIGNAL IGNORED: MAX DAILY LOSS</b>\n"
                                            f"Signal: {action}\n"
                                            f"Current loss: {loss_percent:.2f}%\n"
                                            f"Threshold: {max_loss_percent}%\n"
                                            f"New positions are blocked until loss decreases below threshold")
                    return  # Skip processing this signal
            
        client = initialize_bybit_client()  # Will reuse existing client
        
        # Extract information from alert
        action = alert_data.get('action', '')
        symbol = alert_data.get('ticker', config['TRADING'].get('default_symbol', 'BTCUSDT'))
        price = float(alert_data.get('close', 0))
        
        # For entries, get stop loss if available
        stop_loss = None
        if 'stopLoss' in alert_data:
            stop_loss = float(alert_data.get('stopLoss', 0))
            
        # For entries, get take profit if available
        take_profit = None
        if 'takeProfit' in alert_data:
            take_profit = float(alert_data.get('takeProfit', 0))
            
        # Determine if we should use limit orders from config
        use_limit_orders_entry = config['TRADING'].getboolean('use_limit_orders_entry', False)
        use_limit_orders_exit = config['TRADING'].getboolean('use_limit_orders_exit', False)
        
        # Check if this is an entry signal
        if action.startswith("LONG ENTRY") or action.startswith("SHORT ENTRY"):
            # Check if there's already an active position for this symbol
            if symbol in active_positions:
                logger.info(f"Existing position found for {symbol} - closing before opening new position")
                
                # Get the direction of the existing position
                existing_direction = active_positions[symbol]['direction']
                
                # Determine exit action based on existing position direction
                exit_action = "EXIT LONG" if existing_direction == 'long' else "EXIT SHORT"
                
                # Send notification about position closure
                if config.getboolean('NOTIFICATIONS', 'notify_entries', fallback=True):
                    send_telegram_message(f"🔄 <b>CLOSING EXISTING POSITION</b>\n"
                                        f"Symbol: {symbol}\n"
                                        f"Current Position: {existing_direction.upper()}\n"
                                        f"Reason: New {action} signal received\n"
                                        f"Exit Type: Market Order (forced)")
                
                # IMPORTANT: Force market order (use_limit_order=False) when closing positions due to new entry signals
                # This ensures the position is fully closed before opening a new one
                exit_result = execute_trade_with_retry(client, exit_action, symbol, price, use_limit_order=False, closeOnTrigger=True)
                
                if not exit_result:
                    logger.error(f"Failed to close existing position for {symbol}, aborting new entry")
                    
                    # Send notification for error
                    if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                        send_telegram_message(f"❌ <b>ERROR</b>\n"
                                            f"Failed to close existing position for {symbol}\n"
                                            f"New {action} signal ignored")
                    return
                
                # Small delay to ensure the position is properly closed before opening a new one
                time.sleep(2)  # 2 seconds delay should be sufficient for the exchange to process
                
                logger.info(f"Successfully closed existing {existing_direction} position for {symbol} with market order")
                
                # Ensure the position is removed from tracking
                if symbol in active_positions:
                    logger.warning(f"Position for {symbol} still in active_positions after closure, removing manually")
                    del active_positions[symbol]
                    remove_position(symbol)
        
        # Execute the trade with retry logic and limit order options
        result = execute_trade_with_retry(
            client, 
            action, 
            symbol, 
            price, 
            stop_loss, 
            None,  # position_size will be calculated in execute_trade
            take_profit, 
            use_limit_orders_entry, 
            use_limit_orders_exit
        )
        
        if result:
            logger.info(f"Successfully processed alert: {action} for {symbol}")
        else:
            logger.error(f"Failed to process alert: {action} for {symbol}")
            
    except Exception as e:
        logger.error(f"Error processing TradingView alert: {str(e)}")
        
        # Send notification for error
        if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
            send_telegram_message(f"❌ <b>ERROR</b>\n"
                                f"Error processing TradingView alert: {str(e)}")



# IMPROVEMENT: Comprehensive Health Monitoring
# Get system metrics for health endpoint
def get_system_metrics():
    try:
        process = psutil.Process(os.getpid())
        
        return {
            "cpu_percent": process.cpu_percent(),
            "memory_percent": process.memory_percent(),
            "memory_mb": process.memory_info().rss / (1024 * 1024),
            "threads": len(process.threads()),
            "open_files": len(process.open_files()),
            "api_calls": api_calls,
            "api_errors": api_errors,
            "error_rate": (api_errors / api_calls * 100) if api_calls > 0 else 0
        }
    except Exception as e:
        logger.error(f"Error getting system metrics: {str(e)}")
        return {}

# Get uptime for health endpoint
def get_uptime():
    uptime = datetime.now() - start_time
    days = uptime.days
    hours, remainder = divmod(uptime.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    
    return f"{days}d {hours}h {minutes}m {seconds}s"

# Flask route for webhook
@webhook_app.route(config['WEBHOOK'].get('endpoint', '/webhook'), methods=['POST'])
def webhook():
    if request.method == 'POST':
        try:
            data = request.json
            logger.info(f"Received webhook: {data}")
            
            # Process in a separate thread to avoid blocking
            threading.Thread(target=process_tradingview_alert, args=(data,)).start()
            
            return jsonify({"status": "success", "message": "Alert received"})
        except Exception as e:
            logger.error(f"Error in webhook: {str(e)}")
            return jsonify({"status": "error", "message": str(e)}), 500
    
    return jsonify({"status": "error", "message": "Method not allowed"}), 405

# Enhanced health check endpoint
@app.route('/health', methods=['GET'])
def health_check():
    try:
        # Test API connection
        api_status = "unknown"
        try:
            client = initialize_bybit_client()  # Will reuse existing client
            account_info = client.get_wallet_balance(accountType="UNIFIED")
            if account_info.get('retCode') == 0:
                api_status = "online"
            else:
                api_status = "error"
        except Exception:
            api_status = "offline"
        
        # System metrics
        system_metrics = get_system_metrics()
        
        # Current positions
        positions_info = []
        for symbol, position in active_positions.items():
            positions_info.append({
                "symbol": symbol,
                "direction": position['direction'],
                "entry_price": position['entry_price'],
                "position_size": position['position_size'],
                "entry_time": position['entry_time'],
                "order_type": position.get('order_type', 'Market')
            })
            
        # Add pending orders info
        pending_orders_info = []
        for order_id, order in pending_orders.items():
            pending_orders_info.append({
                "order_id": order_id,
                "symbol": order['symbol'],
                "side": order['side'],
                "price": order['price'],
                "qty": order['qty'],
                "order_type": order['order_type'],
                "purpose": order.get('purpose', 'unknown'),
                "created_time": order.get('created_time', ''),
                "status": order.get('status', 'unknown')
            })
        
        # Check if reconciliation is running
        reconciliation_enabled = config.getboolean('MONITORING', 'enabled', fallback=True) and config.get('MONITORING', 'reconciliation_interval_seconds', fallback='60') != '0'
        
        return jsonify({
            "status": "healthy" if api_status == "online" and bot_status == "running" else "degraded",
            "time": datetime.now().isoformat(),
            "uptime": get_uptime(),
            "bot_status": bot_status,
            "api_status": api_status,
            "internet_status": internet_status,
            "exchange_status": exchange_status,
            "reconciliation_enabled": reconciliation_enabled,
            "system": system_metrics,
            "capital": get_current_capital(),
            "active_positions_count": len(active_positions),
            "active_positions": positions_info,
            "pending_orders_count": len(pending_orders),
            "pending_orders": pending_orders_info,
            "order_types": {
                "entry": "Limit" if config['TRADING'].getboolean('use_limit_orders_entry', False) else "Market",
                "exit": "Limit" if config['TRADING'].getboolean('use_limit_orders_exit', False) else "Market"
            }
        })
    except Exception as e:
        logger.error(f"Error in health check: {str(e)}")
        return jsonify({
            "status": "error",
            "time": datetime.now().isoformat(),
            "error": str(e)
        }), 500

# IMPROVEMENT: Enhanced Monitoring
@app.route('/monitoring', methods=['GET'])
def monitoring():
    try:
        # System metrics
        system_metrics = get_system_metrics()
        
        # Trading performance
        capital_file = 'capital.txt'
        initial_capital = float(config['TRADING'].get('initial_capital', 10000))
        current_capital = get_current_capital()
        
        # Get reconciliation status
        reconciliation_info = {}
        try:
            conn = sqlite3.connect('trading_bot.db')
            cursor = conn.cursor()
            
            # Check if reconciliation table exists
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='reconciliation_log'")
            if cursor.fetchone():
                # Get the latest reconciliation events
                cursor.execute('SELECT event_type, COUNT(*) FROM reconciliation_log GROUP BY event_type')
                event_counts = cursor.fetchall()
                for event_type, count in event_counts:
                    reconciliation_info[f"{event_type}_count"] = count
                
                # Get last reconciliation time
                cursor.execute('SELECT MAX(timestamp) FROM reconciliation_log')
                last_time = cursor.fetchone()[0]
                if last_time:
                    reconciliation_info["last_reconciliation"] = last_time
            
            conn.close()
        except Exception as e:
            logger.error(f"Error getting reconciliation info: {str(e)}")
        
        performance = {
            "initial_capital": initial_capital,
            "current_capital": current_capital,
            "profit_loss": current_capital - initial_capital,
            "profit_loss_percent": ((current_capital - initial_capital) / initial_capital) * 100,
        }
        
        # Get limit order stats
        limit_order_info = {
            "enabled_entry": config['TRADING'].getboolean('use_limit_orders_entry', False),
            "enabled_exit": config['TRADING'].getboolean('use_limit_orders_exit', False),
            "timeout_minutes": int(config['TRADING'].get('limit_order_timeout_minutes', 5)),
            "pending_count": len(pending_orders)
        }
        
        return jsonify({
            "time": datetime.now().isoformat(),
            "uptime": get_uptime(),
            "system": system_metrics,
            "trading_performance": performance,
            "active_positions": len(active_positions),
            "reconciliation": reconciliation_info,
            "limit_orders": limit_order_info
        })
    except Exception as e:
        logger.error(f"Error in monitoring endpoint: {str(e)}")
        return jsonify({
            "status": "error",
            "error": str(e)
        }), 500

# IMPROVEMENT: Performance Tracking Endpoint
@app.route('/performance', methods=['GET'])
def performance():
    try:
        metrics = get_performance_metrics()
        
        # Get recent trades
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Check if order_type column exists
        cursor.execute("PRAGMA table_info(trade_history)")
        columns = [col[1] for col in cursor.fetchall()]
        
        # Build query dynamically based on available columns
        query_columns = ['id', 'symbol', 'direction', 'entry_price', 'exit_price', 'pnl', 'pnl_percent', 'entry_time', 'exit_time']
        
        # Add optional columns if they exist
        if 'reason' in columns:
            query_columns.append('reason')
        if 'order_type' in columns:
            query_columns.append('order_type')
            
        query = f"SELECT {', '.join(query_columns)} FROM trade_history ORDER BY exit_time DESC LIMIT 10"
        cursor.execute(query)
        
        columns = [col[0] for col in cursor.description]
        
        recent_trades = []
        for row in cursor.fetchall():
            trade = {columns[i]: row[i] for i in range(len(columns))}
            recent_trades.append(trade)
        
        # Get trades by reason if the column exists
        trades_by_reason = {}
        if 'reason' in columns:
            cursor.execute('SELECT reason, COUNT(*), AVG(pnl), SUM(pnl) FROM trade_history GROUP BY reason')
            for row in cursor.fetchall():
                reason = row[0] or "Unknown"
                trades_by_reason[reason] = {
                    "count": row[1],
                    "avg_pnl": row[2],
                    "total_pnl": row[3]
                }
                
        # Get trades by order type if the column exists
        trades_by_order_type = {}
        if 'order_type' in columns:
            cursor.execute('SELECT order_type, COUNT(*), AVG(pnl), SUM(pnl) FROM trade_history GROUP BY order_type')
            for row in cursor.fetchall():
                order_type = row[0] or "Market"
                trades_by_order_type[order_type] = {
                    "count": row[1],
                    "avg_pnl": row[2],
                    "total_pnl": row[3]
                }
                
        conn.close()
        
        response_data = {
            'metrics': metrics,
            'recent_trades': recent_trades
        }
        
        if trades_by_reason:
            response_data['trades_by_reason'] = trades_by_reason
            
        if trades_by_order_type:
            response_data['trades_by_order_type'] = trades_by_order_type
            
        return jsonify(response_data)
    except Exception as e:
        logger.error(f"Error in performance endpoint: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

# Add endpoint to view reconciliation log
@app.route('/reconciliation/log', methods=['GET'])
def reconciliation_log():
    try:
        limit = request.args.get('limit', default=100, type=int)
        
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Check if reconciliation_log table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='reconciliation_log'")
        if not cursor.fetchone():
            conn.close()
            return jsonify({
                'status': 'error',
                'message': 'Reconciliation log table not found'
            }), 404
        
        # Get the latest log entries
        cursor.execute('SELECT * FROM reconciliation_log ORDER BY timestamp DESC LIMIT ?', (limit,))
        columns = [col[0] for col in cursor.description]
        
        log_entries = []
        for row in cursor.fetchall():
            entry = {columns[i]: row[i] for i in range(len(columns))}
            # Parse position_data JSON if it exists
            if entry.get('position_data'):
                try:
                    entry['position_data'] = json.loads(entry['position_data'])
                except:
                    pass
            log_entries.append(entry)
            
        conn.close()
        
        return jsonify({
            'status': 'success',
            'count': len(log_entries),
            'log': log_entries
        })
    except Exception as e:
        logger.error(f"Error getting reconciliation log: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

# EXCEL LOGGING: Add endpoints to manage Excel file
@app.route('/excel/export', methods=['GET'])
def export_to_excel():
    """Endpoint to manually trigger database export to Excel"""
    try:
        if excel_logger is None:
            return jsonify({
                'status': 'error',
                'message': 'Excel logging is not enabled'
            }), 400
            
        # Get limit parameter from query string, default to all trades
        limit = request.args.get('limit', default=None, type=int)
        
        # Get the database path
        db_path = 'trading_bot.db'
        
        # Export trades
        excel_logger.export_from_database(db_path, limit)
        
        return jsonify({
            'status': 'success',
            'message': f'Exported trades to Excel file: {excel_logger.excel_path}',
            'limit': limit or 'all'
        })
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/excel/backup', methods=['GET'])
def backup_excel_file():
    """Endpoint to manually trigger Excel file backup"""
    try:
        if excel_logger is None:
            return jsonify({
                'status': 'error',
                'message': 'Excel logging is not enabled'
            }), 400
            
        # Trigger backup
        excel_logger.backup_excel()
        
        return jsonify({
            'status': 'success',
            'message': 'Excel file backup created'
        })
    except Exception as e:
        logger.error(f"Error backing up Excel file: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/risk/max-loss-status', methods=['GET'])
def max_loss_status():
    """API endpoint to check current max loss status"""
    try:
        # Get configuration
        max_loss_percent = float(config['RISK_MANAGEMENT'].get('max_daily_loss_percent', 2.0))
        warning_percent = float(config['RISK_MANAGEMENT'].get('max_daily_loss_warning_percent', 1.5))
        period_hours = int(config['RISK_MANAGEMENT'].get('max_daily_loss_period_hours', 24))
        
        # Check current loss
        exceeded, loss_percent, current_balance, reference_balance = check_max_daily_loss()
        
        # Calculate risk level
        if loss_percent < warning_percent:
            risk_level = "normal"
        elif loss_percent >= warning_percent and loss_percent < max_loss_percent:
            risk_level = "warning"
        else:
            risk_level = "critical"
        
        # Get historical events
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Check if the table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='max_loss_events'")
        if cursor.fetchone():
            cursor.execute('''
            SELECT * FROM max_loss_events
            ORDER BY timestamp DESC
            LIMIT 5
            ''')
            
            columns = [col[0] for col in cursor.description]
            recent_events = [dict(zip(columns, row)) for row in cursor.fetchall()]
        else:
            recent_events = []
        
        conn.close()
        
        return jsonify({
            "status": "success",
            "risk_level": risk_level,
            "loss_percent": loss_percent,
            "max_loss_threshold": max_loss_percent,
            "warning_threshold": warning_percent,
            "current_balance": current_balance,
            "reference_balance": reference_balance,
            "reference_period_hours": period_hours,
            "feature_enabled": config['RISK_MANAGEMENT'].getboolean('max_daily_loss_enabled', True),
            "recent_events": recent_events
        })
    except Exception as e:
        logger.error(f"Error in max loss status endpoint: {str(e)}")
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


@app.route('/risk/reset-max-loss', methods=['POST'])
def reset_max_loss():
    """API endpoint to reset max daily loss by setting current balance as the reference"""
    global last_max_loss_notification_time, last_warning_notification_time, max_loss_first_triggered_time, skipped_signals_count
    
    try:
        # Reset notification counters and timers
        last_max_loss_notification_time = None
        last_warning_notification_time = None
        max_loss_first_triggered_time = None
        skipped_signals_count = 0
        
        # Get current balance
        current_balance = get_bybit_balance()
        
        # Insert a new record in balance_history that will be used as reference
        conn = sqlite3.connect('trading_bot.db')
        cursor = conn.cursor()
        
        # Use current time for the reset record (NOT 24 hours ago)
        current_time = datetime.now().isoformat()
        
        # Insert current balance with current timestamp
        cursor.execute('''
        INSERT INTO balance_history (timestamp, balance, source)
        VALUES (?, ?, ?)
        ''', (
            current_time,  # Use current time instead of 24 hours ago
            current_balance,
            'manual_reset'
        ))
        
        conn.commit()
        conn.close()
        
        # Log the reset action
        logger.info(f"Max daily loss manually reset. New reference balance: ${current_balance}")
        
        # Send notification
        if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
            send_telegram_message(f"🔄 <b>MAX DAILY LOSS RESET</b>\n"
                                f"Reference balance reset to current balance: ${current_balance}\n"
                                f"Trading can now resume")
        
        return jsonify({
            "status": "success",
            "message": "Max daily loss reference has been reset to current balance",
            "current_balance": current_balance
        })
    except Exception as e:
        logger.error(f"Error resetting max daily loss: {str(e)}")
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


# Endpoint to get pending orders
@app.route('/api/pending-orders', methods=['GET'])
def get_pending_orders():
    """API endpoint to get current pending orders"""
    try:
        orders_list = []
        
        for order_id, order_data in pending_orders.items():
            # Convert timestamps
            created_time = order_data.get('created_time', '')
            expiry_time = order_data.get('expiry_time', '')
            
            # Get order type info - check multiple possible field names
            order_type = order_data.get('order_type', order_data.get('orderType', 'Limit'))
            stop_order_type = order_data.get('stop_order_type', order_data.get('stopOrderType', ''))
            trigger_price = order_data.get('trigger_price', order_data.get('triggerPrice', 0))
            
            # Determine if this is a stop loss
            is_stop_loss = order_data.get('is_stop_loss', False)
            if not is_stop_loss:
                # Check other indicators
                is_stop_loss = (
                    stop_order_type in ['StopLoss', 'StopMarket']
                )
            
            orders_list.append({
                'orderId': str(order_id),
                'symbol': order_data.get('symbol', ''),
                'orderType': order_type,
                'side': order_data.get('side', ''),
                'price': order_data.get('price', 0),
                'qty': order_data.get('qty', 0),
                'status': order_data.get('status', 'Unknown'),
                'createdTime': created_time,
                'expiryTime': expiry_time,
                'purpose': order_data.get('purpose', 'entry'),
                'isStopLoss': is_stop_loss,
                'stopOrderType': stop_order_type,
                'triggerPrice': trigger_price
            })
        
        return jsonify({
            'status': 'success',
            'orders': orders_list
        })
    except Exception as e:
        logger.error(f"Error getting pending orders: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


# Endpoint to cancel a pending order
@app.route('/api/cancel-order', methods=['POST'])
def cancel_pending_order():
    """API endpoint to cancel a pending order"""
    try:
        data = request.json
        order_id = data.get('orderId')
        
        if not order_id:
            return jsonify({
                'status': 'error',
                'message': 'Order ID is required'
            }), 400
        
        # Check if order exists in pending orders
        if order_id not in pending_orders:
            return jsonify({
                'status': 'error',
                'message': f'Order ID {order_id} not found in pending orders'
            }), 404
        
        # Get symbol from order data
        symbol = pending_orders[order_id].get('symbol', '')
        
        # Cancel order using Bybit API
        client = initialize_bybit_client()
        result = cancel_order(client, order_id, symbol)
        
        if result:
            # Remove from database and memory
            remove_pending_order(order_id)
            
            # Send notification
            if config.getboolean('NOTIFICATIONS', 'notify_order_status', fallback=True):
                purpose = pending_orders[order_id].get('purpose', 'entry')
                side = pending_orders[order_id].get('side', '')
                price = pending_orders[order_id].get('price', 0)
                
                send_telegram_message(f"❌ <b>LIMIT ORDER CANCELED</b>\n"
                                    f"Symbol: {symbol}\n"
                                    f"Side: {side}\n"
                                    f"Price: {price}\n"
                                    f"Purpose: {purpose.capitalize()}")
            
            return jsonify({
                'status': 'success',
                'message': f'Order {order_id} successfully canceled'
            })
        else:
            return jsonify({
                'status': 'error',
                'message': f'Failed to cancel order {order_id}'
            }), 500
            
    except Exception as e:
        logger.error(f"Error canceling order: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


# Serve the manual entry HTML page
@app.route('/')
def index():
    return send_from_directory('.', 'manual_entry.html')

# API endpoint to get current price
@app.route('/api/price', methods=['GET'])
def get_price():
    try:
        symbol = request.args.get('symbol', config['TRADING'].get('default_symbol', 'BTCUSDT'))
        
        client = initialize_bybit_client()
        ticker_response = client.get_tickers(category="linear", symbol=symbol)
        
        if ticker_response.get('retCode') == 0:
            current_price = float(ticker_response.get('result', {}).get('list', [{}])[0].get('lastPrice', 0))
            return jsonify({
                "status": "success",
                "symbol": symbol,
                "price": current_price
            })
        else:
            return jsonify({
                "status": "error",
                "message": f"Failed to get price: {ticker_response.get('retMsg', 'Unknown error')}"
            })
            
    except Exception as e:
        logger.error(f"Error getting price: {str(e)}")
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

# API endpoint to calculate position size
@app.route('/api/calculate-position', methods=['POST'])
def calculate_position_api():
    try:
        data = request.json
        
        symbol = data.get('symbol', config['TRADING'].get('default_symbol', 'BTCUSDT'))
        direction = data.get('direction', 'long')
        current_price = float(data.get('currentPrice', 0))
        stop_loss = float(data.get('stopLoss', 0))
        take_profit = float(data.get('takeProfit', 0)) if data.get('takeProfit') else None
        
        # Get risk parameters
        risk_mode = data.get('riskMode', 'percentage')
        percent_risk = float(data.get('percentRisk', config['TRADING'].get('risk_percentage', 1)))
        fixed_risk = float(data.get('fixedRisk', 100))
        
        # Get order type preferences - Add this line
        use_limit_entry = data.get('useLimitEntryOrder', False)
        
        # Validate inputs
        if not current_price or not stop_loss:
            return jsonify({
                "status": "error",
                "message": "Missing required parameters"
            }), 400
        
        # Validate stop loss direction
        if direction == 'long' and stop_loss >= current_price:
            return jsonify({
                "status": "error",
                "message": "For LONG positions, stop loss must be below the current price"
            }), 400
        
        if direction == 'short' and stop_loss <= current_price:
            return jsonify({
                "status": "error",
                "message": "For SHORT positions, stop loss must be above the current price"
            }), 400
        
        # Validate take profit direction if provided
        if take_profit:
            if direction == 'long' and take_profit <= current_price:
                return jsonify({
                    "status": "error",
                    "message": "For LONG positions, take profit must be above the current price"
                }), 400
            
            if direction == 'short' and take_profit >= current_price:
                return jsonify({
                    "status": "error",
                    "message": "For SHORT positions, take profit must be below the current price"
                }), 400
        
        # Check for max daily loss if enabled
        if config.get('RISK_MANAGEMENT', 'max_daily_loss_enabled', fallback='True').lower() == 'true':
            max_loss_exceeded, loss_percent, current_balance, reference_balance = check_max_daily_loss()
            
            if max_loss_exceeded:
                return jsonify({
                    "status": "error",
                    "message": f"Max daily loss threshold exceeded ({loss_percent:.2f}%). Cannot open new positions."
                }), 400
        
        # Get symbol info from Bybit
        client = initialize_bybit_client()
        symbol_info = get_symbol_info_with_retry(client, symbol)
        
        if not symbol_info:
            return jsonify({
                "status": "error",
                "message": f"Could not get symbol information for {symbol}"
            }), 400
        
        # Calculate position size based on risk mode - Pass use_limit_entry to both functions
        if risk_mode == 'fixed':
            # Use fixed dollar risk amount - Pass use_limit_entry parameter
            risk_amount = fixed_risk
            position_size = calculate_position_size_fixed_risk(current_price, stop_loss, symbol_info, risk_amount, use_limit_entry)
        else:
            # Use percentage risk (default) - Pass use_limit_entry parameter
            risk_percentage = percent_risk / 100
            position_size = calculate_position_size(current_price, stop_loss, symbol_info, risk_percentage, use_limit_entry)
        
        if position_size <= 0:
            return jsonify({
                "status": "error",
                "message": "Calculated position size is too small based on current risk parameters"
            }), 400
        
        # Calculate risk amount for display
        if risk_mode == 'fixed':
            risk_amount = fixed_risk
        else:
            risk_percentage = percent_risk / 100
            available_capital = get_current_capital() * (1 - float(config['TRADING'].get('balance_buffer_percent', 5)) / 100)
            risk_amount = available_capital * risk_percentage
        
        return jsonify({
            "status": "success",
            "symbol": symbol,
            "direction": direction,
            "current_price": current_price,
            "stop_loss": stop_loss,
            "take_profit": take_profit,
            "position_size": position_size,
            "risk_amount": risk_amount,
            "risk_mode": risk_mode,
            "order_type": "Limit" if use_limit_entry else "Market"
        })
        
    except Exception as e:
        logger.error(f"Error calculating position size: {str(e)}")
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

# API endpoint to execute a manual trade
@app.route('/api/execute-trade', methods=['POST'])
def execute_manual_trade():
    try:
        data = request.json
        
        symbol = data.get('symbol', config['TRADING'].get('default_symbol', 'BTCUSDT'))
        direction = data.get('direction', 'long')
        current_price = float(data.get('currentPrice', 0))
        stop_loss = float(data.get('stopLoss', 0))
        take_profit = float(data.get('takeProfit', 0)) if data.get('takeProfit') else None
        
        # Get risk parameters from the request (these will be passed from the frontend)
        risk_mode = data.get('riskMode', 'percentage')
        percent_risk = float(data.get('percentRisk', config['TRADING'].get('risk_percentage', 1)))
        fixed_risk = float(data.get('fixedRisk', 100))
        
        # Get order type preferences
        use_limit_entry_order = data.get('useLimitEntryOrder', False)
        use_limit_tp_order = data.get('useLimitTpOrder', False)
        
        # Validate inputs
        if not current_price or not stop_loss:
            return jsonify({
                "status": "error",
                "message": "Missing required parameters"
            }), 400
        
        # Validate stop loss direction
        if direction == 'long' and stop_loss >= current_price:
            return jsonify({
                "status": "error",
                "message": "For LONG positions, stop loss must be below the current price"
            }), 400
        
        if direction == 'short' and stop_loss <= current_price:
            return jsonify({
                "status": "error",
                "message": "For SHORT positions, stop loss must be above the current price"
            }), 400
        
        # Validate take profit direction if provided
        if take_profit:
            if direction == 'long' and take_profit <= current_price:
                return jsonify({
                    "status": "error",
                    "message": "For LONG positions, take profit must be above the current price"
                }), 400
            
            if direction == 'short' and take_profit >= current_price:
                return jsonify({
                    "status": "error",
                    "message": "For SHORT positions, take profit must be below the current price"
                }), 400
        
        # Create appropriate action string based on direction
        action = "LONG ENTRY" if direction == "long" else "SHORT ENTRY"
        
        # Initialize Bybit client
        client = initialize_bybit_client()
        
        # Check if there's already an open position for this symbol
        if symbol in active_positions:
            logger.info(f"Existing position found for {symbol} - closing before opening new position")
            
            # Get the direction of the existing position
            existing_direction = active_positions[symbol]['direction']
            
            # Determine exit action based on existing position direction
            exit_action = "EXIT LONG" if existing_direction == 'long' else "EXIT SHORT"
            
            # IMPORTANT: Force market order (use_limit_order=False) when closing positions due to new entry
            # Always use market orders for closing existing positions when opening new ones
            exit_result = execute_trade_with_retry(client, exit_action, symbol, current_price, use_limit_order=False, closeOnTrigger=True)
            
            if not exit_result:
                return jsonify({
                    "status": "error",
                    "message": f"Failed to close existing position for {symbol}"
                }), 500
            
            # Small delay to ensure the position is properly closed
            time.sleep(2)
            
            # Verify the position has been removed
            if symbol in active_positions:
                logger.warning(f"Position for {symbol} still in active_positions after closure, removing manually")
                del active_positions[symbol]
                remove_position(symbol)
        
        # Get symbol info
        symbol_info = get_symbol_info_with_retry(client, symbol)
        if not symbol_info:
            return jsonify({
                "status": "error", 
                "message": f"Could not get symbol information for {symbol}"
            }), 400
            
        # Calculate position size based on risk mode
        if risk_mode == 'fixed':
            # Use fixed dollar risk amount - pass use_limit_entry_order parameter
            position_size = calculate_position_size_fixed_risk(current_price, stop_loss, symbol_info, fixed_risk, use_limit_entry_order)
        else:
            # Use the default risk percentage - pass use_limit_entry_order parameter
            risk_percentage = percent_risk / 100
            position_size = calculate_position_size(current_price, stop_loss, symbol_info, risk_percentage, use_limit_entry_order)
        
        if position_size <= 0:
            return jsonify({
                "status": "error",
                "message": "Calculated position size is too small"
            }), 400
        
        # Execute the trade with the calculated position size
        result = execute_trade_with_retry(
            client, 
            action, 
            symbol, 
            current_price, 
            stop_loss, 
            position_size, 
            take_profit, 
            use_limit_entry_order, 
            use_limit_tp_order
        )
        
        if result:
            # Get the order ID from active positions (for market orders)
            # or from pending orders (for limit orders)
            order_id = "Unknown"
            
            if use_limit_entry_order:
                # For limit orders, find the most recent pending order for this symbol
                for oid, order_data in pending_orders.items():
                    if order_data.get('symbol') == symbol and order_data.get('purpose') == 'entry':
                        order_id = oid
                        break
            else:
                # For market orders, get from active positions
                if symbol in active_positions:
                    order_id = active_positions[symbol]['order_id']
            
            return jsonify({
                "status": "success",
                "message": f"Successfully executed {action} for {symbol}",
                "symbol": symbol,
                "direction": direction,
                "entry_price": current_price,
                "stop_loss": stop_loss,
                "take_profit": take_profit,
                "position_size": position_size,
                "order_id": order_id,
                "order_type": "Limit" if use_limit_entry_order else "Market"
            })
        else:
            return jsonify({
                "status": "error",
                "message": f"Failed to execute {action} for {symbol}"
            }), 500
        
    except Exception as e:
        logger.error(f"Error executing manual trade: {str(e)}")
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

# API endpoint to get active positions with current R multiple
@app.route('/api/positions', methods=['GET'])
def get_active_positions():
    try:
        client = initialize_bybit_client()
        positions_data = []
        
        # Get fee percentages for both maker and taker
        taker_fee_percent = float(config['TRADING'].get('trading_fee_percent', 0.075)) / 100  # For market orders
        maker_fee_percent = float(config['TRADING'].get('maker_fee_percent', 0.02)) / 100     # For limit orders
        
        for symbol, position in active_positions.items():
            # Get current price
            ticker_response = client.get_tickers(category="linear", symbol=symbol)
            if ticker_response.get('retCode') == 0:
                current_price = float(ticker_response.get('result', {}).get('list', [{}])[0].get('lastPrice', 0))
                
                # Extract position data
                direction = position['direction']
                entry_price = position['entry_price']
                position_size = position['position_size']
                stop_loss = position.get('stop_loss', 0)
                
                # Get order types
                entry_order_type = position.get('order_type', 'Market')
                exit_order_type = position.get('tp_order_type', 'Market')
                
                # Determine fee percentages based on order types
                entry_fee_percent = maker_fee_percent if entry_order_type == 'Limit' else taker_fee_percent
                exit_fee_percent = maker_fee_percent if exit_order_type == 'Limit' else taker_fee_percent
                
                # Calculate entry fee (already paid)
                entry_fee = entry_price * position_size * entry_fee_percent
                
                # Calculate estimated exit fee (will be paid on exit)
                exit_fee = current_price * position_size * exit_fee_percent
                
                # Total fees
                total_fees = entry_fee + exit_fee
                
                # Calculate current P&L
                if direction == 'long':
                    gross_pnl = (current_price - entry_price) * position_size
                else:  # short
                    gross_pnl = (entry_price - current_price) * position_size
                
                # Net P&L
                net_pnl = gross_pnl - total_fees
                
                # Calculate risk amount (same as in record_completed_trade)
                if stop_loss:
                    if direction == 'long':
                        # Price risk component
                        price_risk_per_unit = entry_price - stop_loss
                        price_risk = price_risk_per_unit * position_size
                        
                        # Fee risk component
                        fee_cost_per_unit = entry_fee_percent * entry_price + taker_fee_percent * stop_loss
                        fee_risk = fee_cost_per_unit * position_size
                    else:  # short
                        # Price risk component
                        price_risk_per_unit = stop_loss - entry_price
                        price_risk = price_risk_per_unit * position_size
                        
                        # Fee risk component
                        fee_cost_per_unit = entry_fee_percent * entry_price + taker_fee_percent * stop_loss
                        fee_risk = fee_cost_per_unit * position_size
                    
                    risk_amount = price_risk + fee_risk
                    
                    # Calculate R multiple
                    r_multiple = net_pnl / risk_amount if risk_amount > 0 else 0
                else:
                    risk_amount = 0
                    r_multiple = 0
                
                positions_data.append({
                    'symbol': symbol,
                    'direction': direction,
                    'entry_price': entry_price,
                    'current_price': current_price,
                    'position_size': position_size,
                    'stop_loss': stop_loss,
                    'r_multiple': round(r_multiple, 2),
                    'net_pnl': round(net_pnl, 2),
                    'gross_pnl': round(gross_pnl, 2),
                    'entry_time': position.get('entry_time', ''),
                    'order_type': entry_order_type,
                    'tp_order_type': exit_order_type,
                    'entry_fee': round(entry_fee, 2),
                    'exit_fee': round(exit_fee, 2),
                    'total_fees': round(total_fees, 2),
                    'entry_fee_percent': round(entry_fee_percent * 100, 3),
                    'exit_fee_percent': round(exit_fee_percent * 100, 3)
                })
        
        return jsonify({
            'status': 'success',
            'positions': positions_data
        })
        
    except Exception as e:
        logger.error(f"Error getting active positions: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

# API endpoint to close a specific position
@app.route('/api/close-position', methods=['POST'])
def close_position_api():
    try:
        data = request.json
        symbol = data.get('symbol')
        use_limit_order = data.get('useLimitOrder', False)
        
        if not symbol:
            return jsonify({
                'status': 'error',
                'message': 'Symbol is required'
            }), 400
        
        if symbol not in active_positions:
            return jsonify({
                'status': 'error',
                'message': f'No active position found for {symbol}'
            }), 404
        
        # Get position details
        position = active_positions[symbol]
        direction = position['direction']
        
        # Determine exit action
        exit_action = "EXIT LONG" if direction == 'long' else "EXIT SHORT"
        
        # Get current price
        client = initialize_bybit_client()
        ticker_response = client.get_tickers(category="linear", symbol=symbol)
        
        if ticker_response.get('retCode') == 0:
            current_price = float(ticker_response.get('result', {}).get('list', [{}])[0].get('lastPrice', 0))
            
            # Execute the exit trade
            result = execute_trade_with_retry(client, exit_action, symbol, current_price, use_limit_order=use_limit_order, closeOnTrigger=True)
            
            if result:
                # For limit order exits, just return success (actual position close will happen when order fills)
                if use_limit_order:
                    return jsonify({
                        'status': 'success',
                        'message': f'Limit order placed to close {direction} position for {symbol}',
                        'symbol': symbol,
                        'order_type': 'Limit'
                    })
                else:
                    # For market orders, position is closed immediately
                    return jsonify({
                        'status': 'success',
                        'message': f'Successfully closed {direction} position for {symbol}',
                        'symbol': symbol,
                        'exit_price': current_price,
                        'order_type': 'Market'
                    })
            else:
                return jsonify({
                    'status': 'error',
                    'message': f'Failed to close position for {symbol}'
                }), 500
        else:
            return jsonify({
                'status': 'error',
                'message': f'Failed to get current price for {symbol}'
            }), 500
            
    except Exception as e:
        logger.error(f"Error closing position: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

# Function to run the webhook server
def run_webhook_server():
    webhook_port = int(os.getenv('PORT', config['SECURITY'].get('webhook_port', 5000)))
    logger.info(f"Starting webhook server on port {webhook_port}...")
    logger.info(f"Webhook endpoint: /webhook")
    logger.info("Make sure to configure ngrok to forward to this port")
    
    try:
        from waitress import serve
        serve(webhook_app, host='0.0.0.0', port=webhook_port)
    except ImportError:
        logger.warning("Waitress not installed. Using Flask development server instead.")
        webhook_app.run(host='0.0.0.0', port=webhook_port, debug=False)

# Main function
def main():
    try:
        logger.info("Starting trading bot...")
                
        # ADD THIS LINE - Define active_exchange early
        active_exchange = config.get('EXCHANGE', 'active_exchange', fallback='bybit').upper()
        
        # Initialize the database
        init_database()
        
        # Initialize enhanced database for reconciliation
        init_enhanced_database()
        
        # Initialize Excel logger
        init_excel_logger()
        
        # Load positions from database
        global active_positions
        active_positions = load_positions()
        logger.info(f"Loaded {len(active_positions)} positions from database")
        
        # Load pending orders from database
        global pending_orders
        pending_orders = load_pending_orders()
        logger.info(f"Loaded {len(pending_orders)} pending orders from database")
        
        # Initialize the database worker
        start_worker()
        logger.info("Database queue system initialized")
        
        # Initialize Bybit client to verify credentials
        try:
            client = initialize_bybit_client(force_new=True)  # Force new client on startup
            
            # Test API connection
            try:
                account_info = client.get_wallet_balance(accountType="UNIFIED")
                
                if account_info.get('retCode') == 0:
                    balance = account_info.get('result', {}).get('list', [{}])[0].get('totalEquity', 'Unknown')
                    logger.info(f"Successfully connected to Bybit API. Account balance: {balance}")
                    
                    # Log the current capital that will be used for position sizing
                    current_capital = get_current_capital()
                    logger.info(f"Current capital for position sizing: ${current_capital}")
                    
                    # Send notification for bot start
                    if config.getboolean('NOTIFICATIONS', 'notify_system', fallback=False):
                        # Add limit order configuration info to notification
                        limit_order_status = []
                        if config['TRADING'].getboolean('use_limit_orders_entry', False):
                            limit_order_status.append("Entry: Limit")
                        else:
                            limit_order_status.append("Entry: Market")
                            
                        if config['TRADING'].getboolean('use_limit_orders_exit', False):
                            limit_order_status.append("Exit: Limit")
                        else:
                            limit_order_status.append("Exit: Market")
                            
                        limit_order_info = ", ".join(limit_order_status)
                        
                        send_telegram_message(f"🚀 <b>BOT STARTED</b>\n"
                                            f"Exchange: {active_exchange}\n"
                                            f"Connected to {active_exchange} API successfully\n"
                                            f"Account balance: ${balance}\n"
                                            f"Loaded positions: {len(active_positions)}\n"
                                            f"Pending orders: {len(pending_orders)}\n"
                                            f"Order types: {limit_order_info}\n"
                                            f"Limit order timeout: {config['TRADING'].get('limit_order_timeout_minutes', 5)} mins")
                else:
                    logger.error(f"Failed to get account information: {account_info}")
                    
                    # Send notification for error
                    if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                        send_telegram_message(f"❌ <b>STARTUP ERROR</b>\n"
                                            f"Failed to get account information: {account_info.get('retMsg', 'Unknown error')}")
            except Exception as api_error:
                logger.error(f"Error testing API connection: {str(api_error)}")
                logger.warning("Continuing without API verification...")
                
                # Send notification for error
                if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                    send_telegram_message(f"⚠️ <b>STARTUP WARNING</b>\n"
                                        f"Error testing API connection: {str(api_error)}\n"
                                        f"Continuing without API verification")
        except Exception as client_error:
            logger.error(f"Error initializing client: {str(client_error)}")
            logger.warning("Continuing without API client...")
            
            # Send notification for error
            if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
                send_telegram_message(f"❌ <b>STARTUP ERROR</b>\n"
                                    f"Error initializing Bybit client: {str(client_error)}")
        
        # Start position reconciliation thread if enabled
        if config.getboolean('MONITORING', 'enabled', fallback=True) and config.get('MONITORING', 'reconciliation_interval_seconds', fallback='60') != '0':
            start_reconciliation_thread()
        else:
            logger.info("Position reconciliation is disabled in configuration")
        
        # Start pending orders monitor thread
        start_pending_orders_monitor()
        
        # Start cancellation cleanup thread
        start_cancellation_cleanup_thread()
        
        # Start connectivity monitoring
        start_connectivity_monitor()
        
        # Run the Flask app with Waitress for production use
        # Determine port (Railway provides PORT env var)
        server_port = int(os.getenv('PORT', config['SECURITY'].get('webhook_port', 5000)))
        
        # Register webhook routes on the main app so we only need one port
        webhook_endpoint = config['WEBHOOK'].get('endpoint', '/webhook')
        
        @app.route(webhook_endpoint, methods=['POST'])
        def webhook_on_main():
            if request.method == 'POST':
                try:
                    data = request.json
                    logger.info(f"Received webhook: {data}")
                    threading.Thread(target=process_tradingview_alert, args=(data,)).start()
                    return jsonify({"status": "success", "message": "Alert received"})
                except Exception as e:
                    logger.error(f"Error in webhook: {str(e)}")
                    return jsonify({"status": "error", "message": str(e)}), 500
            return jsonify({"status": "error", "message": "Method not allowed"}), 405
        
        logger.info(f"Starting combined server on port {server_port}...")
        logger.info(f"Webhook endpoint: {webhook_endpoint}")
        logger.info(f"Health endpoint: /health")

        try:
            from waitress import serve
            logger.info("Using Waitress production server")
            try:
                serve(app, host='0.0.0.0', port=server_port)
            finally:
                stop_worker()
                logger.info("Database queue system stopped")
        except ImportError:
            logger.warning("Waitress not installed. Using Flask development server.")
            app.run(host='0.0.0.0', port=server_port, debug=False)
                
    except Exception as e:
        logger.error(f"Error in main function: {str(e)}")
        
        # Send notification for error
        if config.getboolean('NOTIFICATIONS', 'notify_errors', fallback=True):
            send_telegram_message(f"❌ <b>CRITICAL ERROR</b>\n"
                                f"Bot startup failed: {str(e)}")

if __name__ == "__main__":
    main()